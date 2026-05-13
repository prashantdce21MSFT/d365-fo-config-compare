"""
D365 FO Config Compare — unified CLI bundled with the VS Code extension.

Two backends:
  - mcp        : JSON-RPC against <baseUrl>/mcp. Auth via 'az account get-access-token'.
  - playwright : Connects to Chrome on localhost:9222 (started by the extension).

Writes one Excel per environment, plus a diff Excel if --diff and >=2 envs.

CLI:
  python extract.py --backend mcp \
      --env UAT=https://...uat.dynamics.com \
      --env Config=https://...config.dynamics.com \
      --le my30 --form smmParameters --out-dir D:\\out [--diff] [--form-path "..."]
"""
from __future__ import annotations
import argparse, json, os, sys, subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# ── Auth (Azure CLI or App Registration / client credentials) ────────────────
def _client_credentials_token(resource: str) -> Optional[str]:
    """If MCP_CLIENT_ID + MCP_TENANT_ID + MCP_CLIENT_SECRET env vars are set,
    use the OAuth2 client_credentials flow against AAD. Returns None when unconfigured."""
    cid = os.environ.get("MCP_CLIENT_ID")
    tid = os.environ.get("MCP_TENANT_ID")
    sec = os.environ.get("MCP_CLIENT_SECRET")
    if not (cid and tid and sec):
        return None
    import requests
    r = requests.post(
        f"https://login.microsoftonline.com/{tid}/oauth2/v2.0/token",
        data={
            "grant_type": "client_credentials",
            "client_id": cid,
            "client_secret": sec,
            "scope": resource.rstrip("/") + "/.default",
        },
        timeout=30,
    )
    if r.status_code != 200:
        raise RuntimeError(f"App-reg token request failed ({r.status_code}): {r.text[:300]}")
    tok = r.json().get("access_token")
    if not tok:
        raise RuntimeError("App-reg token response missing access_token.")
    return tok


def az_token(resource: str, tenant: Optional[str] = None) -> str:
    # Prefer app registration (client credentials) when configured.
    try:
        cc = _client_credentials_token(resource)
        if cc:
            return cc
    except Exception as e:
        print(f"    [auth] client-credentials failed, falling back to az: {e}")

    cmd = ["az", "account", "get-access-token", "--resource", resource, "--query", "accessToken", "-o", "tsv"]
    if tenant:
        cmd += ["--tenant", tenant]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=60, shell=(os.name == "nt"))
    except FileNotFoundError:
        raise RuntimeError("Azure CLI ('az') not found. Install: https://aka.ms/installazurecliwindows")
    if r.returncode != 0:
        raise RuntimeError(f"az get-access-token failed:\n{r.stderr.strip()}\nRun 'az login' (and 'az login --tenant <id>' if multi-tenant).")
    tok = r.stdout.strip()
    if not tok:
        raise RuntimeError("Azure CLI returned empty token.")
    return tok


# ── MCP client ───────────────────────────────────────────────────────────────
class McpClient:
    # Resources to try (in order) when negotiating the audience the MCP endpoint expects.
    _RESOURCE_FALLBACKS = [
        None,  # baseUrl itself
        "https://dynamics.microsoft.com",
        "https://globaldisco.crm.dynamics.com",
    ]

    def __init__(self, base_url: str, env_label: str, tenant: Optional[str] = None, mcp_url: Optional[str] = None):
        import requests
        from urllib.parse import urlparse
        self._req = requests
        self.base_url = base_url.rstrip("/")
        self.mcp_url = (mcp_url or f"{self.base_url}/mcp").rstrip("/")
        # Token resource = origin of the MCP URL (handles cases where MCP is on a different host).
        parsed = urlparse(self.mcp_url)
        self.mcp_resource = f"{parsed.scheme}://{parsed.netloc}"
        self.env_label = env_label
        self.tenant = tenant
        self._resource_idx = 0
        self.token = az_token(self.mcp_resource, tenant)
        self.session_id: Optional[str] = None

    def _refresh_token(self) -> bool:
        """Try the next audience in the fallback list. Returns False when exhausted."""
        self._resource_idx += 1
        if self._resource_idx >= len(self._RESOURCE_FALLBACKS):
            return False
        res = self._RESOURCE_FALLBACKS[self._resource_idx]
        if res is None:
            res = self.base_url
        try:
            print(f"    [auth] retrying with resource={res} tenant={self.tenant or 'default'}")
            self.token = az_token(res, self.tenant)
            self.session_id = None
            return True
        except Exception as e:
            print(f"    [auth] token fetch failed: {e}")
            return self._refresh_token()

    def _rpc(self, method: str, params: Optional[dict] = None, rpc_id: Optional[int] = None):
        while True:
            headers = {
                "Authorization": f"Bearer {self.token}",
                "Content-Type": "application/json",
                "Accept": "application/json",
            }
            if self.session_id:
                headers["mcp-session-id"] = self.session_id
            body: dict = {"jsonrpc": "2.0", "method": method}
            if params is not None: body["params"] = params
            if rpc_id is not None: body["id"] = rpc_id
            r = self._req.post(self.mcp_url, json=body, headers=headers, timeout=120)
            if r.status_code == 401 and method == "initialize":
                www = r.headers.get("WWW-Authenticate", "")
                if www:
                    print(f"    [auth] 401. WWW-Authenticate: {www[:200]}")
                if self._refresh_token():
                    continue
                raise RuntimeError(
                    "401 Unauthorized from /mcp after trying multiple audiences. "
                    "The MCP endpoint may not be enabled on this environment, or your account lacks access. "
                    "Try the Playwright backend instead."
                )
            r.raise_for_status()
            sid = r.headers.get("mcp-session-id")
            if sid: self.session_id = sid
            if rpc_id is not None: return r.json()
            return None

    def connect(self):
        self._rpc("initialize", {"protocolVersion": "2024-11-05", "capabilities": {},
                                 "clientInfo": {"name": "vscode-d365-extractor", "version": "0.1"}}, rpc_id=1)
        self._rpc("notifications/initialized", {})

    def call_tool(self, name: str, args: dict, rpc_id: int = 2) -> str:
        resp = self._rpc("tools/call", {"name": name, "arguments": args}, rpc_id=rpc_id)
        if "error" in resp:
            raise RuntimeError(f"MCP tool {name}: {resp['error'].get('message', resp['error'])}")
        content = resp.get("result", {}).get("content", [])
        return "\n".join(c.get("text", "") for c in content if c.get("type") == "text")

    def tool_json(self, name: str, args: dict) -> dict:
        raw = self.call_tool(name, args)
        try: return json.loads(raw)
        except json.JSONDecodeError: return {"raw": raw}


# ── MCP extraction ───────────────────────────────────────────────────────────
def _parse_input(name: str, info: dict, ftype: str) -> dict:
    return {
        "label": info.get("Label", name),
        "value": info.get("Value", ""),
        "type": ftype.lower(),
        "is_required": info.get("IsRequired", ""),
        "is_editable": info.get("IsEditable", ""),
    }


def _parse_combobox(name: str, info: dict) -> dict:
    value = info.get("ValueText", info.get("Value", ""))
    return {
        "label": info.get("Label", name),
        "value": value,
        "raw_value": info.get("Value", ""),
        "type": "combobox",
    }


def _parse_checkbox(name: str, info: dict) -> dict:
    return {
        "label": info.get("Label", name),
        "value": info.get("IsChecked", info.get("Value", "")),
        "type": "checkbox",
    }


_INPUT_KINDS = ("Input", "RealInput", "IntInput", "DateInput", "TimeInput", "SegmentedEntry")


def _parse_form(form_obj: dict, data: Optional[dict] = None, depth: int = 0) -> dict:
    """Recursively walk FormState.Form.

    Merges fields & grids found at any nesting depth into top-level flat dicts
    (`fields`, `grids`) while ALSO preserving tab structure under `tabs[tk]`.
    """
    if data is None:
        data = {"fields": {}, "grids": {}, "tabs": {}, "fasttabs": {}}
    if not isinstance(form_obj, dict) or depth > 12:
        return data

    # Grids
    for gname, ginfo in (form_obj.get("Grid") or {}).items():
        if not isinstance(ginfo, dict): continue
        cols = [c.get("Label", c.get("Name", "")) for c in ginfo.get("Columns", [])]
        rows = []
        for r in ginfo.get("Rows", []):
            v = {k: v2 for k, v2 in (r.get("Values") or {}).items() if not k.startswith("<")}
            if any(v.values()): rows.append(v)
        pag = ginfo.get("Pagination", {})
        new_grid = {
            "columns": cols, "rows": rows,
            "has_next": str(pag.get("HasNextPage", "False")).lower() == "true",
        }
        existing = data["grids"].get(gname)
        # Preserve a previously-seen grid that already has more rows.
        if existing and len(existing.get("rows", [])) > len(rows):
            if not existing.get("columns"): existing["columns"] = cols
            if new_grid["has_next"]: existing["has_next"] = True
        else:
            data["grids"][gname] = new_grid

    # Fields (Input / RealInput / IntInput / DateInput / TimeInput / SegmentedEntry)
    for kind in _INPUT_KINDS:
        for n, info in (form_obj.get(kind) or {}).items():
            if isinstance(info, dict):
                data["fields"][n] = _parse_input(n, info, kind)
    for n, info in (form_obj.get("Combobox") or {}).items():
        if isinstance(info, dict): data["fields"][n] = _parse_combobox(n, info)
    for n, info in (form_obj.get("Checkbox") or {}).items():
        if isinstance(info, dict): data["fields"][n] = _parse_checkbox(n, info)

    # Tabs & TabPages — capture structure + recurse + merge into flat.
    for container_key in ("Tab", "TabPage"):
        for tn, ti in (form_obj.get(container_key) or {}).items():
            if not isinstance(ti, dict): continue
            tab_text = ti.get("Text", ti.get("Label", tn))
            tab_container = {"fields": {}, "grids": {}, "tabs": {}, "fasttabs": {}}
            _parse_form(ti.get("Children", ti), tab_container, depth + 1)
            if tn not in data["tabs"]:
                data["tabs"][tn] = {"text": tab_text, "fields": {}, "grids": {}}
            data["tabs"][tn]["text"] = tab_text
            data["tabs"][tn]["fields"].update(tab_container["fields"])
            data["tabs"][tn]["grids"].update(tab_container["grids"])
            data["fields"].update(tab_container["fields"])
            data["grids"].update(tab_container["grids"])
            data["fasttabs"].update(tab_container["fasttabs"])

    # FastTabs / Groups — record FastTab state; recurse for nested controls.
    for key in ("Group", "FastTab", "FastTabHeader", "ReferenceGroup"):
        for gn, gi in (form_obj.get(key) or {}).items():
            if not isinstance(gi, dict): continue
            if key in ("FastTab", "FastTabHeader"):
                data["fasttabs"][gn] = {
                    "text": gi.get("Text", gn),
                    "expanded": str(gi.get("IsExpanded", "True")).lower() == "true",
                }
            _parse_form(gi.get("Children", gi), data, depth + 1)

    # Generic Children container.
    children = form_obj.get("Children")
    if isinstance(children, dict):
        _parse_form(children, data, depth + 1)

    return data


def _find_grid_recursive(form_obj: dict, grid_name: str) -> Optional[dict]:
    """Search FormState.Form for a grid by name at any depth."""
    if not isinstance(form_obj, dict):
        return None
    grids = form_obj.get("Grid", {})
    if isinstance(grids, dict) and grid_name in grids and isinstance(grids[grid_name], dict):
        return grids[grid_name]
    for key in ("Tab", "TabPage", "Group", "FastTab", "FastTabHeader", "ReferenceGroup", "Children"):
        container = form_obj.get(key, {})
        if isinstance(container, dict):
            for _n, child in container.items():
                if isinstance(child, dict):
                    for sub in (child.get("Children", {}), child):
                        if sub is not None:
                            r = _find_grid_recursive(sub, grid_name)
                            if r: return r
    return None


def _paginate_grid(client: McpClient, grid_name: str, existing_rows: list) -> list:
    """Walk LoadNextPage until exhausted. Uses recursive grid finder so nested
    Tab/TabPage > Grid responses are handled."""
    seen = {tuple(sorted(r.items())) for r in existing_rows}
    rows = list(existing_rows)
    page = 1
    for _ in range(200):
        try:
            res = client.tool_json("form_click_control",
                                   {"controlName": grid_name, "actionId": "LoadNextPage"})
        except Exception:
            break
        if "raw" in res: break
        form = (res.get("FormState") or {}).get("Form", {})
        gi = _find_grid_recursive(form, grid_name)
        if not gi: break
        added = 0
        for r in gi.get("Rows", []):
            v = {k: v2 for k, v2 in (r.get("Values") or {}).items() if not k.startswith("<")}
            if not any(v.values()): continue
            key = tuple(sorted(v.items()))
            if key in seen: continue
            seen.add(key); rows.append(v); added += 1
        if added == 0: break
        page += 1
        if str((gi.get("Pagination") or {}).get("HasNextPage", "False")).lower() != "true":
            break
    return rows


_FIND_CONTROLS_TERMS = list("abcdefghijklmnopqrstuvwxyz") + [
    "0","1","2","3","4","5","6","7","8","9",
    "Code","Name","Date","Num","Type","Status","Amount","Qty","Unit","Group",
    "Account","Description","Currency","Company","Ledger","Dimension","Table","Line",
    "Purch","Sales","Cust","Vend","Item","Route","Oper","Interest","Fee","Calendar",
    "Paym","Rate","Percent","Invoice","Calculate","Grace","Period","Range","Method",
    "Earnings","Payments","Note","Voucher","Post","Charge","Min","Max","Day","Month",
    "From","To","Base","Value","Debit","Credit","Text","Language","Version","Valid",
    "Warehouse","Site","Config","Color","Size","Style","Batch","Serial","Location",
    "Inventory","Order","Trans","Price","Cost","Discount","Tax","Weight","Address",
    "Contact","Phone","Email","Country","State","Prod","BOM","Resource","Capacity",
    "Schedule","Plan","Budget","Forecast","Actual","Variance","Report","Default",
    "Setup","Parameter","General","Header","Delivery","Shipping","Freight","Carrier",
    "Mode","Terms","Due","Cash","Bank","Check","Return","Reason","Action",
]


def _find_controls_sweep(client: McpClient) -> Dict[str, dict]:
    """Broad sweep using form_find_controls (a-z + domain terms).

    Returns {control_name: {label, value, type}} for any non-grid control
    discovered. Helps capture hidden / lazy-loaded fields the recursive
    FormState walk misses.
    """
    found: Dict[str, dict] = {}
    for term in _FIND_CONTROLS_TERMS:
        try:
            raw = client.call_tool("form_find_controls", {"controlSearchTerm": term})
        except Exception:
            continue
        if not raw or not raw.strip(): continue
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if not isinstance(parsed, list): continue
        for item in parsed:
            cname = item.get("Name", "")
            if not cname or cname in found or cname.startswith("SystemDefined"):
                continue
            props = item.get("Properties") or {}
            if not isinstance(props, dict) or "Columns" in props:
                continue  # skip grids — handled by recursive parse + pagination
            value = props.get("ValueText", props.get("Value", ""))
            found[cname] = {
                "label": props.get("Label", ""),
                "value": str(value).strip() if value else "",
                "type": "swept",
            }
    return found


def _expand_fasttabs(client: McpClient, fasttabs: Dict[str, dict]) -> None:
    for name, info in fasttabs.items():
        if info.get("expanded"): continue
        try: client.tool_json("form_click_control", {"controlName": name})
        except Exception: pass


def extract_mcp(client: McpClient, form: str, le: str) -> dict:
    """Deep MCP extraction: recursive parse + force-open every tab +
    paginate all grids + form_find_controls sweep. Slower than the previous
    shallow walk, but captures lazy-loaded tabs/fields/sub-grids."""
    print(f"  [{client.env_label}/{le}] opening {form} via MCP...")
    client.connect()
    res = client.tool_json("form_open_menu_item",
                           {"name": form, "type": "Display", "companyId": le})
    if "raw" in res and any(s in res["raw"].lower() for s in ("exception", "failed")):
        raise RuntimeError(f"open failed: {res['raw'][:200]}")
    fs = res.get("FormState", {})
    form_obj = fs.get("Form", {})
    caption = fs.get("Caption", "")

    # 1) Initial recursive parse.
    data = _parse_form(form_obj)
    print(f"    Initial parse: {len(data['fields'])} fields, "
          f"{len(data['grids'])} grids, {len(data['tabs'])} tabs, "
          f"{len(data['fasttabs'])} fasttabs")

    # 2) Expand any collapsed FastTabs at top level.
    if data["fasttabs"]:
        print(f"    Expanding {len(data['fasttabs'])} FastTab(s)...")
        _expand_fasttabs(client, data["fasttabs"])

    # 3) Force-open every tab and merge newly-revealed fields/grids.
    tab_keys = list(data["tabs"].keys())
    if tab_keys:
        print(f"    Opening {len(tab_keys)} tab(s)...")
    for idx, tk in enumerate(tab_keys, 1):
        text = data["tabs"][tk].get("text", tk)
        try:
            try:
                res2 = client.tool_json("form_open_or_close_tab",
                                        {"tabName": tk, "tabAction": "Open"})
            except Exception:
                res2 = client.tool_json("form_click_control", {"controlName": tk})
            sub = _parse_form((res2.get("FormState") or {}).get("Form", {}))
            data["tabs"][tk]["fields"].update(sub["fields"])
            data["tabs"][tk]["grids"].update(sub["grids"])
            data["fields"].update(sub["fields"])
            data["grids"].update(sub["grids"])
            inner = sub["tabs"].get(tk, {})
            if inner:
                data["tabs"][tk]["fields"].update(inner.get("fields") or {})
                data["tabs"][tk]["grids"].update(inner.get("grids") or {})
                data["fields"].update(inner.get("fields") or {})
                data["grids"].update(inner.get("grids") or {})
            if sub["fasttabs"]:
                _expand_fasttabs(client, sub["fasttabs"])
                try:
                    res3 = client.tool_json("form_click_control", {"controlName": tk})
                    sub3 = _parse_form((res3.get("FormState") or {}).get("Form", {}))
                    data["tabs"][tk]["fields"].update(sub3["fields"])
                    data["tabs"][tk]["grids"].update(sub3["grids"])
                    data["fields"].update(sub3["fields"])
                    data["grids"].update(sub3["grids"])
                except Exception:
                    pass
            print(f"      [{idx}/{len(tab_keys)}] {text}: "
                  f"{len(data['tabs'][tk]['fields'])} fields, "
                  f"{len(data['tabs'][tk]['grids'])} grid(s)")
        except Exception as e:
            print(f"      [{idx}/{len(tab_keys)}] {text}: open failed ({e})")

    # 4) Paginate every grid (single + nested).
    grids_to_page = [(gn, gi) for gn, gi in data["grids"].items()
                     if gi.get("has_next") or len(gi.get("rows", [])) >= 25]
    if grids_to_page:
        print(f"    Paginating {len(grids_to_page)} grid(s)...")
    for gname, gi in grids_to_page:
        before = len(gi.get("rows", []))
        gi["rows"] = _paginate_grid(client, gname, gi.get("rows", []))
        added = len(gi["rows"]) - before
        if added > 0:
            print(f"      {gname}: +{added} rows (total {len(gi['rows'])})")

    # 5) form_find_controls sweep.
    print(f"    Running form_find_controls sweep ({len(_FIND_CONTROLS_TERMS)} terms)...")
    swept = _find_controls_sweep(client)
    new = 0
    for cname, finfo in swept.items():
        if cname not in data["fields"]:
            data["fields"][cname] = finfo
            new += 1
    print(f"      Sweep added {new} new field(s) "
          f"(total fields: {len(data['fields'])}).")

    try: client.call_tool("form_close_form", {})
    except Exception: pass

    # Build the shape write_excel expects. Avoid duplication: fields/grids
    # attributed to a tab go ONLY into the tab; everything else is "top".
    tab_field_names = set()
    tab_grid_names = set()
    for ti in data["tabs"].values():
        tab_field_names.update((ti.get("fields") or {}).keys())
        tab_grid_names.update((ti.get("grids") or {}).keys())
    top_fields = {k: v for k, v in data["fields"].items() if k not in tab_field_names}
    top_grids = {k: v for k, v in data["grids"].items() if k not in tab_grid_names}

    return {
        "form": form, "caption": caption, "company": le,
        "top_fields": top_fields, "top_grids": top_grids, "tabs": data["tabs"],
    }


# ── Playwright extraction (full tab walk + virtualized grid scroll) ──────────
_ACTION_BUTTONS = {
    "add", "remove", "delete", "new", "edit", "save", "cancel", "close",
    "post", "confirm", "approve", "reject", "submit", "validate",
    "refresh", "print", "export", "import", "copy", "move", "attach",
    "back", "share", "options", "personalize", "personalize this page",
    "power apps", "read mode", "revert",
}

_JS_TABS = """(actionButtons) => {
    const candidates = [
        ...document.querySelectorAll('[class*="verticalTabs"] button'),
        ...document.querySelectorAll('[class*="sysTabControl"] [role="tab"]'),
        ...document.querySelectorAll('[role="tablist"] [role="tab"]'),
        ...document.querySelectorAll('[class*="tabPage"] button'),
        ...document.querySelectorAll('[class*="pivot"] button'),
    ];
    const seen = new Set();
    return candidates.map((el, idx) => {
        const name = (el.textContent || '').trim().replace(/\\s+/g, ' ');
        if (name && name.length < 120 && !seen.has(name)
                && !actionButtons.includes(name.toLowerCase())) {
            seen.add(name);
            return { name, idx };
        }
        return null;
    }).filter(Boolean);
}"""

_JS_CLICK_TAB = """(args) => {
    const actionButtons = args.actionButtons;
    const tabIdx = args.tabIdx;
    const candidates = [
        ...document.querySelectorAll('[class*="verticalTabs"] button'),
        ...document.querySelectorAll('[class*="sysTabControl"] [role="tab"]'),
        ...document.querySelectorAll('[role="tablist"] [role="tab"]'),
        ...document.querySelectorAll('[class*="tabPage"] button'),
        ...document.querySelectorAll('[class*="pivot"] button'),
    ];
    const seen = new Set();
    const tabs = [];
    candidates.forEach(el => {
        const name = (el.textContent || '').trim().replace(/\\s+/g, ' ');
        if (name && name.length < 120 && !seen.has(name)
                && !actionButtons.includes(name.toLowerCase())) {
            seen.add(name);
            tabs.push(el);
        }
    });
    if (tabs[tabIdx]) { tabs[tabIdx].click(); return true; }
    return false;
}"""

_JS_EXPAND_FASTTABS = """() => {
    let n = 0;
    // Only target FastTab/group headers, not arbitrary collapsed buttons
    // (avoids opening menus like Share > Create alert rule).
    const sels = [
        '[class*="fastTab"] button[aria-expanded="false"]',
        '[class*="FastTab"] button[aria-expanded="false"]',
        'button[class*="fastTabHeader"][aria-expanded="false"]',
        'button[class*="FastTabHeader"][aria-expanded="false"]',
        '[class*="groupHeader"] button[aria-expanded="false"]',
    ];
    const seen = new Set();
    sels.forEach(s => document.querySelectorAll(s).forEach(h => {
        if (seen.has(h)) return; seen.add(h);
        const t = (h.textContent || '').trim().toLowerCase();
        if (t.includes('alert') || t.includes('share') || t.includes('option')) return;
        try { h.click(); n++; } catch(e) {}
    }));
    return n;
}"""

_JS_SNAPSHOT_FIELDS = """() => {
    const out = [];
    const seen = new Set();
    document.querySelectorAll('input,[role="combobox"],[role="checkbox"]').forEach(el => {
        const id = el.getAttribute('id') || '';
        const labelByFor = id ? document.querySelector(`label[for="${CSS.escape(id)}"]`) : null;
        const lbl = (labelByFor?.textContent?.trim()
                  || el.getAttribute('aria-label')
                  || el.getAttribute('name') || '').trim();
        if (!lbl || lbl.toLowerCase() === 'filter' || lbl.toLowerCase() === 'search') return;
        const val = (el.value || el.getAttribute('aria-checked') || '').toString().trim();
        const key = lbl + '|' + val;
        if (seen.has(key)) return; seen.add(key);
        out.push({ label: lbl, value: val });
    });
    return out;
}"""

_JS_GRID_CENTERS = """() =>
    [...document.querySelectorAll('[role="grid"], [class*="gridView"]')]
        .filter(g => g.getBoundingClientRect().height > 0)
        .map(g => {
            const r = g.getBoundingClientRect();
            return {x: Math.round(r.left + r.width/2), y: Math.round(r.top + r.height/2)};
        })"""

_JS_SNAPSHOT_GRIDS = """() => {
    const grids = [...document.querySelectorAll('[role="grid"]')].filter(g => g.getBoundingClientRect().height > 0);
    return grids.map(g => {
        const headers = [...g.querySelectorAll('[role="columnheader"]')]
            .map(h => (h.textContent || '').trim()).filter(Boolean);
        const rows = [];
        g.querySelectorAll('[role="row"]').forEach(r => {
            if (r.querySelector('[role="columnheader"]')) return;
            const cells = [...r.querySelectorAll('[role="gridcell"],[role="cell"]')]
                .map(c => {
                    const inp = c.querySelector('input');
                    return inp ? (inp.value || '').trim() : (c.textContent || '').trim();
                });
            if (cells.some(v => v)) rows.push(cells);
        });
        return { headers, rows };
    });
}"""


def _scroll_collect_grids(page) -> List[dict]:
    grid_centers = page.evaluate(_JS_GRID_CENTERS) or []
    if not grid_centers:
        return page.evaluate(_JS_SNAPSHOT_GRIDS) or []

    n = len(grid_centers)
    all_hdrs: List[List[str]] = [[] for _ in range(n)]
    all_rows: List[Dict[str, list]] = [dict() for _ in range(n)]

    for gi, g in enumerate(page.evaluate(_JS_SNAPSHOT_GRIDS) or []):
        if gi >= n: break
        all_hdrs[gi] = g.get("headers", [])
        for row in g.get("rows", []):
            all_rows[gi]["|".join(str(c) for c in row)] = row

    for gi, pt in enumerate(grid_centers):
        page.mouse.move(pt["x"], pt["y"])
        prev = -1
        for _ in range(120):
            page.mouse.wheel(0, 250)
            page.wait_for_timeout(120)
            cur = page.evaluate(_JS_SNAPSHOT_GRIDS) or []
            if gi < len(cur):
                g = cur[gi]
                if not all_hdrs[gi] and g.get("headers"): all_hdrs[gi] = g["headers"]
                for row in g.get("rows", []):
                    all_rows[gi]["|".join(str(c) for c in row)] = row
            count = len(all_rows[gi])
            if count == prev: break
            prev = count

    return [{"headers": all_hdrs[i], "rows": list(all_rows[i].values())} for i in range(n)]


def extract_playwright(base_url: str, env_label: str, form: str, le: str) -> dict:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    host = base_url.replace("https://", "").replace("http://", "").rstrip("/")
    url = f"https://{host}/?cmp={le.lower()}&mi={form}"
    print(f"  [{env_label}/{le}] {url}")

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.connect_over_cdp("http://localhost:9222")
        except Exception:
            raise RuntimeError("Could not connect to Chrome at localhost:9222. Run 'D365 FO Config Compare: Start Chrome with Remote Debugging' first.")
        ctx = browser.contexts[0] if browser.contexts else browser.new_context()
        page = ctx.new_page()
        page.set_viewport_size({"width": 1600, "height": 900})
        page.goto(url)
        try:
            page.wait_for_selector(
                '[class*="formBody"],[class*="verticalTabs"],[class*="pivotItem"],'
                '[role="grid"],[class*="listSection"],[role="row"]',
                timeout=30_000,
            )
        except PWTimeout:
            pass
        if "Sign in" in page.title():
            raise RuntimeError("Not authenticated in Chrome. Sign in to D365, then retry.")
        page.wait_for_timeout(2500)

        try: page.evaluate(_JS_EXPAND_FASTTABS); page.wait_for_timeout(700)
        except Exception: pass

        top_fields_arr = page.evaluate(_JS_SNAPSHOT_FIELDS) or []
        top_grids_arr  = _scroll_collect_grids(page)
        top_fields = {f"f{i}": {"label": f["label"], "value": f["value"], "type": "input"} for i, f in enumerate(top_fields_arr)}
        top_grids: Dict[str, dict] = {}
        for i, g in enumerate(top_grids_arr):
            top_grids[f"Grid{i}"] = {
                "columns": g["headers"],
                "rows": [dict(zip(g["headers"], r)) if g["headers"] else r for r in g["rows"]],
            }

        tabs = page.evaluate(_JS_TABS, list(_ACTION_BUTTONS)) or []
        print(f"    Tabs discovered: {len(tabs)}")
        out_tabs: Dict[str, dict] = {}
        for i, t in enumerate(tabs, 1):
            try:
                ok = page.evaluate(_JS_CLICK_TAB, {"actionButtons": list(_ACTION_BUTTONS), "tabIdx": t["idx"]})
                if not ok: continue
                page.wait_for_timeout(2500)
                try: page.evaluate(_JS_EXPAND_FASTTABS); page.wait_for_timeout(600)
                except Exception: pass

                fields_arr = page.evaluate(_JS_SNAPSHOT_FIELDS) or []
                grids_arr  = _scroll_collect_grids(page)
                fields = {f"f{j}": {"label": f["label"], "value": f["value"], "type": "input"} for j, f in enumerate(fields_arr)}
                grids: Dict[str, dict] = {}
                for j, g in enumerate(grids_arr):
                    grids[f"Grid{j}"] = {
                        "columns": g["headers"],
                        "rows": [dict(zip(g["headers"], r)) if g["headers"] else r for r in g["rows"]],
                    }
                out_tabs[f"tab{i}"] = {"text": t["name"], "fields": fields, "grids": grids}
                print(f"    [{i}/{len(tabs)}] {t['name']}: {len(fields)} fields, {len(grids)} sub-grid(s)")
            except Exception as e:
                print(f"    tab '{t.get('name')}' failed: {e}")

        browser.close()

    return {"form": form, "caption": "", "company": le,
            "top_fields": top_fields, "top_grids": top_grids, "tabs": out_tabs}


# ── Excel output ─────────────────────────────────────────────────────────────
def write_excel(data: dict, env_label: str, out_dir: Path, form: str, le: str) -> Path:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    HDR = PatternFill("solid", fgColor="1F4E79")
    SUB = PatternFill("solid", fgColor="BDD7EE")
    WHITE = Font(color="FFFFFF", bold=True)
    BOLD = Font(bold=True)
    THIN = Side(style="thin", color="CCCCCC")
    BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def cell(ws, r, c, v="", fill=None, font=None, bold=False, wrap=False):
        x = ws.cell(row=r, column=c, value=v if v is not None else "")
        if fill: x.fill = fill
        if font: x.font = font
        elif bold: x.font = BOLD
        x.border = BRD
        x.alignment = Alignment(wrap_text=wrap, vertical="center")
        return x

    ws = wb.create_sheet("Summary")
    cell(ws, 1, 1, f"Form: {form}", bold=True)
    cell(ws, 2, 1, f"Legal Entity: {le}")
    cell(ws, 3, 1, f"Environment: {env_label}")
    cell(ws, 4, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    cell(ws, 5, 1, f"Caption: {data.get('caption', '')}")
    ws.column_dimensions["A"].width = 80

    ws = wb.create_sheet("Fields")
    for c, h in enumerate(["Tab", "Field", "Value", "Type"], 1):
        cell(ws, 1, c, h, fill=HDR, font=WHITE)
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 36; ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"
    row = 2
    for fk, fv in (data.get("top_fields") or {}).items():
        cell(ws, row, 1, "(top)")
        cell(ws, row, 2, fv.get("label", fk), wrap=True)
        cell(ws, row, 3, str(fv.get("value", "")), wrap=True)
        cell(ws, row, 4, fv.get("type", ""))
        row += 1
    for tk, ti in (data.get("tabs") or {}).items():
        for fk, fv in (ti.get("fields") or {}).items():
            cell(ws, row, 1, ti.get("text", tk))
            cell(ws, row, 2, fv.get("label", fk), wrap=True)
            cell(ws, row, 3, str(fv.get("value", "")), wrap=True)
            cell(ws, row, 4, fv.get("type", ""))
            row += 1

    ws = wb.create_sheet("Grids")
    row = 1
    def write_grid(label, g):
        nonlocal row
        cols = g.get("columns", [])
        cell(ws, row, 1, label, fill=HDR, font=WHITE); row += 1
        if cols:
            for c, h in enumerate(cols, 1): cell(ws, row, c, h, fill=SUB, bold=True)
            row += 1
        for r in g.get("rows", []):
            if isinstance(r, dict):
                for c, h in enumerate(cols, 1): cell(ws, row, c, str(r.get(h, "")))
            else:
                for c, v in enumerate(r, 1): cell(ws, row, c, str(v))
            row += 1
        row += 1

    for gname, g in (data.get("top_grids") or {}).items():
        write_grid(f"Grid: {gname}", g)
    for tk, ti in (data.get("tabs") or {}).items():
        for gname, g in (ti.get("grids") or {}).items():
            write_grid(f"Tab: {ti.get('text', tk)} / Grid: {gname}", g)

    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    import re as _re
    safe_label = _re.sub(r'[<>:"/\\|?*]+', '_', env_label).strip('._ ') or "env"
    path = out_dir / f"{form}_{le}_{safe_label}_{ts}.xlsx"
    wb.save(path)
    return path


def write_diff_excel(per_env: List[Tuple[str, dict]], out_dir: Path, form: str, le: str) -> Optional[Path]:
    if len(per_env) < 2: return None
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
    HDR = PatternFill("solid", fgColor="1F4E79")
    DIFF = PatternFill("solid", fgColor="FFE0E0")
    MATCH = PatternFill("solid", fgColor="E2EFDA")
    WHITE = Font(color="FFFFFF", bold=True)
    THIN = Side(style="thin", color="CCCCCC")
    BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def all_field_labels(d):
        out: List[Tuple[str, str, str]] = []
        for fk, fv in (d.get("top_fields") or {}).items():
            out.append(("(top)", fv.get("label", fk), str(fv.get("value", ""))))
        for tk, ti in (d.get("tabs") or {}).items():
            for fk, fv in (ti.get("fields") or {}).items():
                out.append((ti.get("text", tk), fv.get("label", fk), str(fv.get("value", ""))))
        return out

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Diff")
    headers = ["Tab", "Field"] + [e[0] for e in per_env] + ["Match"]
    for c, h in enumerate(headers, 1):
        x = ws.cell(row=1, column=c, value=h); x.fill = HDR; x.font = WHITE; x.border = BRD
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 44
    for i in range(3, 3 + len(per_env) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 28
    ws.freeze_panes = "A2"

    rows_by_key: Dict[Tuple[str, str], Dict[str, str]] = {}
    for env_label, data in per_env:
        for tab, label, value in all_field_labels(data):
            rows_by_key.setdefault((tab, label), {})[env_label] = value

    row = 2
    n_diffs = 0
    for (tab, label), vals in rows_by_key.items():
        env_vals = [vals.get(env_label, "") for env_label, _ in per_env]
        is_diff = len(set(env_vals)) > 1
        if is_diff: n_diffs += 1
        fill = DIFF if is_diff else MATCH
        ws.cell(row=row, column=1, value=tab).fill = fill
        ws.cell(row=row, column=2, value=label).fill = fill
        for i, v in enumerate(env_vals, start=3):
            ws.cell(row=row, column=i, value=v).fill = fill
        ws.cell(row=row, column=3 + len(per_env), value="DIFF" if is_diff else "Match").fill = fill
        row += 1

    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{form}_{le}_DIFF_{ts}.xlsx"
    wb.save(path)
    print(f"\nDiff: {n_diffs} field(s) differ across environments.")
    return path


# ── Form path validation (Playwright via nav search) ────────────────────────
def validate_form_paths(base_url: str, le: str, paths: List[str]) -> None:
    """For each path like 'A > B > C > Form Name', open D365 nav search, type the
    last segment, pick the result whose breadcrumb matches the parents, then
    capture the mi= URL parameter. Prints RESULT|path|menuItem|error per line."""
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    from urllib.parse import urlparse, parse_qs
    host = base_url.replace("https://", "").replace("http://", "").rstrip("/")
    home = f"https://{host}/?cmp={le.lower()}"
    print(f"  Opening {home}")

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.connect_over_cdp("http://localhost:9222")
        except Exception:
            for p in paths:
                print(f"RESULT|{p}||Chrome CDP not running on :9222")
            return
        ctx = browser.contexts[0] if browser.contexts else browser.new_context()
        page = ctx.new_page()
        page.set_viewport_size({"width": 1600, "height": 900})
        page.goto(home)
        try:
            page.wait_for_selector('[class*="dashboard"], [class*="navigationBar"], [aria-label*="search" i]', timeout=30_000)
        except PWTimeout:
            pass
        page.wait_for_timeout(1500)

        for raw in paths:
            segs = [s.strip() for s in raw.split(">") if s.strip()]
            if not segs:
                print(f"RESULT|{raw}||empty path"); continue
            terminal = segs[-1]
            parents = segs[:-1]
            try:
                # Always start each path from the dashboard so a previous failure
                # (error dialog, stale form) cannot poison the next attempt.
                page.goto(home)
                try:
                    page.wait_for_selector('[class*="dashboard"], [class*="navigationBar"], [aria-label*="search" i]', timeout=15_000)
                except PWTimeout:
                    pass
                page.wait_for_timeout(800)

                # Dismiss any flyouts / notifications.
                try: page.mouse.click(10, 300)
                except Exception: pass
                page.keyboard.press("Escape")
                page.wait_for_timeout(120)
                page.keyboard.press("Escape")
                page.wait_for_timeout(200)

                # Ensure focus on the page.
                try: page.bring_to_front()
                except Exception: pass
                try: page.locator("body").first.click(position={"x": 10, "y": 300}, force=True, timeout=2000)
                except Exception: pass

                input_selectors = [
                    'input[placeholder*="Search for a page" i]',
                    'input[aria-label*="navigation search" i]',
                    'input[aria-label*="search for a page" i]',
                    'input[aria-label="Search" i]',
                    '[class*="navigationSearch"] input',
                    '[class*="NavigationSearch"] input',
                    'input[name*="NavigationSearch" i]',
                ]
                button_selectors = [
                    'button[aria-label="Search" i]',
                    '[aria-label="Search" i]',
                    '[aria-label*="Search for a page" i]',
                    '[title*="Search for a page" i]',
                    'button[name*="NavigationSearchButton" i]',
                    '[id*="NavigationSearchButton" i]',
                    '[data-dyn-controlname*="NavigationSearch" i]',
                    '[class*="appBarSearchButton" i]',
                ]

                def find_search_input():
                    for sel in input_selectors:
                        loc = page.locator(sel).first
                        try:
                            if loc.count():
                                # Don't require is_visible() — some inputs are hidden
                                # until focused but still interactable.
                                return loc
                        except Exception:
                            pass
                    return None

                # Attempt 1: Alt+G keyboard shortcut.
                try:
                    page.keyboard.down("Alt")
                    page.keyboard.press("KeyG")
                    page.keyboard.up("Alt")
                except Exception:
                    pass
                page.wait_for_timeout(600)
                search = find_search_input()

                # Attempt 2: click the magnifier / nav-search button.
                if not search:
                    for sel in button_selectors:
                        btn = page.locator(sel).first
                        try:
                            if btn.count():
                                try: btn.click(timeout=2000)
                                except Exception:
                                    try: btn.click(force=True, timeout=2000)
                                    except Exception: continue
                                page.wait_for_timeout(500)
                                search = find_search_input()
                                if search: break
                        except Exception:
                            pass

                if not search:
                    # Dump screenshot + visible elements for diagnostics.
                    try:
                        shot = Path(os.environ.get("TEMP", ".")) / f"d365_search_fail_{datetime.now():%H%M%S}.png"
                        page.screenshot(path=str(shot), full_page=False)
                        sample = page.evaluate(
                            """() => [...document.querySelectorAll('input,button')]
                                .filter(e => e.offsetParent !== null)
                                .slice(0, 25)
                                .map(e => `${e.tagName} aria='${e.getAttribute('aria-label')||''}' ph='${e.getAttribute('placeholder')||''}' title='${e.getAttribute('title')||''}'`)
                                .join(' || ')"""
                        )
                        print(f"  [diag] screenshot: {shot}")
                        print(f"  [diag] inputs/buttons: {sample[:600]}")
                    except Exception:
                        pass
                    print(f"RESULT|{raw}||search box not found"); continue
                search.click()
                search.fill("")
                search.type(terminal, delay=30)
                # Wait for the nav-search results to appear. F&O renders results in a
                # portal/flyout that is NOT inside the navigationSearch container, so
                # we use a broad selector and filter by breadcrumb shape below.
                result_sel = (
                    '[role="option"], '
                    '[role="listbox"] [role="listitem"], '
                    '[class*="searchResultItem"], '
                    '[class*="searchResult"] li, '
                    '[class*="SearchResult"] li'
                )
                try:
                    page.wait_for_function(
                        """sel => document.querySelectorAll(sel).length > 0""",
                        arg=result_sel, timeout=5000,
                    )
                except PWTimeout:
                    pass
                page.wait_for_timeout(500)
                # Collect candidate results, then filter to those that look like
                # navigation search items: contain a breadcrumb ' > ' OR contain the
                # terminal token. This excludes notification flyout listitems.
                raw_results = page.evaluate(
                    """(sel) => {
                        const items = [...document.querySelectorAll(sel)];
                        return items.slice(0, 80).map((el, i) => ({
                            idx: i,
                            text: (el.innerText || '').trim().replace(/\\s+/g, ' '),
                        })).filter(x => x.text);
                    }""", result_sel
                )
                term_lc = terminal.lower()
                results = [
                    r for r in raw_results
                    if (" > " in r["text"]) or (term_lc in r["text"].lower())
                ]
                if not results:
                    top5 = " | ".join(r["text"][:60] for r in raw_results[:5]) or "<none>"
                    print(f"RESULT|{raw}||no search results for '{terminal}' (panel empty). Sample: {top5!r}"); continue
                # Score by parent containment
                def score(text: str) -> int:
                    t = text.lower()
                    return sum(1 for p in parents if p.lower() in t) + (5 if terminal.lower() in t else 0)
                results.sort(key=lambda r: score(r["text"]), reverse=True)
                best = results[0]
                if score(best["text"]) == 0:
                    top5 = " | ".join(r["text"][:60] for r in results[:5])
                    print(f"RESULT|{raw}||no matching result. Top5: {top5!r}"); continue
                # Remember the URL/mi before clicking so we can detect real navigation.
                initial_url = page.url
                initial_mi = (parse_qs(urlparse(initial_url).query).get("mi") or [""])[0]

                # Click the matching item
                clicked = page.evaluate(
                    """(args) => {
                        const items = [...document.querySelectorAll(args.sel)];
                        if (items[args.idx]) { items[args.idx].click(); return true; }
                        return false;
                    }""", {"sel": result_sel, "idx": best["idx"]}
                )
                if not clicked:
                    print(f"RESULT|{raw}||click failed"); continue

                # Wait for the URL's mi= to change to something different from the
                # dashboard (or whatever we started on). If it doesn't change, fall
                # back to keyboard navigation: ArrowDown * idx then Enter.
                def current_mi() -> str:
                    return (parse_qs(urlparse(page.url).query).get("mi") or [""])[0]

                changed = False
                for _ in range(20):
                    page.wait_for_timeout(300)
                    mi_now = current_mi()
                    if mi_now and mi_now != initial_mi:
                        changed = True; break

                if not changed:
                    # Fallback: keyboard navigation inside the search box.
                    try:
                        search.click()
                        for _ in range(best["idx"] + 1):
                            page.keyboard.press("ArrowDown")
                            page.wait_for_timeout(80)
                        page.keyboard.press("Enter")
                    except Exception:
                        pass
                    for _ in range(20):
                        page.wait_for_timeout(300)
                        mi_now = current_mi()
                        if mi_now and mi_now != initial_mi:
                            changed = True; break

                mi = current_mi()
                if not changed or not mi or mi == initial_mi:
                    print(f"RESULT|{raw}||navigation did not occur (still mi={mi or '<none>'})"); continue
                if mi.lower() in ("defaultdashboard", "navigationsearchpage"):
                    print(f"RESULT|{raw}||unexpected mi after click: {mi}"); continue
                print(f"  OK {raw}  ->  {mi}")
                print(f"RESULT|{raw}|{mi}|")
            except Exception as e:
                print(f"RESULT|{raw}||{type(e).__name__}: {e}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description="D365 FO Config Compare")
    p.add_argument("--validate-paths", action="store_true", help="Validate form paths and print RESULT lines")
    p.add_argument("--paths", default="", help="Unit-separator (\\u001f) joined paths for --validate-paths")
    p.add_argument("--backend", choices=["mcp", "playwright"])
    p.add_argument("--env", action="append", required=True, help="LABEL=BASE_URL[|TENANT[|MCP_URL]] (repeatable)")
    p.add_argument("--le", required=True)
    p.add_argument("--form")
    p.add_argument("--form-path", default="")
    p.add_argument("--out-dir")
    p.add_argument("--diff", action="store_true")
    args = p.parse_args()

    envs: List[Tuple[str, str, Optional[str], Optional[str]]] = []
    for e in args.env:
        if "=" not in e:
            print(f"bad --env: {e}", file=sys.stderr); sys.exit(2)
        lbl, rest = e.split("=", 1)
        parts = rest.split("|")
        url = parts[0].strip().rstrip("/")
        tenant = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None
        mcp_url = parts[2].strip() if len(parts) > 2 and parts[2].strip() else None
        envs.append((lbl.strip(), url, tenant, mcp_url))

    if args.validate_paths:
        if not args.paths.strip():
            print("No paths provided.", file=sys.stderr); sys.exit(2)
        paths = [p for p in args.paths.split("\u001f") if p.strip()]
        _, base_url, _, _ = envs[0]
        validate_form_paths(base_url, args.le, paths)
        return

    if not args.backend or not args.form or not args.out_dir:
        print("--backend, --form, --out-dir are required for extraction", file=sys.stderr); sys.exit(2)

    out_dir = Path(args.out_dir)
    print(f"D365 FO Config Compare — backend={args.backend} form={args.form} le={args.le}")
    if args.form_path: print(f"Form path: {args.form_path}")
    print(f"Environments: {', '.join(l for l,_,_,_ in envs)}")

    per_env: List[Tuple[str, dict]] = []
    for label, base_url, tenant, mcp_url in envs:
        print(f"\n=== {label} ===" + (f" (tenant={tenant})" if tenant else "") + (f" mcp={mcp_url}" if mcp_url else ""))
        if args.backend == "mcp":
            client = McpClient(base_url, label, tenant, mcp_url)
            data = extract_mcp(client, args.form, args.le)
        else:
            data = extract_playwright(base_url, label, args.form, args.le)
        path = write_excel(data, label, out_dir, args.form, args.le)
        print(f"Saved: {path}")
        per_env.append((label, data))

    if args.diff and len(per_env) >= 2:
        d = write_diff_excel(per_env, out_dir, args.form, args.le)
        if d: print(f"Diff: {d}")


if __name__ == "__main__":
    main()
