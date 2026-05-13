"""
Compare D365 form controls between UAT and Config environments.
Extracts all controls from both envs for all LEs and produces
a single comparison Excel sheet.
"""
import sys
import os
import json
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from form_control_extractor import FormControlExtractor

OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output\Comparison"


def merge_le_controls(le_data):
    """Merge all controls from a single LE extraction into one flat dict.
    Returns {control_name: {label, value, type, source}}
    """
    merged = {}

    # Grids — store grid rows as a single value string
    for gname, gdata in le_data.get("grids", {}).items():
        rows = gdata.get("rows", [])
        cols = [c["label"] for c in gdata.get("columns", [])]
        # Grid summary
        merged[f"[GRID] {gname}"] = {
            "label": gdata.get("text", gname),
            "value": f"{gdata['total_rows']} rows",
            "type": "grid",
            "source": "FormState",
        }
        # Each grid row as its own entry
        for ri, row in enumerate(rows):
            for col_label in cols:
                key = f"{gname}[{ri}].{col_label}"
                merged[key] = {
                    "label": col_label,
                    "value": str(row.get(col_label, "")),
                    "type": "grid_cell",
                    "source": "FormState",
                }

        # Row detail fields
        for ri, fields in sorted(gdata.get("row_details", {}).items()):
            for fname, fdata in fields.items():
                key = f"{gname}[{ri}].detail.{fname}"
                merged[key] = {
                    "label": fdata.get("label", ""),
                    "value": str(fdata.get("value", "")),
                    "type": fdata.get("type", "field"),
                    "source": "RowDetail",
                }

    # Fields from FormState
    for fname, fdata in le_data.get("fields", {}).items():
        if fname not in merged:
            merged[fname] = {
                "label": fdata.get("label", ""),
                "value": str(fdata.get("value", "")),
                "type": fdata.get("type", "field"),
                "source": "FormState",
            }

    # Extra from find_controls (not already captured)
    for cname, cdata in le_data.get("find_controls", {}).items():
        if cname not in merged and cname not in le_data.get("grids", {}):
            # Skip system/button controls for cleaner comparison
            if cname.startswith("SystemDefined"):
                continue
            merged[cname] = {
                "label": cdata.get("label", ""),
                "value": str(cdata.get("value", "")),
                "type": cdata.get("type", "field"),
                "source": "FindControls",
            }

    return merged


def build_comparison(uat_result, config_result, legal_entities):
    """Build comparison data structure.
    Returns list of rows: [{control, label, LE_UAT_value, LE_Config_value, LE_match, ...}]
    """
    # Collect all unique control keys across all LEs and both envs
    all_keys = set()
    uat_merged = {}
    config_merged = {}

    for le in legal_entities:
        uat_le = uat_result["legal_entities"].get(le, {})
        cfg_le = config_result["legal_entities"].get(le, {})

        uat_merged[le] = merge_le_controls(uat_le) if uat_le.get("status") == "OK" else {}
        config_merged[le] = merge_le_controls(cfg_le) if cfg_le.get("status") == "OK" else {}

        all_keys.update(uat_merged[le].keys())
        all_keys.update(config_merged[le].keys())

    # Sort keys: grids first, then grid cells, then fields
    def sort_key(k):
        if k.startswith("[GRID]"):
            return (0, k)
        if "].detail." in k:
            return (2, k)
        if "]." in k:
            return (1, k)
        return (3, k)

    sorted_keys = sorted(all_keys, key=sort_key)

    rows = []
    for key in sorted_keys:
        # Get label from whichever env has it
        label = ""
        ctrl_type = ""
        for le in legal_entities:
            for merged in [uat_merged[le], config_merged[le]]:
                if key in merged:
                    label = label or merged[key].get("label", "")
                    ctrl_type = ctrl_type or merged[key].get("type", "")

        row = {
            "control": key,
            "label": label,
            "type": ctrl_type,
        }

        any_mismatch = False
        for le in legal_entities:
            uat_val = uat_merged[le].get(key, {}).get("value", "")
            cfg_val = config_merged[le].get(key, {}).get("value", "")

            row[f"{le}_UAT"] = uat_val
            row[f"{le}_Config"] = cfg_val

            # Determine match status
            if not uat_val and not cfg_val:
                row[f"{le}_Match"] = ""  # both empty
            elif uat_val == cfg_val:
                row[f"{le}_Match"] = "Match"
            else:
                row[f"{le}_Match"] = "MISMATCH"
                any_mismatch = True

        row["has_mismatch"] = any_mismatch
        rows.append(row)

    return rows


def to_excel(rows, legal_entities, uat_result, config_result):
    """Write comparison to Excel — D365-style grid tables + field comparison."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    form_name = uat_result.get("form_name") or uat_result.get("menu_item", "Form")
    filepath = os.path.join(OUTPUT_DIR, f"Compare_{form_name}_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # ── Styles ──
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold = Font(bold=True, size=10)
    mismatch_font = Font(bold=True, color="9C0006", size=10)
    uat_hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cfg_hfill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    le_sep_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    le_sep_font = Font(bold=True, size=11, color="1F3864")

    def write_header(ws, headers, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font, c.fill, c.border = hf, hfill, tb
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws, ncols, nrows):
        for ci in range(1, ncols + 1):
            mx = 8
            for r in range(1, min(nrows + 1, 200)):
                v = ws.cell(row=r, column=ci).value
                if v:
                    mx = max(mx, min(len(str(v)), 55))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 58)

    # ── Summary sheet ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_rows = [
        ["Form Name", uat_result.get("form_name", "")],
        ["Form Caption", uat_result.get("form_caption", "")],
        ["Menu Item", uat_result.get("menu_item", "")],
        ["MI Type", uat_result.get("mi_type", "")],
        ["Navigation Path", uat_result.get("navigation_path", "")],
        ["UAT Environment", "Env1 UAT (ENV1)"],
        ["Config Environment", "Env4 Config (ENV4)"],
        ["Extracted At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["Legal Entity", "UAT Status", "Config Status",
         "UAT Grid Rows", "Config Grid Rows",
         "Total Mismatches"],
    ]
    total_mismatches = 0
    for le in legal_entities:
        uat_le = uat_result["legal_entities"].get(le, {})
        cfg_le = config_result["legal_entities"].get(le, {})
        uat_gr = sum(g.get("total_rows", 0) for g in uat_le.get("grids", {}).values())
        cfg_gr = sum(g.get("total_rows", 0) for g in cfg_le.get("grids", {}).values())
        le_mm = sum(1 for r in rows if r.get(f"{le}_Match") == "MISMATCH")
        total_mismatches += le_mm
        sum_rows.append([le, uat_le.get("status", "N/A"), cfg_le.get("status", "N/A"),
                         uat_gr, cfg_gr, le_mm])
    sum_rows.append([])
    sum_rows.append(["Total Mismatches", total_mismatches])

    for ri, rd in enumerate(sum_rows, 1):
        for ci, v in enumerate(rd, 1):
            c = ws_sum.cell(row=ri, column=ci, value=v)
            c.border = tb
            if ri in (1, 10) or ri >= len(sum_rows) - 1:
                c.font = bold
            if ri == 10:
                c.font, c.fill = hf, hfill
    for ci in range(1, 7):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 28

    # ── Grid sheets (D365-style table per grid) ──
    # Collect all grid names from both UAT and Config results
    all_grid_names = set()
    for le in legal_entities:
        for env_result in [uat_result, config_result]:
            le_data = env_result["legal_entities"].get(le, {})
            all_grid_names.update(le_data.get("grids", {}).keys())

    for grid_name in sorted(all_grid_names):
        # Collect column labels from all LEs/envs
        col_labels = []
        col_seen = set()
        for le in legal_entities:
            for env_result in [uat_result, config_result]:
                le_data = env_result["legal_entities"].get(le, {})
                gdata = le_data.get("grids", {}).get(grid_name, {})
                for col in gdata.get("columns", []):
                    lbl = col.get("label", col.get("name", ""))
                    if lbl and lbl not in col_seen:
                        col_labels.append(lbl)
                        col_seen.add(lbl)

        if not col_labels:
            continue

        # Sheet name (max 31 chars for Excel)
        sheet_name = grid_name[:28] if len(grid_name) > 28 else grid_name
        ws_grid = wb.create_sheet(sheet_name)

        # Headers: #, [grid columns...], Env, LE
        grid_headers = ["#"] + col_labels + ["Env", "Legal Entity"]
        ncols = len(grid_headers)

        write_header(ws_grid, grid_headers)

        ri = 2
        seq = 0

        for le in legal_entities:
            uat_le = uat_result["legal_entities"].get(le, {})
            cfg_le = config_result["legal_entities"].get(le, {})
            uat_grid = uat_le.get("grids", {}).get(grid_name, {})
            cfg_grid = cfg_le.get("grids", {}).get(grid_name, {})
            uat_rows = uat_grid.get("rows", [])
            cfg_rows = cfg_grid.get("rows", [])

            if not uat_rows and not cfg_rows:
                continue

            # LE separator row
            for ci in range(1, ncols + 1):
                c = ws_grid.cell(row=ri, column=ci)
                c.fill = le_sep_fill
                c.border = tb
            ws_grid.cell(row=ri, column=1, value=le).font = le_sep_font
            ws_grid.cell(row=ri, column=1).fill = le_sep_fill
            ws_grid.cell(row=ri, column=2,
                         value=f"UAT: {len(uat_rows)} rows | Config: {len(cfg_rows)} rows"
                         ).font = le_sep_font
            ws_grid.cell(row=ri, column=2).fill = le_sep_fill
            ri += 1

            # UAT rows
            for row_data in uat_rows:
                seq += 1
                ws_grid.cell(row=ri, column=1, value=seq).border = tb
                for ci, col_lbl in enumerate(col_labels, 2):
                    val = row_data.get(col_lbl, "")
                    c = ws_grid.cell(row=ri, column=ci, value=val)
                    c.border = tb
                c_env = ws_grid.cell(row=ri, column=ncols - 1, value="UAT")
                c_env.border = tb
                c_env.font = Font(bold=True, color="2F5496", size=10)
                c_le = ws_grid.cell(row=ri, column=ncols, value=le)
                c_le.border = tb
                ri += 1

            # Config rows
            for row_data in cfg_rows:
                seq += 1
                ws_grid.cell(row=ri, column=1, value=seq).border = tb
                for ci, col_lbl in enumerate(col_labels, 2):
                    val = row_data.get(col_lbl, "")
                    c = ws_grid.cell(row=ri, column=ci, value=val)
                    c.border = tb
                c_env = ws_grid.cell(row=ri, column=ncols - 1, value="Config")
                c_env.border = tb
                c_env.font = Font(bold=True, color="548235", size=10)
                c_le = ws_grid.cell(row=ri, column=ncols, value=le)
                c_le.border = tb
                ri += 1

        auto_width(ws_grid, ncols, ri)
        ws_grid.freeze_panes = "B2"
        ws_grid.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ri - 1}"

    # ── Fields Comparison sheet (non-grid controls only) ──
    ws_fields = wb.create_sheet("Fields")
    field_headers = ["#", "Control Name", "Label", "Legal Entity",
                     "UAT Value", "Config Value", "Match"]
    write_header(ws_fields, field_headers)

    fri = 2
    fseq = 0
    for row in rows:
        # Skip grid rows (cells, summaries) — they're in their own sheets
        if row["type"] in ("grid", "grid_cell") or row["control"].startswith("[GRID]"):
            continue
        # Skip row details too — they'll go in a separate sheet
        if "].detail." in row["control"]:
            continue

        has_any = False
        for le in legal_entities:
            if row.get(f"{le}_UAT") or row.get(f"{le}_Config"):
                has_any = True
                break
        if not has_any:
            continue

        for le in legal_entities:
            uat_val = row.get(f"{le}_UAT", "")
            cfg_val = row.get(f"{le}_Config", "")
            match_val = row.get(f"{le}_Match", "")

            if not uat_val and not cfg_val:
                continue

            fseq += 1
            vals = [fseq, row["control"], row["label"], le,
                    uat_val, cfg_val, match_val]

            for ci, v in enumerate(vals, 1):
                c = ws_fields.cell(row=fri, column=ci, value=v)
                c.border = tb
                if match_val == "MISMATCH":
                    c.fill = mismatch_fill
                    if ci in (5, 6, 7):
                        c.font = mismatch_font
                elif match_val == "Match":
                    c.fill = match_fill

            fri += 1

    if fri == 2:
        ws_fields.cell(row=2, column=1, value="No field controls found").font = bold

    auto_width(ws_fields, len(field_headers), fri)
    ws_fields.freeze_panes = "D2"
    if fri > 2:
        ws_fields.auto_filter.ref = f"A1:{get_column_letter(len(field_headers))}{fri - 1}"

    wb.save(filepath)
    print(f"\nExcel saved: {filepath}")
    return filepath


def main():
    legal_entities = ["MY30", "MY60", "SG60"]
    mi_name = "PmfFormulaTable"
    mi_type = "Display"

    print("=" * 70)
    print("D365 Form Control Comparison: UAT vs Config")
    print("=" * 70)
    print(f"  Form:  {mi_name}")
    print(f"  LEs:   {', '.join(legal_entities)}")
    print(f"  UAT:   ENV1 (Env1 UAT)")
    print(f"  Config: ENV4 (Env4 Config)")
    print()

    # Extract from UAT (ENV1)
    print(">>> Extracting from UAT (ENV1)...")
    uat_extractor = FormControlExtractor(env_key="ENV1")
    uat_result = uat_extractor.extract(mi_name, legal_entities, mi_type)

    # Extract from Config (ENV4)
    print("\n>>> Extracting from Config (ENV4)...")
    config_extractor = FormControlExtractor(env_key="ENV4")
    config_result = config_extractor.extract(mi_name, legal_entities, mi_type)

    # Build comparison
    print("\n>>> Building comparison...")
    rows = build_comparison(uat_result, config_result, legal_entities)

    # Print summary
    mismatches = [r for r in rows if r["has_mismatch"]]
    print(f"\n  Total controls compared: {len(rows)}")
    print(f"  Controls with mismatches: {len(mismatches)}")

    if mismatches:
        print(f"\n  MISMATCHES:")
        for r in mismatches:
            print(f"    {r['control']} ({r['label']})")
            for le in legal_entities:
                if r.get(f"{le}_Match") == "MISMATCH":
                    print(f"      {le}: UAT={r[f'{le}_UAT']!r} vs Config={r[f'{le}_Config']!r}")

    # Export
    filepath = to_excel(rows, legal_entities, uat_result, config_result)
    print(f"\nDone! Open: {filepath}")


if __name__ == "__main__":
    main()
