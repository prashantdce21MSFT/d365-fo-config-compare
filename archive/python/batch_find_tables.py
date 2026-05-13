"""
Batch script to find all tables/entities used by each of the 276 CDD forms.
Opens each form via MCP, extracts form name, grids, tabs, and searches for
related CDR/OData entities. Outputs Excel files in batches of 5.
"""
import sys
import os
import json
import time
import traceback
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from d365_mcp_client import D365McpClient
from form_reader import (
    open_form, close_form, mcp_call, resolve_menu_item_from_path,
    find_menu_item, _collect_all_grid_rows, _extract_fields,
)

# ── Config ────────────────────────────────────────────────────────────
CONFIG_FILE = r"C:\D365DataValidator\config.json"
CDD_EXCEL = r"C:\D365 Configuration Drift Analysis\output\MYSG_CDD_FormPaths_20260428_230658.xlsx"
OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output\Tables"
MI_CACHE_FILE = r"C:\D365 Configuration Drift Analysis\output\batch_extract_mi_cache.json"
TABLE_CACHE_FILE = os.path.join(OUTPUT_DIR, "table_cache.json")

BATCH_SIZE = int(os.environ.get("BATCH_SIZE", "5"))
ENV_KEY = "ENV1"  # UAT for table discovery


def load_config():
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)


def load_mi_cache():
    if os.path.exists(MI_CACHE_FILE):
        with open(MI_CACHE_FILE, "r") as f:
            return json.load(f)
    return {}


def load_table_cache():
    if os.path.exists(TABLE_CACHE_FILE):
        with open(TABLE_CACHE_FILE, "r") as f:
            return json.load(f)
    return {}


def save_table_cache(cache):
    with open(TABLE_CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def load_cdd_items():
    """Load the 276 items with Screenshot Path from the CDD Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(CDD_EXCEL, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        sp = d.get("Screenshot Path (from images)", "")
        if sp and str(sp).strip():
            items.append({
                "ado_id": str(d.get("ADO ID", "")),
                "title": str(d.get("Title", "")),
                "screenshot_path": str(sp).strip(),
                "navigation_path": str(d.get("Navigation Path", "") or ""),
                "mi_value": str(d.get("mi= Value", "") or ""),
            })
    return items


def find_grids_and_tabs(form_obj):
    """Recursively find all grid names and tab names in a form."""
    grids = []
    tabs = []

    def _walk(obj, depth=0):
        if not isinstance(obj, dict) or depth > 8:
            return
        for key, val in obj.items():
            if key == "Grid" and isinstance(val, dict):
                for gn, gi in val.items():
                    if isinstance(gi, dict):
                        row_count = len(gi.get("Rows", []))
                        cols = []
                        if gi.get("Rows"):
                            first = gi["Rows"][0]
                            cols = list(first.get("Cells", {}).keys())
                        grids.append({"name": gn, "rows": row_count, "columns": cols})
            elif key in ("Tab", "TabPage") and isinstance(val, dict):
                for tn, ti in val.items():
                    if isinstance(ti, dict):
                        label = ti.get("Label", "")
                        tabs.append({"name": tn, "label": label})
                        _walk(ti, depth + 1)
            elif key == "Children" and isinstance(val, dict):
                _walk(val, depth)
            elif isinstance(val, dict) and key not in (
                "Rows", "Cells", "Button", "MenuButton", "ButtonGroup",
            ):
                _walk(val, depth + 1)

    _walk(form_obj)
    return grids, tabs


def find_detail_fields_from_row(client, form_result, grid_name):
    """Select first row and extract all detail/tab field labels."""
    fields = set()
    try:
        sel = mcp_call(client, "form_select_grid_row", {
            "gridName": grid_name, "rowNumber": 1, "marking": "Unmarked",
        })
        if "raw" not in sel:
            sel_form = sel.get("FormState", {}).get("Form", {})
            detail = {}
            _extract_fields(sel_form, detail)
            for fname, finfo in detail.items():
                fields.add(finfo.get("label", fname))
            # Tab fields
            for tab_name, tab_info in sel_form.get("Tab", {}).items():
                if not isinstance(tab_info, dict):
                    continue
                children = tab_info.get("Children", {})
                tf = {}
                _extract_fields(children, tf)
                for fname, finfo in tf.items():
                    fields.add(finfo.get("label", fname))
    except Exception:
        pass
    return sorted(fields)


def search_entities(client, search_term):
    """Search for OData/CDR entities matching the search term."""
    entities = []
    try:
        result = client.find_entity_types(search_term)
        if isinstance(result, str):
            result = json.loads(result)
        matches = result.get("Matches", [])
        for m in matches:
            entities.append({
                "name": m.get("Name", ""),
                "label": m.get("Label", ""),
            })
    except Exception:
        pass
    return entities


def get_entity_fields(client, entity_name):
    """Get field count and field names for an entity."""
    try:
        result = client.get_entity_metadata(entity_name)
        if isinstance(result, str):
            result = json.loads(result)
        fields = result.get("Fields", {})
        return {
            "field_count": len(fields),
            "field_names": sorted(fields.keys()),
            "entity_real_name": result.get("Name", entity_name),
        }
    except Exception:
        return {"field_count": 0, "field_names": [], "entity_real_name": entity_name}


def extract_tables_for_item(env_config, mi_name, mi_type, ado_id, nav_path):
    """Open a form, extract its structure, and find related entities/tables."""
    client = D365McpClient(env_config)
    client.connect()

    result = {
        "ado_id": ado_id,
        "nav_path": nav_path,
        "mi_name": mi_name,
        "mi_type": mi_type,
        "form_name": "",
        "form_caption": "",
        "grids": [],
        "tabs": [],
        "detail_fields": [],
        "entities": [],
        "primary_table": "",
        "all_tables": [],
        "status": "",
    }

    try:
        # Open form
        form_result = open_form(client, mi_name, mi_type, "MY30")
        fs = form_result.get("FormState", {})
        result["form_name"] = fs.get("Name", "")
        result["form_caption"] = fs.get("Caption", "")
        form_obj = fs.get("Form", {})

        # Find grids and tabs
        grids, tabs = find_grids_and_tabs(form_obj)
        result["grids"] = grids
        result["tabs"] = tabs

        # Primary table = form name (in D365, form name usually matches main table)
        primary_table = result["form_name"]
        result["primary_table"] = primary_table
        all_tables = {primary_table}

        # If grid has rows, get detail fields from first row
        grid_with_rows = None
        grid_name_for_detail = None
        for g in grids:
            if g["rows"] > 0:
                grid_with_rows = g
                grid_name_for_detail = g["name"]
                break
            elif not grid_name_for_detail:
                grid_name_for_detail = g["name"]

        if grid_with_rows:
            detail_fields = find_detail_fields_from_row(
                client, form_result, grid_with_rows["name"]
            )
            result["detail_fields"] = detail_fields

        # Grid names that differ from "Grid" are often table names
        for g in grids:
            if g["name"] != "Grid":
                all_tables.add(g["name"])

        # Search for CDR entities using the form/table name
        search_terms = set()
        search_terms.add(primary_table)
        # Also search the mi_name if different
        if mi_name.lower() != primary_table.lower():
            search_terms.add(mi_name)

        all_entities = []
        for st in search_terms:
            entities = search_entities(client, st)
            # Filter relevant ones
            for e in entities:
                ename_lower = e["name"].lower()
                st_lower = st.lower()
                if (st_lower in ename_lower or
                    ename_lower.startswith(st_lower[:6])):
                    all_entities.append(e)
                    all_tables.add(e["name"])

        # Get field info for CDR entities
        for e in all_entities:
            if "CDR" in e["name"]:
                finfo = get_entity_fields(client, e["name"])
                e["field_count"] = finfo["field_count"]
                e["real_name"] = finfo["entity_real_name"]
                # The CDR entity real_name usually = {TableName}CDREntity
                real = finfo["entity_real_name"]
                if real.endswith("CDREntity"):
                    base_table = real[:-9]  # Remove CDREntity suffix
                    all_tables.add(base_table)

        result["entities"] = all_entities
        result["all_tables"] = sorted(all_tables)
        result["status"] = "OK"

        close_form(client)

    except Exception as e:
        result["status"] = f"ERROR: {str(e)[:200]}"
        try:
            close_form(client)
        except Exception:
            pass

    return result


def generate_batch_excel(batch_results, batch_num, batch_size):
    """Generate an Excel file for a batch of results."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    start = (batch_num - 1) * batch_size + 1
    end = start + len(batch_results) - 1
    filename = f"Tables_batch{batch_num}_{start}-{end}_{ts}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    err_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # ── Summary Sheet ──
    ws_sum = wb.create_sheet("Summary")
    sum_headers = [
        "ADO ID", "Navigation Path", "Menu Item", "Form Name", "Caption",
        "Primary Table", "Grid Count", "Tab Count", "Detail Field Count",
        "Entity Count", "All Tables", "Status",
    ]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, r in enumerate(batch_results, 2):
        vals = [
            r["ado_id"],
            r["nav_path"],
            r["mi_name"],
            r["form_name"],
            r["form_caption"],
            r["primary_table"],
            len(r["grids"]),
            len(r["tabs"]),
            len(r["detail_fields"]),
            len(r["entities"]),
            ", ".join(r["all_tables"]),
            r["status"],
        ]
        fill = ok_fill if r["status"] == "OK" else err_fill
        for ci, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            cell.fill = fill

    # Auto-width
    for ci in range(1, len(sum_headers) + 1):
        max_len = len(str(sum_headers[ci - 1]))
        for ri in range(2, len(batch_results) + 2):
            val = ws_sum.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 2, 60)

    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(sum_headers))}{len(batch_results) + 1}"

    # ── Detail Sheet — one row per table per ADO item ──
    ws_det = wb.create_sheet("Tables Detail")
    det_headers = [
        "ADO ID", "Navigation Path", "Form Name", "Table/Entity Name",
        "Entity Label", "Entity Type", "Field Count",
    ]
    for ci, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    for r in batch_results:
        # Primary table
        ws_det.cell(row=row_num, column=1, value=r["ado_id"]).border = thin_border
        ws_det.cell(row=row_num, column=2, value=r["nav_path"]).border = thin_border
        ws_det.cell(row=row_num, column=3, value=r["form_name"]).border = thin_border
        ws_det.cell(row=row_num, column=4, value=r["primary_table"]).border = thin_border
        ws_det.cell(row=row_num, column=5, value=r["form_caption"]).border = thin_border
        ws_det.cell(row=row_num, column=6, value="Primary (Form)").border = thin_border
        ws_det.cell(row=row_num, column=7, value="").border = thin_border
        row_num += 1

        # Grid-based tables
        for g in r["grids"]:
            if g["name"] != "Grid" and g["name"] != r["primary_table"]:
                ws_det.cell(row=row_num, column=1, value=r["ado_id"]).border = thin_border
                ws_det.cell(row=row_num, column=2, value=r["nav_path"]).border = thin_border
                ws_det.cell(row=row_num, column=3, value=r["form_name"]).border = thin_border
                ws_det.cell(row=row_num, column=4, value=g["name"]).border = thin_border
                ws_det.cell(row=row_num, column=5, value="").border = thin_border
                ws_det.cell(row=row_num, column=6, value="Grid").border = thin_border
                ws_det.cell(row=row_num, column=7, value=len(g["columns"])).border = thin_border
                row_num += 1

        # CDR/OData entities
        for e in r["entities"]:
            if e["name"] != r["primary_table"]:
                ws_det.cell(row=row_num, column=1, value=r["ado_id"]).border = thin_border
                ws_det.cell(row=row_num, column=2, value=r["nav_path"]).border = thin_border
                ws_det.cell(row=row_num, column=3, value=r["form_name"]).border = thin_border
                ws_det.cell(row=row_num, column=4, value=e["name"]).border = thin_border
                ws_det.cell(row=row_num, column=5, value=e.get("label", "")).border = thin_border
                etype = "CDR Entity" if "CDR" in e["name"] else "OData Entity"
                ws_det.cell(row=row_num, column=6, value=etype).border = thin_border
                ws_det.cell(row=row_num, column=7, value=e.get("field_count", "")).border = thin_border
                row_num += 1

    # Auto-width for detail sheet
    for ci in range(1, len(det_headers) + 1):
        max_len = len(str(det_headers[ci - 1]))
        for ri in range(2, min(row_num, 100)):
            val = ws_det.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws_det.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 2, 60)

    ws_det.freeze_panes = "A2"

    # ── Fields Sheet — detail fields found for each form ──
    ws_fld = wb.create_sheet("Form Fields")
    fld_headers = ["ADO ID", "Form Name", "Caption", "Field Label"]
    for ci, h in enumerate(fld_headers, 1):
        cell = ws_fld.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    for r in batch_results:
        for fl in r.get("detail_fields", []):
            ws_fld.cell(row=row_num, column=1, value=r["ado_id"]).border = thin_border
            ws_fld.cell(row=row_num, column=2, value=r["form_name"]).border = thin_border
            ws_fld.cell(row=row_num, column=3, value=r["form_caption"]).border = thin_border
            ws_fld.cell(row=row_num, column=4, value=fl).border = thin_border
            row_num += 1

    ws_fld.freeze_panes = "A2"

    # Move Summary first
    wb.move_sheet("Summary", offset=-2)

    wb.save(filepath)
    print(f"  Excel saved: {filepath}")
    return filepath


def main():
    config = load_config()
    env_config = config["environments"][ENV_KEY]
    mi_cache = load_mi_cache()
    table_cache = load_table_cache()

    items = load_cdd_items()
    print(f"Total CDD items with Screenshot Path: {len(items)}")
    print(f"MI cache entries: {len(mi_cache)}")
    print(f"Table cache entries: {len(table_cache)}")

    # Filter to items not yet processed
    pending = []
    for item in items:
        ado_id = item["ado_id"]
        if ado_id not in table_cache:
            pending.append(item)

    print(f"Pending: {len(pending)}")
    if not pending:
        print("All items processed!")
        return

    # Process batch
    batch = pending[:BATCH_SIZE]
    batch_num = (len(table_cache) // BATCH_SIZE) + 1
    print(f"\n{'='*60}")
    print(f"Processing batch {batch_num}: items {len(table_cache)+1} to {len(table_cache)+len(batch)}")
    print(f"{'='*60}")

    batch_results = []

    for idx, item in enumerate(batch, 1):
        ado_id = item["ado_id"]
        nav_path = item["screenshot_path"]
        print(f"\n[{idx}/{len(batch)}] ADO {ado_id}: {nav_path}")

        # Resolve MI from cache
        # Try screenshot_path first, then navigation_path
        mi_info = mi_cache.get(nav_path) or mi_cache.get(item["navigation_path"])
        if not mi_info and item["mi_value"]:
            mi_name = item["mi_value"]
            mi_type = "Display"
            mi_info = {"mi_name": mi_name, "mi_type": mi_type}

        if not mi_info:
            # Try to resolve
            print("  Resolving menu item...")
            try:
                client = D365McpClient(env_config)
                client.connect()
                mi_name, mi_type, hint = resolve_menu_item_from_path(
                    client, nav_path, "MY30"
                )
                close_form(client)
                if mi_name:
                    mi_info = {"mi_name": mi_name, "mi_type": mi_type}
                    print(f"  Resolved: {mi_name} ({mi_type})")
                else:
                    print("  Could not resolve menu item")
            except Exception as e:
                print(f"  Resolve error: {e}")
                try:
                    close_form(client)
                except Exception:
                    pass

        if not mi_info:
            result = {
                "ado_id": ado_id,
                "nav_path": nav_path,
                "mi_name": "",
                "mi_type": "",
                "form_name": "",
                "form_caption": "",
                "grids": [],
                "tabs": [],
                "detail_fields": [],
                "entities": [],
                "primary_table": "",
                "all_tables": [],
                "status": "SKIP: MI not resolved",
            }
            batch_results.append(result)
            table_cache[ado_id] = result
            save_table_cache(table_cache)
            continue

        mi_name = mi_info["mi_name"]
        mi_type = mi_info.get("mi_type", "Display")
        print(f"  MI: {mi_name} ({mi_type})")

        result = extract_tables_for_item(
            env_config, mi_name, mi_type, ado_id, nav_path
        )
        batch_results.append(result)
        table_cache[ado_id] = result
        save_table_cache(table_cache)

        print(f"  Form: {result['form_name']} | Tables: {result['all_tables']}")
        print(f"  Grids: {len(result['grids'])}, Tabs: {len(result['tabs'])}, "
              f"Fields: {len(result['detail_fields'])}, Entities: {len(result['entities'])}")
        print(f"  Status: {result['status']}")

        time.sleep(0.5)

    # Generate Excel
    print(f"\nGenerating Excel for batch {batch_num}...")
    generate_batch_excel(batch_results, batch_num, BATCH_SIZE)

    # Print summary
    print(f"\n{'='*60}")
    print(f"BATCH {batch_num} COMPLETE")
    print(f"{'='*60}")
    ok = sum(1 for r in batch_results if r["status"] == "OK")
    err = sum(1 for r in batch_results if r["status"].startswith("ERROR"))
    skip = sum(1 for r in batch_results if r["status"].startswith("SKIP"))
    print(f"  OK: {ok}, Errors: {err}, Skipped: {skip}")
    print(f"  Total processed: {len(table_cache)} / {len(items)}")
    print(f"  Remaining: {len(items) - len(table_cache)}")


if __name__ == "__main__":
    main()
