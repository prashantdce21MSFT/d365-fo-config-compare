"""
merge_results_to_excel.py
=========================
Reads H + M JSON results and writes Correct_MI / Correct_URL / Search_Status
columns into CDD_FormMapping_WithURLs.xlsx against each ADO_ID.
Overwrites any previously written columns of the same name.
"""

import json
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

EXCEL_IN     = r"C:/D365 Configuration Drift Analysis/output/CDD_FormMapping_WithURLs.xlsx"
EXCEL_OUT    = r"C:/D365 Configuration Drift Analysis/output/CDD_FormMapping_WithURLs.xlsx"
JSON_H       = r"C:/D365 Configuration Drift Analysis/output/correct_mi_results.json"
JSON_M       = r"C:/D365 Configuration Drift Analysis/output/correct_mi_results_M.json"
HEADER_ROW   = 11

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")


def load_results(*paths):
    combined = {}
    for p in paths:
        path = Path(p)
        if path.exists():
            raw = json.loads(path.read_text(encoding='utf-8'))
            combined.update({int(k): v for k, v in raw.items()})
            print(f"Loaded {len(raw)} results from {path.name}")
        else:
            print(f"WARNING: {path.name} not found, skipping")
    return combined


def find_or_create_col(ws, header_row, name):
    """Find existing column with this header, or return next available column."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == name:
            return col
    return ws.max_column + 1


def main():
    results = load_results(JSON_H, JSON_M)
    print(f"Total results loaded: {len(results)}")

    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb.active

    # Find or create the 3 output columns
    mi_col     = find_or_create_col(ws, HEADER_ROW, 'Correct_MI')
    url_col    = find_or_create_col(ws, HEADER_ROW, 'Correct_URL')
    status_col = find_or_create_col(ws, HEADER_ROW, 'Search_Status')

    for col, hdr in [(mi_col, 'Correct_MI'), (url_col, 'Correct_URL'), (status_col, 'Search_Status')]:
        cell = ws.cell(row=HEADER_ROW, column=col, value=hdr)
        cell.font = Font(bold=True)

    written = 0
    for excel_row in range(HEADER_ROW + 1, ws.max_row + 1):
        raw = ws.cell(row=excel_row, column=1).value
        try:
            ado_id = int(raw) if raw else None
        except (ValueError, TypeError):
            continue
        if ado_id not in results:
            continue

        r = results[ado_id]
        ws.cell(row=excel_row, column=mi_col,     value=r.get('mi', ''))
        ws.cell(row=excel_row, column=url_col,     value=r.get('url', ''))
        ws.cell(row=excel_row, column=status_col,  value=r.get('status', ''))

        sc = ws.cell(row=excel_row, column=status_col)
        s  = r.get('status', '')
        if s == 'FOUND':
            sc.fill = GREEN
        elif s in ('TIMEOUT', 'SESSION_EXPIRED') or 'ERROR' in s:
            sc.fill = RED
        else:
            sc.fill = YELLOW

        written += 1

    for col in [mi_col, url_col, status_col]:
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or ''))
             for r in range(HEADER_ROW, ws.max_row + 1)), default=10
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)

    wb.save(EXCEL_OUT)
    found = sum(1 for v in results.values() if v['status'] == 'FOUND')
    not_found = len(results) - found
    print(f"\nWrote {written} rows | {found} FOUND | {not_found} not found")
    print(f"Saved: {EXCEL_OUT}")


if __name__ == '__main__':
    main()
