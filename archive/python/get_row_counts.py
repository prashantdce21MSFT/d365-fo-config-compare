"""
Query row counts for all 505 tables from CDD_Form_Tables_Merged.xlsx.
D365 AOT names -> uppercase DB names (e.g. LedgerParameters -> LEDGERPARAMETERS).
"""
import pyodbc
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

CONN_STR = (
    'DRIVER={SQL Server};'
    'SERVER=spartan-srv-apac-d365opsprod-0d7f250ccb67.database.windows.net;'
    'DATABASE=db_d365opsprod_cbasiadynuat_ax_20260117_02184426_3946;'
    'UID=JIT-Prashan-nldn1fmf;'
    'PWD=9G6dobO7cP9fmKP-JaOQWpvF-lz2_ksAaa7iX1wEZpHgvQdV5mABaBIMIfzHC60DkR3d1qcniOviXYjlo3GggimwoA;'
    'Encrypt=yes;TrustServerCertificate=yes;'
)

# ── 1. Extract all unique table names from merged Excel ───────────────────────
wb = load_workbook('output/CDD_Form_Tables_Merged.xlsx', read_only=True, data_only=True)
ws = wb['By Form']
aot_tables = set()
for r in ws.iter_rows(min_row=4, values_only=True):
    if r[8]:
        for t in str(r[8]).split('|'):
            t = t.strip()
            if t:
                aot_tables.add(t)
wb.close()
print(f'AOT table names extracted: {len(aot_tables)}')

# ── 2. Get all actual table names from DB (uppercase lookup dict) ─────────────
conn = pyodbc.connect(CONN_STR)
cur  = conn.cursor()
cur.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE'")
db_tables_upper = {r[0].upper(): r[0] for r in cur.fetchall()}
print(f'DB tables found: {len(db_tables_upper)}')

# ── 3. Match AOT names -> DB names ────────────────────────────────────────────
matched   = {}   # aot_name -> actual_db_name
unmatched = []

for aot in sorted(aot_tables):
    upper = aot.upper()
    if upper in db_tables_upper:
        matched[aot] = db_tables_upper[upper]
    elif upper + 'TABLE' in db_tables_upper:
        matched[aot] = db_tables_upper[upper + 'TABLE']
    elif upper + 'S' in db_tables_upper:
        matched[aot] = db_tables_upper[upper + 'S']
    else:
        unmatched.append(aot)

print(f'Matched: {len(matched)}  |  Unmatched: {len(unmatched)}')
if unmatched:
    print('Unmatched:', unmatched[:20])

# ── 4. Get row counts using fast sys.partitions approach ─────────────────────
# Get all row counts in one query (much faster than COUNT(*) per table)
cur.execute("""
    SELECT
        UPPER(t.name) as tbl_upper,
        t.name        as tbl_name,
        SUM(p.rows)   as row_count
    FROM sys.tables t
    JOIN sys.partitions p ON t.object_id = p.object_id
    WHERE p.index_id IN (0, 1)
    GROUP BY t.name
""")
db_row_counts = {r[0]: r[2] for r in cur.fetchall()}  # UPPER(name) -> count
conn.close()

# ── 5. Build results ──────────────────────────────────────────────────────────
results = []
for aot in sorted(aot_tables):
    upper = aot.upper()
    db_name = matched.get(aot)
    if db_name:
        count = db_row_counts.get(db_name.upper(), 0)
        results.append((aot, db_name, count, 'Found'))
    else:
        results.append((aot, '', None, 'Not in DB'))

results.sort(key=lambda x: (0 if x[3] == 'Found' else 1, -(x[2] or 0)))

# ── 6. Write Excel ────────────────────────────────────────────────────────────
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Row Counts"

hdr_font  = Font(bold=True, color="FFFFFF", size=10)
hdr_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
orange    = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
thin      = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'),  bottom=Side(style='thin'))

found_count   = sum(1 for r in results if r[3] == 'Found')
missing_count = len(results) - found_count

ws2.merge_cells('A1:E1')
ws2['A1'] = (f"D365 UAT — Row counts for CDD-related tables  |  "
             f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
             f"Found: {found_count}  |  Not in DB: {missing_count}")
ws2['A1'].font = Font(bold=True, size=12)

headers = ["AOT Table Name (D365)", "DB Table Name", "Row Count", "Status", "Notes"]
for ci, h in enumerate(headers, 1):
    c = ws2.cell(row=3, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = thin
    c.alignment = Alignment(horizontal='center')

for ri, (aot, db_name, count, status) in enumerate(results, 4):
    is_cba = aot.upper().startswith('CBA')
    note = ''
    if status == 'Not in DB':
        note = 'CBA custom — may not exist in this env' if is_cba else 'Name mismatch or view'

    vals = [aot, db_name, count, status, note]
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.border = thin
        c.font = Font(size=9)
        c.alignment = Alignment(vertical='top')

    if status == 'Not in DB':
        fill = yellow if is_cba else red_fill
    elif is_cba:
        fill = orange
    else:
        fill = green
    for ci in range(1, 6):
        ws2.cell(row=ri, column=ci).fill = fill

for ci, w in enumerate([38, 38, 14, 14, 40], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.auto_filter.ref = f"A3:E{3 + len(results)}"
ws2.freeze_panes = "A4"

# Summary sheet
ws3 = wb2.create_sheet("Summary")
ws3['A1'] = "Summary"; ws3['A1'].font = Font(bold=True, size=14)
total_rows = sum(r[2] for r in results if r[2] is not None)
top10 = sorted([r for r in results if r[2]], key=lambda x: -x[2])[:10]
summary = [
    ("Total AOT tables", len(results)),
    ("Found in DB", found_count),
    ("Not found in DB", missing_count),
    ("", ""),
    ("Total rows across all tables", total_rows),
    ("", ""),
    ("Top 10 largest tables", ""),
]
for lbl, val in summary:
    ri = len(summary) + 3
for idx, (lbl, val) in enumerate(summary, 3):
    ws3.cell(row=idx, column=1, value=lbl).font = Font(bold=bool(lbl))
    ws3.cell(row=idx, column=2, value=val)
ws3.cell(row=idx+2, column=1, value="Top 10 largest tables").font = Font(bold=True)
for i, (aot, db, count, _) in enumerate(top10, idx+3):
    ws3.cell(row=i, column=1, value=aot)
    ws3.cell(row=i, column=2, value=count)
ws3.column_dimensions['A'].width = 38
ws3.column_dimensions['B'].width = 20

out = r'output\CDD_Table_RowCounts.xlsx'
wb2.save(out)
print(f'\nSaved: {out}')
print(f'Total rows across all found tables: {total_rows:,}')
print('\nTop 10 largest:')
for aot, db, count, _ in top10:
    print(f'  {aot:<45} {count:>10,}')
print(f'\nNot in DB ({missing_count}):')
for aot, db, count, status in results:
    if status == 'Not in DB':
        print(f'  {aot}')
