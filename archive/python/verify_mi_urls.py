"""
CDD Form MI Name URL Verifier
==============================
Reads CDD_FormMapping.xlsx, adds a URL column for H-confidence items,
then uses Playwright to verify each URL loads a valid D365 form.

Usage — parallel batches (5 windows, each signs in independently):
    python verify_mi_urls.py --skip 0   --limit 92  --batch-json batch_0.json
    python verify_mi_urls.py --skip 92  --limit 92  --batch-json batch_1.json
    python verify_mi_urls.py --skip 184 --limit 92  --batch-json batch_2.json
    python verify_mi_urls.py --skip 276 --limit 92  --batch-json batch_3.json
    python verify_mi_urls.py --skip 368 --limit 92  --batch-json batch_4.json

Usage — merge all batches into final Excel:
    python verify_mi_urls.py --merge

Other:
    python verify_mi_urls.py --dry-run
    python verify_mi_urls.py --out custom_output.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

BASE_URL   = "https://example-uat.sandbox.operations.dynamics.com/?cmp=my30&mi={mi}"
EXCEL_IN   = r"C:/D365 Configuration Drift Analysis/output/CDD_FormMapping.xlsx"
EXCEL_OUT  = r"C:/D365 Configuration Drift Analysis/output/CDD_FormMapping_WithURLs.xlsx"
BATCH_DIR  = r"C:/D365 Configuration Drift Analysis/output/batches"
HEADER_ROW = 11


def load_data(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0, header=10)
    df.columns = [
        'ADO_ID', 'ADO_Title', 'Display_Name', 'MI_Name',
        'Table_Name_Manual', 'Confidence', 'State', 'Custom_Standard', 'Verified'
    ]
    df = df.dropna(subset=['ADO_ID'])
    df['ADO_ID'] = df['ADO_ID'].astype(int)
    return df


def build_url(mi_name) -> str:
    if pd.isna(mi_name) or str(mi_name).strip() == '':
        return ''
    return BASE_URL.format(mi=str(mi_name).strip())


def verify_urls(df: pd.DataFrame, skip: int = 0, limit: int | None = None) -> dict:
    """Returns {ado_id: {url, status, title}} for the batch."""
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

    h_df = df[df['Confidence'] == 'H'].copy()
    h_df = h_df.iloc[skip:]
    if limit:
        h_df = h_df.head(limit)

    total = len(h_df)
    batch_label = f"items {skip+1}–{skip+total}"
    print(f"\nVerifying {total} H-confidence URLs ({batch_label}) via Playwright...")
    print("A browser window will open — sign in when prompted.\n")

    results = {}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, args=["--start-maximized"])
        ctx     = browser.new_context(viewport={"width": 1600, "height": 900})
        page    = ctx.new_page()

        first_url = h_df.iloc[0]['URL']
        print(f"  Navigating to first URL for sign-in: {first_url}")
        page.goto(first_url)

        try:
            page.wait_for_url(
                lambda u: "operations.dynamics.com" in u and "login" not in u.lower(),
                timeout=600_000
            )
            page.wait_for_timeout(3000)
        except PWTimeout:
            print("  Timed out waiting for login. Exiting.")
            browser.close()
            sys.exit(1)

        print("  Signed in. Starting verification loop...\n")

        for i, (_, row) in enumerate(h_df.iterrows()):
            url     = row['URL']
            mi_name = row['MI_Name']
            ado_id  = int(row['ADO_ID'])

            if not url:
                results[ado_id] = {'url': '', 'status': 'BLANK_MI', 'title': ''}
                print(f"  [{i+1}/{total}] SKIP (blank MI): {mi_name}")
                continue

            try:
                resp = page.goto(url, timeout=15_000, wait_until="domcontentloaded")
                page.wait_for_timeout(2000)

                current_url = page.url
                title       = page.title()

                if "login.microsoftonline" in current_url:
                    status = 'SESSION_EXPIRED'
                elif resp and resp.status >= 400:
                    status = f'HTTP_{resp.status}'
                elif 'Error' in title or 'error' in title.lower():
                    status = 'ERROR'
                elif "operations.dynamics.com" in current_url and title:
                    status = 'OK'
                elif not title:
                    status = 'BLANK_TITLE'
                else:
                    status = 'UNKNOWN'

                clean_title = title
                for suffix in [' -- Finance and Operations', ' - Finance and Operations',
                                ' | Finance and Operations', ' | Dynamics 365']:
                    clean_title = clean_title.replace(suffix, '')
                clean_title = clean_title.strip()

                results[ado_id] = {'url': url, 'status': status, 'title': clean_title}
                marker = 'OK' if status == 'OK' else 'XX'
                print(f"  [{i+1}/{total}] {marker} {status:15} | {str(mi_name):40} | {clean_title}")

            except PWTimeout:
                results[ado_id] = {'url': url, 'status': 'TIMEOUT', 'title': ''}
                print(f"  [{i+1}/{total}] XX TIMEOUT         | {mi_name}")

            except Exception as e:
                results[ado_id] = {'url': url, 'status': 'ERROR', 'title': str(e)[:60]}
                print(f"  [{i+1}/{total}] XX ERROR           | {mi_name} | {e}")

        browser.close()

    return results


GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
GREY   = PatternFill("solid", fgColor="EEEEEE")


def save_excel(df: pd.DataFrame, out_path: str):
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb.active

    last_col      = ws.max_column
    url_col       = last_col + 1
    status_col    = last_col + 2
    ver_title_col = last_col + 3

    ws.cell(row=HEADER_ROW, column=url_col,       value="URL")
    ws.cell(row=HEADER_ROW, column=status_col,    value="URL_Status")
    ws.cell(row=HEADER_ROW, column=ver_title_col, value="Verified_Title")

    for col in [url_col, status_col, ver_title_col]:
        ws.cell(row=HEADER_ROW, column=col).font = Font(bold=True)

    id_to_row = {int(row['ADO_ID']): row for _, row in df.iterrows()}

    for excel_row in range(HEADER_ROW + 1, ws.max_row + 1):
        ado_id_cell = ws.cell(row=excel_row, column=1).value
        try:
            ado_id = int(ado_id_cell) if ado_id_cell else None
        except (ValueError, TypeError):
            continue
        if ado_id not in id_to_row:
            continue

        row    = id_to_row[ado_id]
        url    = row.get('URL', '')
        status = row.get('URL_Status', '')
        title  = row.get('Verified_Title', '')

        ws.cell(row=excel_row, column=url_col,       value=url    if url    else '')
        ws.cell(row=excel_row, column=status_col,    value=status if pd.notna(status) else '')
        ws.cell(row=excel_row, column=ver_title_col, value=title  if pd.notna(title)  else '')

        sc = ws.cell(row=excel_row, column=status_col)
        if status == 'OK':
            sc.fill = GREEN
        elif status in ('TIMEOUT', 'ERROR', 'NOT_FOUND') or (isinstance(status, str) and status.startswith('HTTP_4')):
            sc.fill = RED
        elif status in ('BLANK_MI', 'UNKNOWN', 'SESSION_EXPIRED', 'BLANK_TITLE'):
            sc.fill = YELLOW
        elif row.get('Confidence') != 'H':
            sc.fill = GREY

    for col in [url_col, status_col, ver_title_col]:
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or ''))
             for r in range(HEADER_ROW, ws.max_row + 1)), default=10
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)

    wb.save(out_path)
    print(f"\nSaved: {out_path}")


def merge_batches(df: pd.DataFrame, out_path: str):
    batch_dir = Path(BATCH_DIR)
    batch_files = sorted(batch_dir.glob("batch_*.json"))
    if not batch_files:
        print(f"No batch_*.json files found in {BATCH_DIR}")
        sys.exit(1)

    merged = {}
    for f in batch_files:
        data = json.loads(f.read_text(encoding='utf-8'))
        merged.update({int(k): v for k, v in data.items()})
        print(f"  Loaded {len(data)} results from {f.name}")

    print(f"  Total merged: {len(merged)} results")

    df['URL']            = df.apply(lambda r: build_url(r['MI_Name']) if r['Confidence'] == 'H' else '', axis=1)
    df['URL_Status']     = ''
    df['Verified_Title'] = ''

    for _, row in df.iterrows():
        ado_id = int(row['ADO_ID'])
        if ado_id in merged:
            df.loc[df['ADO_ID'] == ado_id, 'URL_Status']     = merged[ado_id]['status']
            df.loc[df['ADO_ID'] == ado_id, 'Verified_Title'] = merged[ado_id]['title']

    ok  = (df['URL_Status'] == 'OK').sum()
    err = df[df['Confidence'] == 'H']['URL_Status'].isin(['ERROR','TIMEOUT','NOT_FOUND']).sum()
    print(f"\nResults: {ok} OK  |  {err} failed  |  {len(merged) - ok - err} other")

    save_excel(df, out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run',    action='store_true')
    parser.add_argument('--skip',       type=int, default=0,    help='Skip first N H-rows')
    parser.add_argument('--limit',      type=int, default=None, help='Verify only N H-rows after skip')
    parser.add_argument('--batch-json', default=None,           help='Save results to JSON file (for parallel merge)')
    parser.add_argument('--merge',      action='store_true',    help='Merge all batch JSONs into final Excel')
    parser.add_argument('--out',        default=EXCEL_OUT)
    args = parser.parse_args()

    print(f"Reading: {EXCEL_IN}")
    df = load_data(EXCEL_IN)
    print(f"  {len(df)} total rows  |  {(df['Confidence']=='H').sum()} H-confidence")

    if args.merge:
        merge_batches(df, args.out)
        return

    df['URL'] = df.apply(
        lambda r: build_url(r['MI_Name']) if r['Confidence'] == 'H' else '', axis=1
    )

    if args.dry_run:
        print("\nDry-run mode: skipping browser verification.")
        df['URL_Status']     = df.apply(lambda r: 'NOT_VERIFIED' if r['Confidence'] == 'H' else '', axis=1)
        df['Verified_Title'] = ''
        save_excel(df, args.out)
        return

    results = verify_urls(df, skip=args.skip, limit=args.limit)

    if args.batch_json:
        Path(BATCH_DIR).mkdir(parents=True, exist_ok=True)
        out_file = Path(BATCH_DIR) / args.batch_json
        out_file.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding='utf-8')
        ok  = sum(1 for v in results.values() if v['status'] == 'OK')
        err = sum(1 for v in results.values() if v['status'] in ('ERROR','TIMEOUT','NOT_FOUND'))
        print(f"\nBatch saved: {out_file}  ({ok} OK, {err} failed, {len(results)} total)")
    else:
        # Single-run mode: write directly to Excel
        h_mask = df['Confidence'] == 'H'
        df['URL_Status']     = ''
        df['Verified_Title'] = ''
        for _, row in df[h_mask].iterrows():
            ado_id = int(row['ADO_ID'])
            if ado_id in results:
                df.loc[df['ADO_ID'] == ado_id, 'URL_Status']     = results[ado_id]['status']
                df.loc[df['ADO_ID'] == ado_id, 'Verified_Title'] = results[ado_id]['title']
        ok  = (df['URL_Status'] == 'OK').sum()
        err = df[h_mask]['URL_Status'].isin(['ERROR','TIMEOUT','NOT_FOUND']).sum()
        print(f"\nResults: {ok} OK  |  {err} failed")
        save_excel(df, args.out)


if __name__ == "__main__":
    main()
