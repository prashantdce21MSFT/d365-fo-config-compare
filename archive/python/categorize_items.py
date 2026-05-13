"""
Categorize 606 MY/SG Config Deliverables via D365 ERP MCP Server.
Checks OData entities + Form menu items, generates Excel report.
"""
import json, sys, re, time, os
sys.path.insert(0, r"C:\D365DataValidator")
from d365_mcp_client import D365McpClient

CONFIG_PATH = r"C:\D365DataValidator\config.json"
ITEMS_PATH = r"C:\D365 Configuration Drift Analysis\ado_items.json"
CHECKPOINT = r"C:\D365 Configuration Drift Analysis\cat_checkpoint.json"
EXCEL_PATH = r"C:\D365 Configuration Drift Analysis\MY_SG_606_Categorization.xlsx"

def clean_title(title):
    t = title.strip()
    for _ in range(4):
        t = re.sub(r'^\[.*?\]\s*', '', t)
    for _ in range(3):
        t = re.sub(r'\s*\[.*?\]\s*$', '', t)
    paren_hint = None
    m = re.search(r'\(([^)]+)\)\s*$', t)
    if m:
        candidate = m.group(1)
        if not re.match(r'^[A-Z]{2}(\d+)?(/[A-Z]{2}\d+)*$', candidate):
            paren_hint = candidate
            t = t[:m.start()].strip()
    t = t.replace('->', '>').replace('\u2013', '>').replace(': ', ' > ')
    parts = [p.strip() for p in t.split('>') if p.strip()]
    last = parts[-1] if parts else t
    return parts, last, paren_hint

def get_search_terms(title):
    parts, last, paren_hint = clean_title(title)
    terms = []
    if paren_hint:
        terms.append(paren_hint)
    if last and last not in terms:
        terms.append(last)
    generic = {"parameters","setup","posting","configuration","settings",
               "number sequences","categories","types","codes","groups",
               "methods","profiles","journal names","terms of payment"}
    if last.lower() in generic and len(parts) >= 2:
        combined = f"{parts[-2]} {last}"
        if combined not in terms:
            terms.append(combined)
    # Also try second-to-last if available
    if len(parts) >= 2 and parts[-2] not in terms:
        terms.append(parts[-2])
    return terms

def score_odata(search_term, raw):
    try:
        data = json.loads(raw)
        matches = data.get("Matches", [])
        if not matches:
            return 0, None, None
        sl = search_term.lower().strip()
        sw = set(re.findall(r'\w+', sl))
        best_sc, best_n, best_l = 0, None, None
        for m in matches:
            label = m.get("Label", "").lower()
            score = 0
            if label == sl:
                score = 100
            elif sl in label or label in sl:
                score = 70
            else:
                lw = set(re.findall(r'\w+', label))
                ov = sw & lw
                if ov:
                    score = len(ov) / max(len(sw), 1) * 50
            if score > best_sc:
                best_sc, best_n, best_l = score, m.get("Name"), m.get("Label")
        return best_sc, best_n, best_l
    except:
        return 0, None, None

def score_form(search_term, raw):
    try:
        data = json.loads(raw)
        menu = data.get("Menu", {})
        menu_items = menu.get("MenuItems", {})
        sl = search_term.lower().strip()
        sw = set(re.findall(r'\w+', sl))
        best_sc, best_name, best_text, best_type = 0, None, None, None
        for item_type, items in menu_items.items():
            if not isinstance(items, list):
                continue
            for item in items:
                name = item.get("Name", "")
                text = item.get("Text", "").lower()
                score = 0
                if text == sl:
                    score = 100
                elif sl in text or text in sl:
                    score = 70
                else:
                    tw = set(re.findall(r'\w+', text))
                    ov = sw & tw
                    if ov:
                        score = len(ov) / max(len(sw), 1) * 50
                if score > best_sc:
                    best_sc = score
                    best_name = name
                    best_text = item.get("Text", "")
                    best_type = item_type
        return best_sc, best_name, best_text, best_type
    except:
        return 0, None, None, None

def main():
    with open(CONFIG_PATH) as f:
        config = json.load(f)
    with open(ITEMS_PATH) as f:
        items = json.load(f)

    # Resume from checkpoint
    results = []
    done_ids = set()
    try:
        with open(CHECKPOINT) as f:
            results = json.load(f)
            done_ids = {r['id'] for r in results}
            print(f"Resuming: {len(done_ids)} done")
    except:
        pass

    remaining = len(items) - len(done_ids)
    print(f"Total: {len(items)}, Remaining: {remaining}")

    env = config["environments"]["ENV1"]
    client = D365McpClient(env)
    client.connect()

    total = len(items)
    for idx, item in enumerate(items):
        aid = item['id']
        title = item['title']
        if aid in done_ids:
            continue

        terms = get_search_terms(title)
        pct = (idx + 1) * 100 // total
        print(f"[{idx+1}/{total} {pct}%] {aid}: {title[:60]}")

        o_score, o_entity, o_label = 0, None, None
        f_score, f_name, f_text, f_type = 0, None, None, None

        # OData search
        for term in terms:
            try:
                raw = client.find_entity_types(term, top=5)
                if raw and "No entities found" not in raw:
                    sc, ent, lbl = score_odata(term, raw)
                    if sc > o_score:
                        o_score, o_entity, o_label = sc, ent, lbl
                time.sleep(0.15)
            except Exception as e:
                print(f"  OData err: {e}")

        # Form search (correct param: menuItemFilter)
        for term in terms:
            try:
                raw = client.call_tool("form_find_menu_item", {
                    "menuItemFilter": term,
                    "topHitCount": "5"
                })
                if raw and "required" not in raw.lower():
                    sc, nm, tx, tp = score_form(term, raw)
                    if sc > f_score:
                        f_score, f_name, f_text, f_type = sc, nm, tx, tp
                time.sleep(0.15)
            except Exception as e:
                print(f"  Form err: {e}")

        # Categorize
        if o_score >= 50 and f_score >= 50:
            cat = "BOTH"
        elif o_score >= 50:
            cat = "ODATA"
        elif f_score >= 50:
            cat = "FORM"
        elif o_score >= 20 or f_score >= 20:
            cat = "UNCERTAIN"
        else:
            cat = "NOT_FOUND"

        print(f"  OData={o_score:.0f}({o_entity}) Form={f_score:.0f}({f_name}) -> {cat}")

        results.append({
            'id': aid, 'title': title,
            'area': item.get('area',''), 'state': item.get('state',''),
            'category': cat,
            'odata_score': o_score, 'odata_entity': o_entity, 'odata_label': o_label,
            'form_score': f_score, 'form_menu_item': f_name, 'form_text': f_text, 'form_type': f_type,
        })
        done_ids.add(aid)

        if len(results) % 25 == 0:
            with open(CHECKPOINT, 'w') as f:
                json.dump(results, f, indent=2)
            cats = {}
            for r in results:
                cats[r['category']] = cats.get(r['category'], 0) + 1
            print(f"  ** Checkpoint {len(results)}/{total}: {cats} **")

    with open(CHECKPOINT, 'w') as f:
        json.dump(results, f, indent=2)

    # ── Generate Excel ──────────────────────────────────────────────
    print("\nGenerating Excel report...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        os.system("pip install openpyxl")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Summary sheet ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    cats = {}
    for r in results:
        cats[r['category']] = cats.get(r['category'], 0) + 1

    ws.append(["MY/SG Configuration Drift - 606 Items Categorization"])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    cat_colors = {
        "ODATA": "C6EFCE", "FORM": "BDD7EE", "BOTH": "D9E2F3",
        "UNCERTAIN": "FFF2CC", "NOT_FOUND": "FFC7CE"
    }
    cat_desc = {
        "ODATA": "Accessible via Data Entity (OData)",
        "FORM": "Accessible via Form Tools (MCP)",
        "BOTH": "Accessible via BOTH Data Entity and Form Tools",
        "UNCERTAIN": "Partial match - needs manual review",
        "NOT_FOUND": "No match found - custom/complex config"
    }
    ws.append(["Category", "Count", "%", "Description"])
    ws[f'A3'].font = Font(bold=True)
    ws[f'B3'].font = Font(bold=True)
    ws[f'C3'].font = Font(bold=True)
    ws[f'D3'].font = Font(bold=True)
    row = 4
    for cat in ["BOTH", "ODATA", "FORM", "UNCERTAIN", "NOT_FOUND"]:
        cnt = cats.get(cat, 0)
        pct = f"{cnt*100//len(results)}%"
        ws.append([cat, cnt, pct, cat_desc.get(cat, "")])
        fill = PatternFill(start_color=cat_colors.get(cat,"FFFFFF"), fill_type="solid")
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = fill
        row += 1
    ws.append(["TOTAL", len(results), "100%", ""])
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 50

    # ── Detail sheet ────────────────────────────────────────────────
    wd = wb.create_sheet("All Items")
    headers = ["ADO ID", "Title", "Area Path", "State", "Category",
               "OData Score", "OData Entity", "OData Label",
               "Form Score", "Form Menu Item", "Form Text", "Form Type",
               "Recommendation"]
    wd.append(headers)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = wd.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for r in sorted(results, key=lambda x: (
        {"BOTH":0,"ODATA":1,"FORM":2,"UNCERTAIN":3,"NOT_FOUND":4}.get(x['category'],5),
        -x.get('odata_score',0), -x.get('form_score',0)
    )):
        rec = ""
        if r['category'] == "BOTH":
            rec = "Use OData (faster) or Form tools"
        elif r['category'] == "ODATA":
            rec = f"Use data entity: {r.get('odata_entity','')}"
        elif r['category'] == "FORM":
            rec = f"Use form: {r.get('form_menu_item','')} ({r.get('form_type','')})"
        elif r['category'] == "UNCERTAIN":
            rec = "Manual review needed - partial match"
        else:
            rec = "Cannot auto-access - manual config or custom form"

        row_data = [
            r['id'], r['title'], r.get('area',''), r.get('state',''),
            r['category'],
            r.get('odata_score',0), r.get('odata_entity',''), r.get('odata_label',''),
            r.get('form_score',0), r.get('form_menu_item',''), r.get('form_text',''), r.get('form_type',''),
            rec
        ]
        wd.append(row_data)
        row_num = wd.max_row
        fill = PatternFill(start_color=cat_colors.get(r['category'],"FFFFFF"), fill_type="solid")
        for col_idx in range(1, len(headers)+1):
            cell = wd.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.border = thin

    # Column widths
    col_widths = [10, 70, 40, 18, 12, 12, 25, 30, 12, 25, 30, 10, 50]
    for i, w in enumerate(col_widths, 1):
        wd.column_dimensions[wd.cell(row=1, column=i).column_letter].width = w

    wd.auto_filter.ref = f"A1:M{wd.max_row}"
    wd.freeze_panes = "A2"

    # ── Per-category sheets ─────────────────────────────────────────
    for cat in ["ODATA", "FORM", "BOTH", "UNCERTAIN", "NOT_FOUND"]:
        cat_items = [r for r in results if r['category'] == cat]
        if not cat_items:
            continue
        ws_cat = wb.create_sheet(f"{cat} ({len(cat_items)})")
        if cat in ("ODATA", "BOTH"):
            ws_cat.append(["ADO ID", "Title", "OData Entity", "OData Label", "OData Score", "Area Path"])
            for c in range(1,7):
                ws_cat.cell(row=1, column=c).font = hdr_font
                ws_cat.cell(row=1, column=c).fill = hdr_fill
            for r in cat_items:
                ws_cat.append([r['id'], r['title'], r.get('odata_entity',''), r.get('odata_label',''), r.get('odata_score',0), r.get('area','')])
        elif cat == "FORM":
            ws_cat.append(["ADO ID", "Title", "Form Menu Item", "Form Text", "Form Type", "Form Score", "Area Path"])
            for c in range(1,8):
                ws_cat.cell(row=1, column=c).font = hdr_font
                ws_cat.cell(row=1, column=c).fill = hdr_fill
            for r in cat_items:
                ws_cat.append([r['id'], r['title'], r.get('form_menu_item',''), r.get('form_text',''), r.get('form_type',''), r.get('form_score',0), r.get('area','')])
        else:
            ws_cat.append(["ADO ID", "Title", "OData Score", "OData Entity", "Form Score", "Form Menu Item", "Area Path"])
            for c in range(1,8):
                ws_cat.cell(row=1, column=c).font = hdr_font
                ws_cat.cell(row=1, column=c).fill = hdr_fill
            for r in cat_items:
                ws_cat.append([r['id'], r['title'], r.get('odata_score',0), r.get('odata_entity',''), r.get('form_score',0), r.get('form_menu_item',''), r.get('area','')])

        ws_cat.auto_filter.ref = f"A1:{ws_cat.cell(row=1, column=ws_cat.max_column).column_letter}{ws_cat.max_row}"
        ws_cat.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    print(f"\nExcel saved: {EXCEL_PATH}")

    # Final summary
    print(f"\n{'='*60}")
    print(f"CATEGORIZATION COMPLETE: {len(results)} items")
    print(f"{'='*60}")
    for cat in ["BOTH","ODATA","FORM","UNCERTAIN","NOT_FOUND"]:
        cnt = cats.get(cat, 0)
        print(f"  {cat:12s}: {cnt:4d} ({cnt*100//len(results)}%)")

if __name__ == "__main__":
    main()
