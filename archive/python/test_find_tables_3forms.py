"""
Test script: Find tables for 3 forms using form_find_controls approach.
Extracts table/datasource names from control name prefixes (TableName_FieldName pattern).
"""
import sys
import os
import json
import re
import time
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from d365_mcp_client import D365McpClient
from form_reader import open_form, close_form, mcp_call

CONFIG_FILE = r"C:\D365DataValidator\config.json"
OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output\Tables"
ENV_KEY = "ENV1"  # UAT

# Generic prefixes to filter out (not real table/datasource names)
SKIP_PREFIXES = {
    "Grid", "Tab", "TabPage", "Button", "MenuButton", "Group", "SystemDefined",
    "Overview", "General", "Details", "Header", "Lines", "Line", "Setup",
    "Action", "ActionPane", "ActionPaneTab", "ButtonGroup", "CommandButton",
    "Filter", "QuickFilter", "Navigation", "NavigationList", "FormControl",
    "Dialog", "DialogHeader", "DialogContent", "StatusBar", "InfoPart",
    "FactBox", "Part", "FormPart", "SysFilter", "Separator",
    "DropDialog", "DropDialogButton", "MenuItemButton", "ToggleButton",
    "Image", "StaticText", "RealEdit", "IntEdit", "StringEdit",
    "DateEdit", "TimeEdit", "CheckBox", "ComboBox", "RadioButton",
    "SegmentedEntry", "ReferenceGroup", "Dimension", "Financial",
    "OK", "Cancel", "Yes", "No", "Close", "Apply", "Save",
    "DateEffectivenessButton", "RunNow",
    "New", "Delete", "Edit", "View", "Find", "Print", "Export",
    "Refresh", "Copy", "Paste", "Cut", "Undo", "Redo",
    "InterestFee",  # common sub-form label, not a datasource
}

# 3 test forms
TEST_FORMS = [
    {
        "nav_path": "Procurement and sourcing > Setup > Purchase orders > Return action",
        "mi_name": "purchReturnActionDefaults",
        "mi_type": "Display",
    },
    {
        "nav_path": "Credit and collections > Setup > Set up interest codes",
        "mi_name": "Interest",
        "mi_type": "Display",
    },
    {
        "nav_path": "Production control > Operations > All routes",
        "mi_name": "RouteTable",
        "mi_type": "Display",
    },
]

# Search terms — a-z plus common D365 field fragments
SEARCH_TERMS = list("abcdefghijklmnopqrstuvwxyz") + [
    "0", "1", "2", "Code", "Name", "Date", "Num", "Type", "Status",
    "Amount", "Qty", "Unit", "Group", "Account", "Description",
    "Currency", "Company", "Ledger", "Dimension", "Table", "Line",
    "Purch", "Sales", "Cust", "Vend", "Item", "Route", "Oper",
    "Interest", "Fee", "Calendar", "Paym", "Return", "Prod",
]


def find_all_controls(client, extra_terms=None):
    """
    Call form_find_controls with many search terms to collect ALL control names.
    The MCP returns a JSON list of {Name, Properties} per call, max ~5 per search.
    We search broadly to capture as many controls as possible.
    """
    all_controls = {}  # name -> properties

    search_terms = list(SEARCH_TERMS)
    if extra_terms:
        search_terms.extend(extra_terms)

    for term in search_terms:
        try:
            raw = client.call_tool("form_find_controls", {"controlSearchTerm": term})
            if isinstance(raw, str):
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    for item in parsed:
                        cname = item.get("Name", "")
                        if cname and cname not in all_controls:
                            all_controls[cname] = item.get("Properties", {})
                elif isinstance(parsed, dict):
                    # Could be a single control or error
                    result_msg = parsed.get("Result", "")
                    if "required" in str(result_msg).lower():
                        continue
                    cname = parsed.get("Name", "")
                    if cname:
                        all_controls[cname] = parsed.get("Properties", {})
        except Exception:
            pass  # Some terms will fail or return empty

    return all_controls


def extract_table_prefixes(control_names):
    """
    Extract table/datasource names from control name prefixes.
    D365 names controls as {DataSourceAlias}_{FieldName}.
    Grid columns also follow this pattern.
    """
    prefixes = set()
    for name in control_names:
        if not name or "_" not in name:
            continue
        parts = name.split("_", 1)
        prefix = parts[0]

        # Skip generic prefixes
        if prefix in SKIP_PREFIXES:
            continue
        # Skip very short
        if len(prefix) < 3:
            continue
        # Skip all-lowercase generic words that are likely not datasource aliases
        generic_lower = {
            "new", "delete", "edit", "view", "find", "print", "export",
            "refresh", "copy", "paste", "cut", "undo", "redo",
            "ok", "cancel", "yes", "no", "close", "apply", "save",
            "true", "false", "none", "null",
        }
        if prefix.lower() in generic_lower:
            continue
        # Skip if ends with common button/control suffixes
        if prefix.endswith("Button") or prefix.endswith("MenuItem"):
            continue

        prefixes.add(prefix)

    return sorted(prefixes)


def _find_first_grid(form_obj, depth=0):
    """Find the first grid name in a form structure."""
    if not isinstance(form_obj, dict) or depth > 6:
        return None
    for key, val in form_obj.items():
        if key == "Grid" and isinstance(val, dict):
            for gname in val:
                return gname
        elif isinstance(val, dict) and key not in ("Rows", "Cells", "Button"):
            found = _find_first_grid(val, depth + 1)
            if found:
                return found
    return None


def _form_specific_terms(mi_name, form_name):
    """Generate extra search terms from form/MI name."""
    terms = set()
    for name in (mi_name, form_name):
        if not name:
            continue
        words = re.findall(r'[A-Z][a-z]+', name)
        terms.update(w.lower() for w in words if len(w) > 2)
    return list(terms)


def _extract_names_from_form_obj(form_obj, controls_dict, depth=0):
    """
    Recursively walk the initial FormState.Form JSON to extract
    control/column names that form_find_controls might miss.
    """
    if not isinstance(form_obj, dict) or depth > 8:
        return
    for key, val in form_obj.items():
        if not isinstance(val, dict):
            continue
        if key == "Grid":
            for gname, ginfo in val.items():
                if not isinstance(ginfo, dict):
                    continue
                # Grid columns have TableName_FieldName pattern
                for col in ginfo.get("Columns", []):
                    cname = col.get("Name", "")
                    if cname:
                        controls_dict[cname] = col
        elif key in ("Input", "Combobox", "Checkbox", "RealInput", "IntInput",
                      "DateInput", "TimeInput", "SegmentedEntry"):
            for cname, cinfo in val.items():
                if cname:
                    controls_dict[cname] = cinfo if isinstance(cinfo, dict) else {}
        elif key in ("Tab", "TabPage", "Children", "Group"):
            if isinstance(val, dict):
                _extract_names_from_form_obj(val, controls_dict, depth + 1)
        elif isinstance(val, dict) and key not in ("Rows", "Cells", "Button",
                "MenuButton", "ButtonGroup", "Pagination"):
            _extract_names_from_form_obj(val, controls_dict, depth + 1)


def extract_tables_for_form(env_config, mi_name, mi_type, nav_path, company="MY30"):
    """Open a form, find all controls, extract table prefixes."""
    client = D365McpClient(env_config)
    client.connect()

    result = {
        "nav_path": nav_path,
        "mi_name": mi_name,
        "mi_type": mi_type,
        "form_name": "",
        "form_caption": "",
        "control_count": 0,
        "all_control_names": [],
        "table_prefixes": [],
        "status": "",
    }

    try:
        print(f"  Opening {mi_name} ({mi_type}) on {company}...")
        form_result = open_form(client, mi_name, mi_type, company)
        fs = form_result.get("FormState", {})
        result["form_name"] = fs.get("Name", "")
        result["form_caption"] = fs.get("Caption", "")
        print(f"  Form: {result['form_name']} — {result['form_caption']}")

        # Try selecting first grid row to expand detail/tab controls
        form_obj = fs.get("Form", {})
        grid_name = _find_first_grid(form_obj)
        if grid_name:
            try:
                print(f"  Selecting first row in grid '{grid_name}'...")
                client.call_tool("form_select_grid_row", {
                    "gridName": grid_name, "rowNumber": "1", "marking": "Unmarked",
                })
                time.sleep(0.5)
            except Exception:
                pass

        # Also try opening tabs to expose more controls
        for tab_term in ["Earnings", "Payments", "General", "Setup", "Details",
                         "Overview", "Header", "Lines", "Dimension"]:
            try:
                client.call_tool("form_open_or_close_tab", {"tabName": tab_term})
            except Exception:
                pass

        # Collect ALL controls using broad searches
        print(f"  Finding all controls (broad search)...")
        extra = _form_specific_terms(mi_name, result["form_name"])
        all_controls = find_all_controls(client, extra)

        # Also extract control/column names from initial form structure
        _extract_names_from_form_obj(form_obj, all_controls)

        result["control_count"] = len(all_controls)
        result["all_control_names"] = sorted(all_controls.keys())
        print(f"  Found {len(all_controls)} unique controls")

        # Extract column names from grid controls (they contain TableName_FieldName)
        grid_names = set()
        for cname, props in all_controls.items():
            if isinstance(props, dict) and "Columns" in props:
                grid_names.add(cname)
                for col in props["Columns"]:
                    col_name = col.get("Name", "")
                    if col_name and col_name not in all_controls:
                        result["all_control_names"].append(col_name)

        # Extract table prefixes
        prefixes = extract_table_prefixes(result["all_control_names"])
        # Grid container names that appear as prefixes — keep only if they
        # also have DataSource_Field controls confirming them as datasources
        for gn in list(grid_names):
            if gn in prefixes:
                has_ds = any(
                    c.startswith(gn + "_") for c in result["all_control_names"]
                    if c != gn
                )
                if not has_ds:
                    prefixes.remove(gn)
        result["table_prefixes"] = prefixes
        print(f"  Tables found: {prefixes}")

        result["status"] = "OK"
        close_form(client)

    except Exception as e:
        result["status"] = f"ERROR: {str(e)[:200]}"
        print(f"  ERROR: {e}")
        try:
            close_form(client)
        except Exception:
            pass

    return result


def generate_excel(results):
    """Generate Excel with results."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Tables_test_3forms_{ts}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    er = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    def write_header(ws, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, Alignment(horizontal="center"), tb

    def auto_width(ws, headers, max_row):
        for ci in range(1, len(headers) + 1):
            mx = len(str(headers[ci-1]))
            for ri in range(2, min(max_row+1, 200)):
                v = ws.cell(row=ri, column=ci).value
                if v: mx = max(mx, min(len(str(v)), 80))
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(mx + 2, 80)

    # ── Summary ──
    ws = wb.create_sheet("Summary")
    hdrs = ["Navigation Path", "Menu Item", "Form Name", "Caption",
            "Total Controls", "Tables Found", "Table Names", "Status"]
    write_header(ws, hdrs)
    for ri, r in enumerate(results, 2):
        vals = [r["nav_path"], r["mi_name"], r["form_name"], r["form_caption"],
                r["control_count"], len(r["table_prefixes"]),
                ", ".join(r["table_prefixes"]), r["status"]]
        fill = ok if r["status"] == "OK" else er
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v); c.border = tb; c.fill = fill
    auto_width(ws, hdrs, len(results)+1)
    ws.freeze_panes = "A2"

    # ── Tables Detail ──
    ws2 = wb.create_sheet("Tables Detail")
    hdrs2 = ["Navigation Path", "Form Name", "Table/Datasource Name"]
    write_header(ws2, hdrs2)
    rn = 2
    for r in results:
        for tbl in r["table_prefixes"]:
            ws2.cell(row=rn, column=1, value=r["nav_path"]).border = tb
            ws2.cell(row=rn, column=2, value=r["form_name"]).border = tb
            ws2.cell(row=rn, column=3, value=tbl).border = tb
            rn += 1
    auto_width(ws2, hdrs2, rn)
    ws2.freeze_panes = "A2"

    # ── All Controls ──
    ws3 = wb.create_sheet("All Controls")
    hdrs3 = ["Navigation Path", "Form Name", "Control Name", "Extracted Table Prefix"]
    write_header(ws3, hdrs3)
    rn = 2
    for r in results:
        for cname in r["all_control_names"]:
            prefix = ""
            if "_" in cname:
                p = cname.split("_", 1)[0]
                if p in r["table_prefixes"]:
                    prefix = p
            ws3.cell(row=rn, column=1, value=r["nav_path"]).border = tb
            ws3.cell(row=rn, column=2, value=r["form_name"]).border = tb
            ws3.cell(row=rn, column=3, value=cname).border = tb
            ws3.cell(row=rn, column=4, value=prefix).border = tb
            rn += 1
    auto_width(ws3, hdrs3, rn)
    ws3.freeze_panes = "A2"

    wb.save(filepath)
    print(f"\nExcel saved: {filepath}")
    return filepath


def main():
    with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
    env_config = config["environments"][ENV_KEY]
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    results = []
    for idx, form in enumerate(TEST_FORMS, 1):
        print(f"\n{'='*60}")
        print(f"[{idx}/3] {form['nav_path']}")
        print(f"{'='*60}")
        r = extract_tables_for_form(
            env_config, form["mi_name"], form["mi_type"], form["nav_path"]
        )
        results.append(r)
        time.sleep(1)

    # Generate Excel
    print(f"\n{'='*60}")
    print("Generating Excel...")
    filepath = generate_excel(results)

    # Print summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for r in results:
        print(f"\n  {r['nav_path']}")
        print(f"    Form: {r['form_name']} | Controls: {r['control_count']}")
        print(f"    Tables: {r['table_prefixes']}")
        print(f"    Status: {r['status']}")


if __name__ == "__main__":
    main()
