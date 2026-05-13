"""
Sample run for ADO 36293 — Warehouse management > Setup > Warehouse > Warehouses
Reads form data from both UAT (ENV1) and Config (ENV4) for MY30, MY60, SG60, DAT.
Uses form tools with pagination (>25 rows).
Outputs an Excel file.
"""
import sys
import os
import json
import time
import traceback
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from d365_mcp_client import D365McpClient
from form_reader import (
    find_menu_item, open_form, close_form, read_form_records,
    resolve_menu_item_from_path,
)

# ── Config ────────────────────────────────────────────────────────────
CONFIG_FILE = r"C:\D365DataValidator\config.json"
OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output"

ADO_ID = 36293
TITLE = "Warehouse management > Setup > Warehouse > Warehouses"
SCREENSHOT_PATH = "Warehouse management > Setup > Warehouse > Warehouses [MY]"

ENVS = {
    # "ENV1": "Env1 UAT Asia",   # UAT MCP endpoint returning 404 — temporarily offline
    "ENV4": "Env4 Config",
}
COMPANIES = ["MY30", "MY60", "SG60", "DAT"]


def load_config():
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)


def read_form_for_env_company(env_config, company, menu_item_name, menu_item_type="Display"):
    """Connect to env, open form for company, read all records with pagination.

    Two-pass approach for forms with >25 rows:
      Pass 1: Paginate grid to collect all row keys (grid-only).
      Pass 2: Re-open form, select each row to read detail fields.
              For multi-page grids, paginate and select row-by-row per page.
    """
    from form_reader import (
        mcp_call, _collect_all_grid_rows, _extract_fields, extract_form_data,
    )

    client = D365McpClient(env_config)
    client.connect()

    try:
        print(f"  Opening form '{menu_item_name}' ({menu_item_type}) for {company}...")
        form_result = open_form(client, menu_item_name, menu_item_type, company)

        # Pass 1: collect all grid rows via pagination
        grid_name, grid_rows, last_form = _collect_all_grid_rows(client, form_result)
        total_rows = len(grid_rows)
        print(f"  Grid '{grid_name}': {total_rows} rows found")

        if total_rows == 0:
            # Try extracting as parameter/settings form (no grid)
            form_data = extract_form_data(client, form_result)
            record = {}
            for fname, finfo in form_data["fields"].items():
                record[finfo.get("label", fname)] = finfo.get("value", "")
            for tab_name, tab_info in form_data["tabs"].items():
                for fname, finfo in tab_info["fields"].items():
                    record[finfo.get("label", fname)] = finfo.get("value", "")
            close_form(client)
            if record:
                return [record], list(record.keys())
            return [], []

        if total_rows <= 25:
            # Small form — detail read works directly (no re-open needed)
            print(f"  Reading details for {total_rows} rows (single page)...")
            records = []
            columns = set()
            for r in grid_rows:
                columns.update(r.keys())

            for row_idx in range(1, total_rows + 1):
                try:
                    sel_result = mcp_call(client, "form_select_grid_row", {
                        "gridName": grid_name,
                        "rowNumber": row_idx,
                        "marking": "Unmarked",
                    })
                    record = dict(grid_rows[row_idx - 1])

                    if "raw" not in sel_result:
                        sel_form = sel_result.get("FormState", {}).get("Form", {})
                        detail_fields = {}
                        _extract_fields(sel_form, detail_fields)
                        for fname, finfo in detail_fields.items():
                            label = finfo.get("label", fname)
                            val = finfo.get("value", "")
                            if label not in record:
                                record[label] = val
                                columns.add(label)
                        # Tab fields
                        for tab_name, tab_info in sel_form.get("Tab", {}).items():
                            if not isinstance(tab_info, dict):
                                continue
                            children = tab_info.get("Children", {})
                            tab_fields = {}
                            _extract_fields(children, tab_fields)
                            for fname, finfo in tab_fields.items():
                                label = finfo.get("label", fname)
                                val = finfo.get("value", "")
                                if label not in record:
                                    record[label] = val
                                    columns.add(label)
                    records.append(record)
                    time.sleep(0.15)
                except Exception as e:
                    records.append(dict(grid_rows[row_idx - 1]))

            close_form(client)
            print(f"  Got {len(records)} records, {len(columns)} columns")
            return records, sorted(columns)

        # >25 rows — need to re-open and walk page by page
        close_form(client)
        print(f"  Re-opening form for detail read ({total_rows} rows across pages)...")

        PAGE_SIZE = 25
        all_records = []
        all_columns = set()
        for r in grid_rows:
            all_columns.update(r.keys())

        page_num = 0
        rows_processed = 0

        while rows_processed < total_rows:
            # Open fresh connection for each page to avoid session issues
            client2 = D365McpClient(env_config)
            client2.connect()
            form_result2 = open_form(client2, menu_item_name, menu_item_type, company)

            # Navigate to the correct page
            if page_num > 0:
                for p in range(page_num):
                    # Find grid name from current form
                    frm = form_result2.get("FormState", {}).get("Form", {})
                    gn = None
                    for g, gi in frm.get("Grid", {}).items():
                        if isinstance(gi, dict) and "Rows" in gi:
                            gn = g
                            break
                    if not gn:
                        break
                    next_r = mcp_call(client2, "form_click_control", {
                        "controlName": gn,
                        "actionId": "LoadNextPage",
                    })
                    if "raw" in next_r:
                        break
                    form_result2 = next_r

            # Now select rows on this page
            page_start = page_num * PAGE_SIZE
            page_end = min(page_start + PAGE_SIZE, total_rows)
            rows_on_page = page_end - page_start

            for row_idx in range(1, rows_on_page + 1):
                global_idx = page_start + row_idx - 1
                try:
                    sel_result = mcp_call(client2, "form_select_grid_row", {
                        "gridName": grid_name,
                        "rowNumber": row_idx,
                        "marking": "Unmarked",
                    })
                    record = dict(grid_rows[global_idx]) if global_idx < len(grid_rows) else {}

                    if "raw" not in sel_result:
                        sel_form = sel_result.get("FormState", {}).get("Form", {})
                        detail_fields = {}
                        _extract_fields(sel_form, detail_fields)
                        for fname, finfo in detail_fields.items():
                            label = finfo.get("label", fname)
                            val = finfo.get("value", "")
                            if label not in record:
                                record[label] = val
                                all_columns.add(label)
                        for tab_name, tab_info in sel_form.get("Tab", {}).items():
                            if not isinstance(tab_info, dict):
                                continue
                            children = tab_info.get("Children", {})
                            tab_fields = {}
                            _extract_fields(children, tab_fields)
                            for fname, finfo in tab_fields.items():
                                label = finfo.get("label", fname)
                                val = finfo.get("value", "")
                                if label not in record:
                                    record[label] = val
                                    all_columns.add(label)
                    all_records.append(record)
                    time.sleep(0.15)
                except Exception as e:
                    if global_idx < len(grid_rows):
                        all_records.append(dict(grid_rows[global_idx]))
                    else:
                        all_records.append({})

            rows_processed = page_end
            page_num += 1
            close_form(client2)
            print(f"    Page {page_num}: rows {page_start+1}-{page_end} done")

        print(f"  Got {len(all_records)} records, {len(all_columns)} columns")
        return all_records, sorted(all_columns)

    except Exception as e:
        print(f"  ERROR: {e}")
        traceback.print_exc()
        try:
            close_form(client)
        except:
            pass
        return [], []


def resolve_menu_item(env_config, company="DAT"):
    """Resolve the menu item name from the navigation path."""
    client = D365McpClient(env_config)
    client.connect()

    try:
        mi_name, mi_type, hint = resolve_menu_item_from_path(client, SCREENSHOT_PATH, company)
        print(f"  Resolved: menu_item='{mi_name}', type='{mi_type}', hint='{hint}'")
        close_form(client)
        return mi_name, mi_type
    except Exception as e:
        print(f"  Resolve error: {e}")
        return None, None


def write_excel(all_data, output_path):
    """Write all collected data to a single Excel sheet with unified columns."""
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], capture_output=True)
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    # Alternating company row colors for readability
    company_fills = {
        "MY30": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
        "MY60": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "SG60": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "DAT":  PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    }

    # ── Build superset of all columns across all companies ──
    all_columns = set()
    for info in all_data.values():
        for rec in info["records"]:
            all_columns.update(rec.keys())

    # Sort columns but put key identifiers first
    priority_cols = ["Warehouse", "Name", "Site", "Type"]
    sorted_cols = []
    for pc in priority_cols:
        if pc in all_columns:
            sorted_cols.append(pc)
            all_columns.discard(pc)
    sorted_cols.extend(sorted(all_columns))

    # Full header: Company + all data columns
    header = ["Company"] + sorted_cols

    # ── Single data sheet ──
    ws = wb.create_sheet("Warehouses Data")

    # Header row
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows — grouped by company
    row_num = 2
    for company in COMPANIES:
        for key, info in all_data.items():
            if info["company"] != company:
                continue
            fill = company_fills.get(company)
            for rec in info["records"]:
                ws.cell(row=row_num, column=1, value=company).border = thin_border
                if fill:
                    ws.cell(row=row_num, column=1).fill = fill
                for ci, col_name in enumerate(sorted_cols, 2):
                    val = rec.get(col_name, "")
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill
                row_num += 1

    # Auto-width
    for ci, col_name in enumerate(header, 1):
        max_len = len(str(col_name))
        for r in range(2, min(row_num, 52)):
            val = ws.cell(row=r, column=ci).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 2, 50)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(header))}{row_num - 1}"

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["ADO ID", ADO_ID])
    ws_sum.append(["Title", TITLE])
    ws_sum.append(["Screenshot Path", SCREENSHOT_PATH])
    ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_sum.append(["Total columns (superset)", len(sorted_cols)])
    ws_sum.append([])
    ws_sum.append(["Company", "Record Count", "Status"])
    for key, info in all_data.items():
        ws_sum.append([info["company"], len(info["records"]), info.get("status", "OK")])
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 50
    ws_sum.column_dimensions["C"].width = 20

    # Move Summary to first position
    wb.move_sheet("Summary", offset=-1)

    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")
    print(f"  Single sheet with {len(sorted_cols)} columns (superset), {row_num - 2} total rows")


def main():
    config = load_config()

    # Step 1: Resolve menu item using ENV1/DAT
    print("=" * 60)
    print(f"ADO {ADO_ID}: {TITLE}")
    print("=" * 60)
    print("\nStep 1: Resolving menu item...")
    # Use first available env for resolution
    first_env_key = list(ENVS.keys())[0]
    resolve_config = config["environments"][first_env_key]
    mi_name, mi_type = resolve_menu_item(resolve_config, "DAT")

    if not mi_name:
        # Try direct search for "Warehouses"
        print("  Trying direct search for 'Warehouses'...")
        client = D365McpClient(resolve_config)
        client.connect()
        items = find_menu_item(client, "Warehouses", "DAT")
        print(f"  Found {len(items)} candidates:")
        for item in items[:10]:
            print(f"    {item['name']} ({item['type']}): {item['text']}")
        close_form(client)

        if items:
            # Pick best match — InventLocation is the standard Warehouses form
            for item in items:
                if "inventlocation" in item["name"].lower():
                    mi_name = item["name"]
                    mi_type = item["type"]
                    break
            if not mi_name:
                mi_name = items[0]["name"]
                mi_type = items[0]["type"]
            print(f"  Selected: {mi_name} ({mi_type})")

    if not mi_name:
        print("ERROR: Could not resolve menu item!")
        return

    # Step 2: Read form data from both environments for all companies
    print(f"\nStep 2: Reading form data (menu item: {mi_name}, type: {mi_type})")
    all_data = {}

    for env_key, env_display in ENVS.items():
        env_config = config["environments"][env_key]
        for company in COMPANIES:
            key = f"{env_key}_{company}"
            print(f"\n--- {env_display} / {company} ---")

            try:
                records, columns = read_form_for_env_company(
                    env_config, company, mi_name, mi_type
                )
                all_data[key] = {
                    "env_key": env_key,
                    "env_name": env_display,
                    "env_short": "UAT" if env_key == "ENV1" else "CFG",
                    "company": company,
                    "records": records,
                    "columns": sorted(columns) if columns else [],
                    "status": f"OK ({len(records)} rows)" if records else "No data",
                }
            except Exception as e:
                print(f"  FAILED: {e}")
                all_data[key] = {
                    "env_key": env_key,
                    "env_name": env_display,
                    "env_short": "UAT" if env_key == "ENV1" else "CFG",
                    "company": company,
                    "records": [],
                    "columns": [],
                    "status": f"ERROR: {str(e)[:100]}",
                }

            time.sleep(1)  # brief pause between calls

    # Step 3: Write Excel
    print("\nStep 3: Writing Excel...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"ADO_{ADO_ID}_Warehouses_{ts}.xlsx")
    write_excel(all_data, output_path)

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for key, info in all_data.items():
        print(f"  {info['env_name']:20s} {info['company']:6s} -> {len(info['records']):4d} records  [{info['status']}]")


if __name__ == "__main__":
    main()
