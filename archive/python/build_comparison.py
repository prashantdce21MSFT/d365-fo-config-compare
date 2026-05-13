"""
Build combined comparison Excel for ADO 36764 smmParameters
across UAT + Config environments and MY30/MY60/SG60 legal entities.
Columns: Tab | Field | MY30_UAT | MY60_UAT | SG60_UAT | MY30_CONFIG | MY60_CONFIG | SG60_CONFIG
Yellow = any difference across all 6 values.
"""
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ADO_ID    = 36764
ADO_TITLE = "Sales and marketing -> Setup -> Sales and marketing parameters"
MI        = "smmParameters"

COLS = [
    ('MY30', 'UAT'), ('MY60', 'UAT'), ('SG60', 'UAT'),
    ('MY30', 'CONFIG'), ('MY60', 'CONFIG'), ('SG60', 'CONFIG'),
]

INPUT_DIR = Path("C:/D365 Configuration Drift Analysis/output")
OUT_FILE  = INPUT_DIR / "ADO_36764_smmParameters_UAT_vs_Config.xlsx"


def parse_text(path: Path) -> dict:
    tabs = {}
    current_tab = None
    for line in path.read_text(encoding='utf-8', errors='replace').splitlines():
        m = re.match(r'^=== TAB: (.+) ===$', line.strip())
        if m:
            current_tab = m.group(1).strip()
            tabs[current_tab] = []
            continue
        if current_tab and ':' in line:
            parts = line.strip().split(':', 1)
            field = parts[0].strip()
            value = parts[1].strip() if len(parts) > 1 else ''
            if field:
                tabs[current_tab].append((field, value))
    return tabs


def main():
    data = {}
    for le, env in COLS:
        path = INPUT_DIR / f"smmParameters_{le}_{env}.xlsx"
        data[(le, env)] = parse_text(path)
        print(f"Loaded {le} {env}: {len(data[(le,env)])} tabs")

    all_tabs, seen_tabs = [], set()
    for le, env in COLS:
        for tab in data[(le, env)]:
            if tab not in seen_tabs:
                all_tabs.append(tab)
                seen_tabs.add(tab)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UAT vs Config"

    H_FILL  = PatternFill("solid", fgColor="1F4E79")
    H_FONT  = Font(bold=True, color="FFFFFF", size=11)
    T_FILL  = PatternFill("solid", fgColor="2E75B6")
    T_FONT  = Font(bold=True, color="FFFFFF")
    F_FILL  = PatternFill("solid", fgColor="D6E4F0")
    D_FILL  = PatternFill("solid", fgColor="FFE699")
    U_FILL  = PatternFill("solid", fgColor="E2EFDA")
    C_FILL  = PatternFill("solid", fgColor="FCE4D6")
    WRAP    = Alignment(wrap_text=True, vertical='top')

    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value=f"ADO ID: {ADO_ID}  |  {ADO_TITLE}  |  MI: {MI}")
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = H_FILL
    c.alignment = WRAP

    headers = ['Tab', 'Field'] + [f"{le} {env}" for le, env in COLS]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    row = 3
    for tab in all_tabs:
        all_fields, seen_fields = [], set()
        for le, env in COLS:
            for field, _ in data[(le, env)].get(tab, []):
                if field not in seen_fields:
                    all_fields.append(field)
                    seen_fields.add(field)
        if not all_fields:
            continue

        ws.merge_cells(f'A{row}:H{row}')
        tc = ws.cell(row=row, column=1, value=f"  {tab}")
        tc.fill = T_FILL
        tc.font = T_FONT
        row += 1

        for field in all_fields:
            vals = {(le, env): dict(data[(le, env)].get(tab, [])).get(field, '') for le, env in COLS}
            has_diff = len(set(v for v in vals.values() if v)) > 1

            ws.cell(row=row, column=1, value=tab).fill = F_FILL
            ws.cell(row=row, column=2, value=field).fill = F_FILL

            for col_idx, (le, env) in enumerate(COLS, 3):
                cell = ws.cell(row=row, column=col_idx, value=vals[(le, env)])
                cell.alignment = WRAP
                cell.fill = D_FILL if has_diff else (U_FILL if env == 'UAT' else C_FILL)

            row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    for i in range(3, 9):
        ws.column_dimensions[get_column_letter(i)].width = 28
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'C3'

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")
    print(f"Tabs: {len(all_tabs)}, Field rows: {row - 3 - len(all_tabs)}")


if __name__ == '__main__':
    main()
