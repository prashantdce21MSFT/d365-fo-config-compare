"""
Build final 2-sheet Excel for ADO 36764 smmParameters.

Source: DeepCompare_smmParameters_20260501_170853.xlsx
  - "All Controls" sheet  → 743 rows: #, Control, Label, Value, Type, Tab, Env, LE
  - "Number sequences_Grid" → grid rows: #, Reference, Num seq code, Desc, Env, LE

Output:
  C:/D365 Configuration Drift Analysis/output/ADO_36764_smmParameters_FinalReport.xlsx

  Sheet 1 "Summary (Differences)"  — only rows where UAT ≠ Config for any LE
  Sheet 2 "All Values"             — every field + grid row

  Layout (rows):  Tab | Field | Legal Entity | UAT Value | Config Value | Match
  One row per (Field, Legal Entity) — so each field appears 3× (MY30, MY60, SG60).
"""
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SRC = Path("C:/D365 Configuration Drift Analysis/output/Comparison"
           "/DeepCompare_smmParameters_20260501_170853.xlsx")
OUT = Path("C:/D365 Configuration Drift Analysis/output"
           "/ADO_36764_smmParameters_FinalReport.xlsx")

LES  = ['MY30', 'MY60', 'SG60']
ENVS = ['UAT', 'Config']

TAB_ORDER = ['General', 'Prospects', 'Documents', 'Campaigns',
             'Activities', 'Transaction log', 'Acme Asia', 'Number sequences']

# Override labels for controls whose D365 label is generic (e.g. lookup display fields)
LABEL_OVERRIDES = {
    'smmParametersTable_CBAAttributeRecId_Name': 'Attribute for Alcohol content',
}

# Fields missing from DeepCompare source — values captured manually via Playwright
# Format: (tab, control_key, label) -> {LE_ENV: value}
MANUAL_FIELDS = [
    (
        'General',
        '_manual_Period_PeriodFrom',
        'Start date',
        {'MY30_UAT': '', 'MY60_UAT': '', 'SG60_UAT': '',
         'MY30_Config': '', 'MY60_Config': '', 'SG60_Config': ''},
    ),
    (
        'General',
        '_manual_Period_PeriodTo',
        'End date',
        {'MY30_UAT': '', 'MY60_UAT': '', 'SG60_UAT': '',
         'MY30_Config': '', 'MY60_Config': '', 'SG60_Config': ''},
    ),
    (
        'Prospects',
        '_manual_OpenFromTime',
        'Opening time',
        {'MY30_UAT': '12:00:00 AM', 'MY60_UAT': '12:00:00 AM', 'SG60_UAT': '08:00:00 AM',
         'MY30_Config': '12:00:00 AM', 'MY60_Config': '12:00:00 AM', 'SG60_Config': '08:00:00 AM'},
    ),
    (
        'Prospects',
        '_manual_OpenToTime',
        'Closing time',
        {'MY30_UAT': '12:00:00 AM', 'MY60_UAT': '12:00:00 AM', 'SG60_UAT': '04:00:00 PM',
         'MY30_Config': '12:00:00 AM', 'MY60_Config': '12:00:00 AM', 'SG60_Config': '04:00:00 PM'},
    ),
    (
        'Documents',
        '_manual_DocuAgeLimitForDisplay',
        'Document age limit',
        {'MY30_UAT': '365', 'MY60_UAT': '365', 'SG60_UAT': '365',
         'MY30_Config': '365', 'MY60_Config': '365', 'SG60_Config': '365'},
    ),
    (
        'Campaigns',
        '_manual_CampaignNumOfDaysExpiryDate',
        'Days campaign expires',
        {'MY30_UAT': '21', 'MY60_UAT': '21', 'SG60_UAT': '21',
         'MY30_Config': '21', 'MY60_Config': '21', 'SG60_Config': '21'},
    ),
    (
        'Campaigns',
        '_manual_NumOfCampaignDaysFollowUpDate',
        'Days before follow-up',
        {'MY30_UAT': '3', 'MY60_UAT': '3', 'SG60_UAT': '3',
         'MY30_Config': '3', 'MY60_Config': '3', 'SG60_Config': '3'},
    ),
    (
        'Transaction log',
        '_manual_GroupOption_TransLogInterval',
        'Interval',
        {'MY30_UAT': '365', 'MY60_UAT': '0', 'SG60_UAT': '0',
         'MY30_Config': '365', 'MY60_Config': '0', 'SG60_Config': '0'},
    ),
    (
        'Acme Asia',
        '_manual_CBASalesOrderGroup',
        'Sales order group',
        {'MY30_UAT': '', 'MY60_UAT': '', 'SG60_UAT': '',
         'MY30_Config': '', 'MY60_Config': '', 'SG60_Config': ''},
    ),
    (
        'Acme Asia',
        '_manual_CBAReleaseBeforeShipDate',
        'Release before ship date',
        {'MY30_UAT': '2', 'MY60_UAT': '2', 'SG60_UAT': '0',
         'MY30_Config': '2', 'MY60_Config': '2', 'SG60_Config': '0'},
    ),
    (
        'Acme Asia',
        '_manual_CBAEastMalaysiaDisplayDim',
        'East Malaysia display dimensions',
        {'MY30_UAT': '', 'MY60_UAT': 'L;M;S;S2;T', 'SG60_UAT': '',
         'MY30_Config': '', 'MY60_Config': '', 'SG60_Config': ''},
    ),
    (
        'Acme Asia',
        '_manual_CBAFMandatoryGeographyCopy1',
        'Mandatory geography',
        {'MY30_UAT': 'No', 'MY60_UAT': 'No', 'SG60_UAT': 'No',
         'MY30_Config': 'No', 'MY60_Config': 'No', 'SG60_Config': 'No'},
    ),
    (
        'Acme Asia',
        '_manual_CBAFMandatoryChannelCopy1',
        'Mandatory channel',
        {'MY30_UAT': 'No', 'MY60_UAT': 'No', 'SG60_UAT': 'No',
         'MY30_Config': 'No', 'MY60_Config': 'No', 'SG60_Config': 'No'},
    ),
]


# ── read "All Controls" ──────────────────────────────────────────────────────

def read_all_controls(ws) -> list[dict]:
    """
    Returns flat list of rows: one per (Tab, Control, LE).
    Columns: tab, label, le, uat_value, config_value, has_diff
    """
    # (tab, control) -> {label, MY30_UAT, MY30_Config, MY60_UAT, ...}
    pivot = {}
    tab_ctrl_order = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        num, control, label, value, ctype, tab, env, le = (list(row) + [None]*8)[:8]
        if not control or not tab or tab == '(Sweep)':
            continue
        le  = str(le  or '').strip()
        env = str(env or '').strip()
        if le not in LES or env not in ENVS:
            continue

        key = (tab, str(control))
        if key not in pivot:
            lbl = LABEL_OVERRIDES.get(str(control), str(label or control))
            pivot[key] = {'tab': tab, 'label': lbl}
            tab_ctrl_order[tab].append(key)
        pivot[key][f"{le}_{env}"] = str(value) if value is not None else ''

    # Inject manually-captured fields
    for tab, ctrl_key, label, values in MANUAL_FIELDS:
        key = (tab, ctrl_key)
        if key not in pivot:
            pivot[key] = {'tab': tab, 'label': label, **values}
            tab_ctrl_order[tab].append(key)
    rows = []
    for tab in TAB_ORDER + [t for t in tab_ctrl_order if t not in TAB_ORDER]:
        for key in tab_ctrl_order.get(tab, []):
            d = pivot[key]
            for le in LES:
                uat_val = d.get(f'{le}_UAT', '')
                cfg_val = d.get(f'{le}_Config', '')
                rows.append({
                    'tab':    tab,
                    'label':  d['label'],
                    'desc':   '',
                    'le':     le,
                    'uat':    uat_val if uat_val not in ('None',) else '',
                    'cfg':    cfg_val if cfg_val not in ('None',) else '',
                })
    return rows


# ── read Number sequences grid ────────────────────────────────────────────────

def read_grid(ws) -> list[dict]:
    """
    Grid: #, Reference, Number sequence code, Description, Env, LE
    Returns flat rows: one per (Reference, LE) with uat/config codes.
    """
    pivot = {}   # (le, ref) -> {uat_code, cfg_code, desc}
    order = []
    current_le = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        num, ref, code, desc, env, le = (list(row) + [None]*6)[:6]
        if isinstance(num, str) and num in LES:
            current_le = num
            continue
        if num is None or ref is None:
            continue
        le_used  = str(le or current_le or '').strip()
        env_used = str(env or '').strip()
        ref_str  = str(ref).strip()
        code_str = str(code or '').strip()
        desc_str = str(desc or '').strip()

        k = (le_used, ref_str)
        if k not in pivot:
            pivot[k] = {'desc': desc_str, 'uat_code': '', 'cfg_code': ''}
            order.append(k)
        if env_used == 'UAT':
            pivot[k]['uat_code'] = code_str
        else:
            pivot[k]['cfg_code'] = code_str

    # unique refs in order
    ref_order = list(dict.fromkeys(ref for _, ref in order))

    rows = []
    for ref in ref_order:
        for le in LES:
            d = pivot.get((le, ref), {})
            desc = d.get('desc', '')
            rows.append({
                'tab':   'Number sequences',
                'label': ref,
                'desc':  desc,
                'le':    le,
                'uat':   d.get('uat_code', ''),
                'cfg':   d.get('cfg_code', ''),
            })
    return rows


# ── styles ───────────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
TAB_FILL  = PatternFill("solid", fgColor="2E75B6")
TAB_FONT  = Font(bold=True, color="FFFFFF")
MY30_FILL = PatternFill("solid", fgColor="D9E1F2")
MY60_FILL = PatternFill("solid", fgColor="E2EFDA")
SG60_FILL = PatternFill("solid", fgColor="FCE4D6")
DIFF_FILL = PatternFill("solid", fgColor="FFE699")
MATCH_FILL= PatternFill("solid", fgColor="FFFFFF")
WRAP      = Alignment(wrap_text=True, vertical='top')
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)

LE_FILLS  = {'MY30': MY30_FILL, 'MY60': MY60_FILL, 'SG60': SG60_FILL}


def write_sheet(ws, all_rows: list[dict], title_text: str):
    # 7 columns: Tab | Field/Label | Description | Legal Entity | UAT Value | Config Value | Match
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = HDR_FILL
    c.alignment = CENTER

    headers = ['Tab', 'Field / Label', 'Description', 'Legal Entity', 'UAT Value', 'Config Value', 'Match']
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER

    row = 3
    current_tab   = None
    current_label = None

    for d in all_rows:
        tab   = d['tab']
        label = d['label']
        desc  = d.get('desc', '')
        le    = d['le']
        uat   = d['uat']
        cfg   = d['cfg']

        if tab != current_tab:
            ws.merge_cells(f'A{row}:G{row}')
            tc = ws.cell(row=row, column=1, value=f"  {tab}")
            tc.fill = TAB_FILL
            tc.font = TAB_FONT
            tc.alignment = WRAP
            row += 1
            current_tab   = tab
            current_label = None

        has_diff = bool(uat or cfg) and uat != cfg
        fill     = DIFF_FILL if has_diff else MATCH_FILL
        le_fill  = LE_FILLS.get(le, MATCH_FILL)
        status   = 'DIFF' if has_diff else ('Match' if (uat or cfg) else '-')

        show_label = label != current_label

        tab_cell = ws.cell(row=row, column=1, value=tab if show_label else '')
        tab_cell.fill = le_fill

        lbl_cell = ws.cell(row=row, column=2, value=label if show_label else '')
        lbl_cell.fill = le_fill
        lbl_cell.alignment = WRAP

        desc_cell = ws.cell(row=row, column=3, value=desc if show_label else '')
        desc_cell.fill = le_fill
        desc_cell.alignment = WRAP

        ws.cell(row=row, column=4, value=le).fill = le_fill
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=4).font = Font(bold=True)

        uat_c = ws.cell(row=row, column=5, value=uat)
        uat_c.fill = fill
        uat_c.alignment = WRAP

        cfg_c = ws.cell(row=row, column=6, value=cfg)
        cfg_c.fill = fill
        cfg_c.alignment = WRAP

        st = ws.cell(row=row, column=7, value=status)
        st.fill  = DIFF_FILL if has_diff else le_fill
        st.font  = Font(bold=True, color="C00000" if has_diff else "375623")
        st.alignment = CENTER

        current_label = label
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 10
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'E3'
    return row - 3


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    src_wb = openpyxl.load_workbook(SRC)

    ctrl_rows = read_all_controls(src_wb['All Controls'])
    print(f"All Controls: {len(ctrl_rows)} rows (fields × LEs)")

    grid_rows = read_grid(src_wb['Number sequences_Grid'])
    print(f"Number sequences grid: {len(grid_rows)} rows")

    all_rows  = ctrl_rows + grid_rows
    diff_rows = [r for r in all_rows
                 if (r['uat'] or r['cfg']) and r['uat'] != r['cfg']]
    print(f"\nTotal rows: {len(all_rows)}")
    print(f"Differences (UAT != Config): {len(diff_rows)}")

    wb  = openpyxl.Workbook()
    ts  = datetime.now().strftime('%Y-%m-%d %H:%M')

    ws1 = wb.active
    ws1.title = "Summary (Differences)"
    n1 = write_sheet(ws1, diff_rows,
        f"ADO 36764 | smmParameters | UAT vs Config - DIFFERENCES ONLY | {ts}")
    print(f"\nSummary sheet: {n1} rows")

    ws2 = wb.create_sheet("All Values")
    n2 = write_sheet(ws2, all_rows,
        f"ADO 36764 | smmParameters | ALL VALUES - UAT vs Config | {ts}")
    print(f"All Values sheet: {n2} rows")

    wb.save(OUT)
    print(f"\nSaved: {OUT}")


if __name__ == '__main__':
    main()
