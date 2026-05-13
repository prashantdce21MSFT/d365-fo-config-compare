"""
find_correct_mi.py
==================
Connects to Chrome via CDP, searches D365 F&O nav for each item by ADO Title,
captures the correct MI Name from the resulting URL, then writes
Correct_MI / Correct_URL / Search_Status columns to the Excel.

SETUP (one-time, before running this script):
  1. Close ALL Chrome windows completely
  2. Open a cmd and run:
       "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\temp\cdp_profile"
  3. In that Chrome window, navigate to D365 UAT and sign in with Acme.asia account
  4. Once signed in and on the D365 home page, run:
       python find_correct_mi.py

Results are saved after every item -- safe to Ctrl+C and resume later.
"""

import json, re, sys, time
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

CDP_URL      = "http://localhost:9222"
HOME_URL     = "https://example-uat.sandbox.operations.dynamics.com/?cmp=my30"
EXCEL_FILE   = r"C:/D365 Configuration Drift Analysis/output/CDD_FormMapping.xlsx"
RESULTS_JSON = r"C:/D365 Configuration Drift Analysis/output/correct_mi_results_ALL.json"
HEADER_ROW   = 11
BASE_URL_TPL = "https://example-uat.sandbox.operations.dynamics.com/?cmp=my30&mi={mi}"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")


def load_items():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    items = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        try:
            ado_id = int(raw) if raw else None
        except (ValueError, TypeError):
            continue
        if not ado_id:
            continue
        ado_title = str(ws.cell(row=row, column=2).value or '').strip()
        items.append({'ado_id': ado_id, 'ado_title': ado_title, 'excel_row': row})
    return items


def build_search_term(ado_title: str) -> str:
    parts = [p.strip() for p in re.split(r'[>\\-]+', ado_title) if p.strip()]
    return parts[-1] if parts else ado_title


def search_d365(page, search_term: str, ado_title: str) -> dict:
    empty = {'mi': '', 'url': '', 'found_title': '', 'status': 'NOT_FOUND'}
    try:
        page.keyboard.press('Escape')
        page.wait_for_timeout(400)

        # Click Search button by aria-label
        page.locator('button[aria-label="Search"]').click(timeout=10000)
        page.wait_for_timeout(800)

        # Fill search input — try multiple selectors
        search_input = None
        for sel in ['#NavigationSearchBox_searchBoxInput_input',
                    'input[id*="search" i]',
                    'input[placeholder*="earch" i]']:
            loc = page.locator(sel).first
            try:
                loc.wait_for(timeout=4000)
                search_input = loc
                break
            except Exception:
                continue

        if search_input is None:
            return {**empty, 'status': 'NO_SEARCH_BOX'}

        # Clear first, wait for dropdown to reset, then type
        search_input.fill('')
        page.wait_for_timeout(500)
        search_input.fill(search_term)

        # Wait until dropdown results contain the search term (not stale results)
        deadline = time.time() + 5
        while time.time() < deadline:
            page.wait_for_timeout(500)
            check = page.evaluate(f"""() => {{
                const items = Array.from(document.querySelectorAll('li, [role="option"]'));
                return items.some(el => {{
                    const r = el.getBoundingClientRect();
                    return r.width > 0 && r.height > 0 && el.innerText.toLowerCase().includes('{search_term[:20].lower()}');
                }});
            }}""")
            if check:
                break

        # Read all visible dropdown results with their DOM index
        result_data = page.evaluate("""() => {
            const results = [];
            document.querySelectorAll('li, [role="option"]').forEach((el, idx) => {
                const rect = el.getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) return;
                const txt = el.innerText.trim();
                if (txt && txt.length > 2 && txt.length < 200) {
                    results.push({text: txt, index: idx});
                }
            });
            return results;
        }""")

        if not result_data:
            return {**empty, 'status': 'NO_RESULTS'}

        # Score each result against full ADO title
        ado_words = set(ado_title.lower().replace('>', ' ').replace('-', ' ').split())
        best = None
        best_score = -1
        for item in result_data:
            words = set(item['text'].lower().split())
            score = len(ado_words & words)
            if score > best_score:
                best_score = score
                best = item

        if best and best_score > 0:
            # Click via DOM — no pixel coordinates
            page.evaluate(f"""() => {{
                const all = Array.from(document.querySelectorAll('li, [role="option"]'));
                const el = all[{best['index']}];
                if (el) el.click();
            }}""")
        else:
            # Click first visible result via DOM
            page.evaluate("""() => {
                const all = Array.from(document.querySelectorAll('li, [role="option"]'));
                const visible = all.find(el => {
                    const r = el.getBoundingClientRect();
                    return r.width > 0 && r.height > 0;
                });
                if (visible) visible.click();
            }""")

        page.wait_for_load_state('domcontentloaded', timeout=15000)
        page.wait_for_timeout(2000)

        current_url = page.url
        if 'login.microsoftonline' in current_url:
            return {**empty, 'status': 'SESSION_EXPIRED'}

        mi_match = re.search(r'[?&]mi=([^&\s]+)', current_url)
        if not mi_match:
            return {**empty, 'status': 'NO_MI_IN_URL', 'url': current_url}

        mi = mi_match.group(1)
        return {
            'mi': mi,
            'url': BASE_URL_TPL.format(mi=mi),
            'found_title': best['text'] if best else '',
            'status': 'FOUND'
        }

    except PWTimeout:
        return {**empty, 'status': 'TIMEOUT'}
    except Exception as e:
        return {**empty, 'status': f'ERROR: {str(e)[:80]}'}


def save_excel(items, results):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    def find_or_create_col(name):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=HEADER_ROW, column=col).value == name:
                return col
        new_col = ws.max_column + 1
        ws.cell(row=HEADER_ROW, column=new_col, value=name).font = Font(bold=True)
        return new_col

    mi_col     = find_or_create_col('Correct_MI')
    url_col    = find_or_create_col('Correct_URL')
    status_col = find_or_create_col('Search_Status')

    for item in items:
        ado_id    = item['ado_id']
        excel_row = item['excel_row']
        if ado_id not in results:
            continue

        r = results[ado_id]
        ws.cell(row=excel_row, column=mi_col,    value=r.get('mi', ''))
        ws.cell(row=excel_row, column=url_col,   value=r.get('url', ''))
        ws.cell(row=excel_row, column=status_col, value=r.get('status', ''))

        s  = r.get('status', '')
        sc = ws.cell(row=excel_row, column=status_col)
        if s == 'FOUND':
            sc.fill = GREEN
        elif s in ('TIMEOUT', 'SESSION_EXPIRED', 'NO_SEARCH_BOX') or 'ERROR' in s:
            sc.fill = RED
        else:
            sc.fill = YELLOW

    for col in [mi_col, url_col, status_col]:
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or ''))
             for r in range(HEADER_ROW, ws.max_row + 1)), default=10
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)

    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")


def main():
    items = load_items()
    total = len(items)
    print(f"Loaded {total} items from Excel")

    results_path = Path(RESULTS_JSON)
    results = {}
    if results_path.exists():
        raw = json.loads(results_path.read_text(encoding='utf-8'))
        results = {int(k): v for k, v in raw.items()}
        done = len(results)
        print(f"Resuming: {done} already done, {total - done} remaining")

    with sync_playwright() as pw:
        print(f"\nConnecting to Chrome via CDP at {CDP_URL} ...")
        try:
            browser = pw.chromium.connect_over_cdp(CDP_URL)
        except Exception as e:
            print(f"\nERROR: Cannot connect to Chrome CDP:\n  {e}")
            sys.exit(1)

        ctx  = browser.contexts[0]
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        page.set_viewport_size({'width': 1280, 'height': 720})

        print("Connected! Verifying D365 session ...")
        page.bring_to_front()
        page.wait_for_timeout(2000)

        if 'login.microsoftonline' in page.url:
            print("ERROR: Not signed in. Sign in to D365 in Chrome first, then re-run.")
            sys.exit(1)

        # Navigate home only if not already on D365
        if 'operations.dynamics.com' not in page.url:
            page.goto(HOME_URL, wait_until='domcontentloaded', timeout=30000)
            page.wait_for_timeout(2000)

        print("Session OK. Starting search loop ...\n")
        # Wait for D365 to fully load before starting
        page.wait_for_selector('button[aria-label="Search"]', timeout=15000)
        page.wait_for_timeout(2000)

        consecutive_failures = 0

        for i, item in enumerate(items):
            ado_id      = item['ado_id']
            ado_title   = item['ado_title']
            search_term = build_search_term(ado_title)

            if ado_id in results:
                print(f"  [{i+1:3}/{total}] SKIP  | {search_term[:50]}")
                continue

            result = search_d365(page, search_term, ado_title)
            results[ado_id] = result

            results_path.write_text(
                json.dumps({str(k): v for k, v in results.items()}, ensure_ascii=False, indent=2),
                encoding='utf-8'
            )

            status = result['status']
            mi     = result.get('mi', '')
            marker = 'OK' if status == 'FOUND' else 'XX'
            print(f"  [{i+1:3}/{total}] {marker} {status:18} | {search_term:50} | {mi}")

            if status == 'SESSION_EXPIRED':
                print("\nSession expired. Sign back in to Chrome and re-run.")
                break

            if status == 'FOUND':
                consecutive_failures = 0
            else:
                consecutive_failures += 1
                if consecutive_failures >= 5:
                    print("  >> 5 consecutive failures — resetting to D365 home ...")
                    page.goto(HOME_URL, wait_until='domcontentloaded', timeout=20000)
                    page.wait_for_timeout(2000)
                    consecutive_failures = 0

        browser.close()

    found     = sum(1 for v in results.values() if v['status'] == 'FOUND')
    not_found = len(results) - found
    print(f"\nResults: {found} FOUND | {not_found} not found | {total} total")

    save_excel(items, results)


if __name__ == '__main__':
    main()
