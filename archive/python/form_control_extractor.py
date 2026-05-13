"""
D365 F&O Form Control Extractor
================================
Opens a D365 form via MCP, walks every control (grids with pagination, fields,
tabs, sections, buttons) and extracts all values.  Repeats for each legal entity.

Usage:
    python form_control_extractor.py
        --path "Inventory Management -> Setup -> Quality Control -> Test Instrument"
        --le MY30 MY60 SG60
        --env ENV1

    python form_control_extractor.py
        --mi Interest --mi-type Display
        --le MY30 MY60
        --env ENV1

    # Interactive mode (no args)
    python form_control_extractor.py
"""
import sys
import os
import json
import time
import argparse
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from d365_mcp_client import D365McpClient
from form_reader import (
    open_form, close_form, mcp_call,
    _collect_all_grid_rows, _extract_fields,
    resolve_menu_item_from_path, find_menu_item,
)

CONFIG_FILE = r"C:\D365DataValidator\config.json"
MI_CACHE_FILE = r"C:\D365 Configuration Drift Analysis\output\batch_extract_mi_cache.json"

# Broad search terms for form_find_controls sweep
FIND_CONTROLS_TERMS = list("abcdefghijklmnopqrstuvwxyz") + [
    "0", "1", "2", "3", "4", "5", "6", "7", "8", "9",
    "Code", "Name", "Date", "Num", "Type", "Status",
    "Amount", "Qty", "Unit", "Group", "Account", "Description",
    "Currency", "Company", "Ledger", "Dimension", "Table", "Line",
    "Purch", "Sales", "Cust", "Vend", "Item", "Route", "Oper",
    "Interest", "Fee", "Calendar", "Paym", "Rate", "Percent",
    "Invoice", "Calculate", "Grace", "Period", "Range", "Method",
    "Earnings", "Payments", "Note", "Voucher", "Post", "Charge",
    "Min", "Max", "Day", "Month", "From", "To", "Base", "Value",
    "Debit", "Credit", "Text", "Language", "Version", "Valid",
    "Warehouse", "Site", "Config", "Color", "Size", "Style",
    "Batch", "Serial", "Location", "Inventory", "Order", "Trans",
    "Price", "Cost", "Discount", "Tax", "Charge", "Weight",
    "Address", "Contact", "Phone", "Email", "Country", "State",
    "Prod", "BOM", "Resource", "Capacity", "Schedule", "Plan",
    "Budget", "Forecast", "Actual", "Variance", "Report",
    "Default", "Setup", "Parameter", "General", "Header",
    "Delivery", "Shipping", "Freight", "Carrier", "Mode",
    "Payment", "Terms", "Due", "Cash", "Bank", "Check",
    "Return", "Reason", "Action", "Disposition", "Replace",
]


class FormControlExtractor:
    """
    Extracts all control values from a D365 F&O form via MCP.

    Usage:
        extractor = FormControlExtractor(config_path="...", env_key="ENV1")
        result = extractor.extract(
            path_or_mi="Credit and collections > Setup > Set up interest codes",
            legal_entities=["MY30", "MY60"],
        )
    """

    def __init__(self, config_path=CONFIG_FILE, env_key="ENV1"):
        with open(config_path) as f:
            config = json.load(f)
        self.env_config = config["environments"][env_key]
        self.env_key = env_key
        self.env_name = self.env_config.get("name", env_key)
        self.mi_cache = self._load_mi_cache()

    def _load_mi_cache(self):
        if os.path.exists(MI_CACHE_FILE):
            with open(MI_CACHE_FILE) as f:
                return json.load(f)
        return {}

    # ── Public API ──────────────────────────────────────────────────────

    def extract(self, path_or_mi, legal_entities, mi_type="Display"):
        """
        Extract all form control values for given legal entities.

        Args:
            path_or_mi: Navigation path OR direct menu item name
            legal_entities: List of company IDs (e.g., ["MY30", "MY60"])
            mi_type: Menu item type — "Display", "Action", or "Output"

        Returns:
            dict with form info + per-LE control data
        """
        # Resolve menu item
        mi_name, resolved_type = self._resolve_mi(path_or_mi, mi_type, legal_entities[0])
        if resolved_type:
            mi_type = resolved_type

        result = {
            "form_name": "",
            "form_caption": "",
            "menu_item": mi_name,
            "mi_type": mi_type,
            "navigation_path": path_or_mi,
            "environment": self.env_name,
            "extracted_at": datetime.now().isoformat(),
            "legal_entities": {},
        }

        for le in legal_entities:
            print(f"\n{'-'*60}")
            print(f"  Legal Entity: {le}")
            print(f"{'-'*60}")
            le_data = self._extract_for_le(mi_name, mi_type, le)
            result["legal_entities"][le] = le_data

            # Set form-level info from first successful LE
            if not result["form_name"] and le_data.get("form_name"):
                result["form_name"] = le_data["form_name"]
                result["form_caption"] = le_data["form_caption"]

        return result

    # ── Menu Item Resolution ────────────────────────────────────────────

    def _resolve_mi(self, path_or_mi, mi_type, company):
        """Resolve a navigation path or MI name to (mi_name, mi_type)."""
        # Normalize separators
        normalized = path_or_mi.replace("->", ">").replace(" > ", ">").replace("> ", ">").replace(" >", ">")

        # Check MI cache (try various key formats)
        for key_fmt in [path_or_mi, normalized, normalized.replace(">", "->")]:
            info = self.mi_cache.get(key_fmt)
            if info:
                print(f"  MI cache hit: {info['mi_name']} ({info['mi_type']})")
                return info["mi_name"], info["mi_type"]

        # If no separator, assume it's a direct MI name
        if ">" not in path_or_mi and "->" not in path_or_mi:
            print(f"  Using direct MI name: {path_or_mi}")
            return path_or_mi, mi_type

        # Resolve via MCP
        print(f"  Resolving menu item from path...")
        try:
            client = D365McpClient(self.env_config)
            client.connect()
            mi_name, resolved_type, hint = resolve_menu_item_from_path(
                client, path_or_mi, company
            )
            close_form(client)
            if mi_name:
                print(f"  Resolved: {mi_name} ({resolved_type})")
                return mi_name, resolved_type
        except Exception as e:
            print(f"  Resolution failed: {e}")
            try:
                close_form(client)
            except Exception:
                pass

        # Fallback: use last segment as MI name
        segments = normalized.split(">")
        fallback = segments[-1].strip()
        print(f"  Fallback MI name: {fallback}")
        return fallback, mi_type

    # ── Per-LE Extraction ───────────────────────────────────────────────

    def _extract_for_le(self, mi_name, mi_type, company):
        """Extract all controls for one legal entity."""
        data = {
            "company": company,
            "form_name": "",
            "form_caption": "",
            "status": "",
            "grids": {},
            "fields": {},
            "tabs": {},
            "buttons": {},
            "find_controls": {},
            "summary": {},
        }

        # Connect with retry for transient network issues
        client = D365McpClient(self.env_config)
        for attempt in range(3):
            try:
                client.connect()
                break
            except Exception as e:
                if attempt < 2:
                    wait = 10 * (attempt + 1)
                    print(f"  Connection failed (attempt {attempt+1}/3), retrying in {wait}s...")
                    time.sleep(wait)
                else:
                    raise

        try:
            # 1. Open form
            print(f"  Opening {mi_name} ({mi_type}) on {company}...")
            form_result = open_form(client, mi_name, mi_type, company)
            fs = form_result.get("FormState", {})
            data["form_name"] = fs.get("Name", "")
            data["form_caption"] = fs.get("Caption", "")
            form_obj = fs.get("Form", {})
            print(f"  Form: {data['form_name']} - {data['form_caption']}")

            # 2. Parse initial FormState recursively
            print(f"  Parsing form structure...")
            self._parse_form_obj(form_obj, data)

            # 3. Grid pagination — collect all rows for each grid
            print(f"  Collecting grid rows (with pagination)...")
            for gname, ginfo in data["grids"].items():
                if ginfo["pagination_has_next"] or len(ginfo["rows"]) == 25:
                    print(f"    Paginating grid '{gname}'...")
                    all_rows = self._paginate_grid(client, gname, ginfo["rows"], form_obj)
                    ginfo["rows"] = all_rows
                    ginfo["total_rows"] = len(all_rows)

            # 4. Select each grid row to get detail/tab fields
            print(f"  Extracting per-row detail fields...")
            for gname, ginfo in data["grids"].items():
                if ginfo["total_rows"] > 0:
                    row_details = self._extract_row_details(
                        client, mi_name, mi_type, company, gname, ginfo
                    )
                    ginfo["row_details"] = row_details

            # 4b. OData enrichment for CustTable (Acme Asia FastTab fields)
            #     form_find_controls returns stale values for FastTab fields;
            #     OData CustomersV3 gives correct per-customer CBA/CHB values
            if mi_name.lower() in ("custtable", "cust table"):
                for gname, ginfo in data["grids"].items():
                    if ginfo["total_rows"] > 0:
                        print(f"  Enriching {gname} with OData CBA/CHB fields...")
                        ginfo["row_details"] = self._enrich_custtable_with_odata(
                            client, company, ginfo["rows"], ginfo["row_details"]
                        )

            # 5. Open known tabs and re-read
            print(f"  Opening tabs...")
            for tname in list(data["tabs"].keys()):
                try:
                    client.call_tool("form_open_or_close_tab", {"tabName": tname, "tabAction": "Open"})
                except Exception:
                    pass

            # 6. form_find_controls sweep
            print(f"  Running form_find_controls sweep...")
            data["find_controls"] = self._find_controls_sweep(client)
            print(f"    Found {len(data['find_controls'])} controls via sweep")

            # 7. Build summary
            data["summary"] = self._build_summary(data)
            data["status"] = "OK"

            close_form(client)

        except Exception as e:
            data["status"] = f"ERROR: {str(e)[:300]}"
            print(f"  ERROR: {e}")
            try:
                close_form(client)
            except Exception:
                pass

        return data

    # ── FormState Parsing ───────────────────────────────────────────────

    def _parse_form_obj(self, form_obj, data, depth=0):
        """Recursively parse FormState.Form into grids, fields, tabs, buttons."""
        if not isinstance(form_obj, dict) or depth > 10:
            return

        # Buttons
        for bname, binfo in form_obj.get("Button", {}).items():
            if isinstance(binfo, dict):
                data["buttons"][bname] = {
                    "label": binfo.get("Label", ""),
                    "help_text": binfo.get("HelpText", ""),
                }

        # Grids
        for gname, ginfo in form_obj.get("Grid", {}).items():
            if not isinstance(ginfo, dict):
                continue
            columns = []
            for col in ginfo.get("Columns", []):
                columns.append({
                    "name": col.get("Name", ""),
                    "label": col.get("Label", ""),
                })
            rows = []
            for row in ginfo.get("Rows", []):
                vals = row.get("Values", {})
                clean = {k: v for k, v in vals.items() if not k.startswith("<")}
                if any(v for v in clean.values()):
                    rows.append(clean)
            pag = ginfo.get("Pagination", {})
            data["grids"][gname] = {
                "text": ginfo.get("Text", ""),
                "columns": columns,
                "rows": rows,
                "total_rows": len(rows),
                "pagination_has_next": str(pag.get("HasNextPage", "False")).lower() == "true",
                "row_details": {},
            }

        # Fields (Input, Checkbox, Combobox, etc.)
        field_types = {
            "Input": self._parse_input,
            "Combobox": self._parse_combobox,
            "Checkbox": self._parse_checkbox,
            "RealInput": self._parse_input,
            "IntInput": self._parse_input,
            "DateInput": self._parse_input,
            "TimeInput": self._parse_input,
            "SegmentedEntry": self._parse_input,
        }
        for ftype, parser in field_types.items():
            for fname, finfo in form_obj.get(ftype, {}).items():
                if isinstance(finfo, dict):
                    data["fields"][fname] = parser(fname, finfo, ftype)

        # Tabs
        for tname, tinfo in form_obj.get("Tab", {}).items():
            if isinstance(tinfo, dict):
                tab_data = {
                    "label": tinfo.get("Label", tinfo.get("Text", "")),
                    "fields": {},
                    "grids": {},
                }
                # Recurse into tab children
                children = tinfo.get("Children", tinfo)
                tab_container = {"grids": {}, "fields": {}, "tabs": {}, "buttons": {}}
                self._parse_form_obj(children, tab_container, depth + 1)
                tab_data["fields"] = tab_container["fields"]
                tab_data["grids"] = tab_container["grids"]
                data["tabs"][tname] = tab_data
                # Also merge tab fields and grids into main data
                data["fields"].update(tab_container["fields"])
                data["grids"].update(tab_container["grids"])

        # TabPage
        for tname, tinfo in form_obj.get("TabPage", {}).items():
            if isinstance(tinfo, dict):
                tab_data = {
                    "label": tinfo.get("Label", tinfo.get("Text", "")),
                    "fields": {},
                    "grids": {},
                }
                children = tinfo.get("Children", tinfo)
                tab_container = {"grids": {}, "fields": {}, "tabs": {}, "buttons": {}}
                self._parse_form_obj(children, tab_container, depth + 1)
                tab_data["fields"] = tab_container["fields"]
                tab_data["grids"] = tab_container["grids"]
                data["tabs"][tname] = tab_data
                data["fields"].update(tab_container["fields"])
                data["grids"].update(tab_container["grids"])

        # Groups
        for gname, ginfo in form_obj.get("Group", {}).items():
            if isinstance(ginfo, dict):
                self._parse_form_obj(ginfo, data, depth + 1)

        # ReferenceGroup
        for gname, ginfo in form_obj.get("ReferenceGroup", {}).items():
            if isinstance(ginfo, dict):
                self._parse_form_obj(ginfo, data, depth + 1)

        # Children (generic container)
        children = form_obj.get("Children", {})
        if isinstance(children, dict):
            self._parse_form_obj(children, data, depth + 1)

    def _parse_input(self, name, info, ftype="Input"):
        return {
            "label": info.get("Label", name),
            "value": info.get("Value", ""),
            "type": ftype.lower(),
            "is_required": info.get("IsRequired", ""),
            "is_editable": info.get("IsEditable", ""),
            "has_lookup": info.get("HasLookup", ""),
        }

    def _parse_combobox(self, name, info, ftype="Combobox"):
        value = info.get("ValueText", info.get("Value", ""))
        options = info.get("Options", [])
        return {
            "label": info.get("Label", name),
            "value": value,
            "raw_value": info.get("Value", ""),
            "type": "combobox",
            "is_required": info.get("IsRequired", ""),
            "is_editable": info.get("IsEditable", ""),
            "options": [{"label": o.get("Label", ""), "value": o.get("Value", "")} for o in options],
        }

    def _parse_checkbox(self, name, info, ftype="Checkbox"):
        return {
            "label": info.get("Label", name),
            "value": info.get("IsChecked", info.get("Value", "")),
            "type": "checkbox",
            "is_required": info.get("IsRequired", ""),
            "is_editable": info.get("IsEditable", ""),
        }

    # ── Grid Pagination ─────────────────────────────────────────────────

    def _find_grid_in_form(self, form_obj, grid_name):
        """Recursively find a grid by name in the FormState tree.
        Grids can be nested inside Tab > Children > Grid, TabPage > Children > Grid, etc.
        """
        if not isinstance(form_obj, dict):
            return None
        # Check direct Grid children
        grids = form_obj.get("Grid", {})
        if isinstance(grids, dict) and grid_name in grids:
            gi = grids[grid_name]
            if isinstance(gi, dict):
                return gi
        # Recurse into Tab, TabPage, Group, Children
        for container_key in ("Tab", "TabPage", "Group", "Children"):
            container = form_obj.get(container_key, {})
            if isinstance(container, dict):
                for _name, child in container.items():
                    if isinstance(child, dict):
                        # Try Children sub-key first, then the child itself
                        for sub in [child.get("Children", {}), child]:
                            result = self._find_grid_in_form(sub, grid_name)
                            if result:
                                return result
        return None

    def _paginate_grid(self, client, grid_name, initial_rows, form_obj):
        """Paginate a grid to collect all rows."""
        all_rows = list(initial_rows)
        page = 1
        max_pages = 200  # safety limit

        while page < max_pages:
            try:
                next_result = mcp_call(client, "form_click_control", {
                    "controlName": grid_name,
                    "actionId": "LoadNextPage",
                })
                if "raw" in next_result:
                    break
                form = next_result.get("FormState", {}).get("Form", {})
                # Search recursively — grid may be nested inside Tabs
                gi = self._find_grid_in_form(form, grid_name)
                if not gi:
                    break
                raw_rows = gi.get("Rows", [])
                page_rows = []
                for row in raw_rows:
                    vals = row.get("Values", {})
                    clean = {k: v for k, v in vals.items() if not k.startswith("<")}
                    if any(v for v in clean.values()):
                        page_rows.append(clean)
                if not page_rows:
                    break
                all_rows.extend(page_rows)
                page += 1
                print(f"      Page {page}: +{len(page_rows)} rows (total: {len(all_rows)})")
                pag = gi.get("Pagination", {})
                if str(pag.get("HasNextPage", "False")).lower() != "true":
                    break
            except Exception as e:
                print(f"      Pagination error: {e}")
                break

        if page > 1:
            print(f"    Grid '{grid_name}': {page} pages, {len(all_rows)} total rows")
        return all_rows

    # ── Row Detail Extraction ───────────────────────────────────────────

    def _extract_row_details(self, client, mi_name, mi_type, company, grid_name, ginfo):
        """Select each grid row, open ALL sub-tabs, and extract detail fields.

        After processing each page of 25 rows (with sub-tab opens), the MCP
        server's grid state becomes corrupted. We work around this by closing
        and reopening the form between pages, then paginating to the next page.
        """
        row_details = {}
        row_sub_grids = {}  # {row_idx: {sub_grid_name: {columns, rows}}}
        total = ginfo["total_rows"]

        page_size = 25
        total_pages = (total + page_size - 1) // page_size

        # Track discovered sub-tabs across rows
        all_tab_names = set()

        for page in range(total_pages):
            start_row = page * page_size
            end_row = min(start_row + page_size, total)

            # For page > 0, reopen the form and paginate to the right page
            if page > 0:
                print(f"      Reopening form for page {page + 1}...")
                try:
                    close_form(client)
                    time.sleep(0.5)
                    form_result = open_form(client, mi_name, mi_type, company)
                    # Paginate to the target page
                    for p in range(page):
                        mcp_call(client, "form_click_control", {
                            "controlName": grid_name,
                            "actionId": "LoadNextPage",
                        })
                    print(f"      On page {page + 1}, rows {start_row}-{end_row - 1}")
                except Exception as e:
                    print(f"      Failed to reopen/paginate to page {page + 1}: {e}")
                    break

            consecutive_fails = 0
            for i in range(start_row, end_row):
                row_in_page = (i % page_size) + 1  # 1-based within page

                # Mid-page recovery: if we hit consecutive failures, reopen
                if consecutive_fails >= 2:
                    print(f"      Mid-page recovery at row {i} (page {page + 1})...")
                    try:
                        close_form(client)
                        time.sleep(0.5)
                        open_form(client, mi_name, mi_type, company)
                        for p in range(page):
                            mcp_call(client, "form_click_control", {
                                "controlName": grid_name,
                                "actionId": "LoadNextPage",
                            })
                        consecutive_fails = 0
                        print(f"      Recovered, resuming from row {i}")
                    except Exception as e:
                        print(f"      Mid-page recovery failed: {e}")
                        break

                try:
                    sel_raw = client.call_tool("form_select_grid_row", {
                        "gridName": grid_name,
                        "rowNumber": str(row_in_page),
                        "marking": "Unmarked",
                    })
                    if not (isinstance(sel_raw, str) and sel_raw.strip()):
                        consecutive_fails += 1
                        continue
                    sel = json.loads(sel_raw)
                    if not isinstance(sel, dict):
                        consecutive_fails += 1
                        continue
                    sel_form = sel.get("FormState", {}).get("Form", {})
                    if not sel_form:
                        consecutive_fails += 1
                        continue

                    consecutive_fails = 0  # reset on success

                    # Parse initial FormState (active tab fields)
                    detail = {"fields": {}, "grids": {}, "tabs": {}, "buttons": {}}
                    self._parse_form_obj(sel_form, detail, depth=0)
                    row_fields = {}
                    row_grids = {}

                    # Build tab-name → tab-label map from discovered tabs
                    tab_label_map = {}
                    for tname, tinfo in detail["tabs"].items():
                        tab_label_map[tname] = tinfo.get("label", tname)

                    # Collect fields, tagging each with its actual parent tab
                    # First, build a map of field→tab from tab-specific children
                    field_to_tab = {}
                    for tname, tinfo in detail["tabs"].items():
                        tab_lbl = tinfo.get("label", tname)
                        for fname in tinfo.get("fields", {}):
                            field_to_tab[fname] = tab_lbl

                    # Now tag each field with its actual parent tab
                    active_tab_label = tab_label_map.get("General", "General")
                    for fname, fdata in detail["fields"].items():
                        val = fdata.get("value", "")
                        if val and str(val).strip():
                            fdata["_tab"] = field_to_tab.get(fname, active_tab_label)
                            row_fields[fname] = fdata

                    # Collect sub-grids from active tab
                    for sgname, sgdata in detail["grids"].items():
                        if sgname != grid_name and sgdata.get("rows"):
                            row_grids[sgname] = sgdata

                    # Discover sub-tabs visible in the row detail area
                    row_tabs = set(detail["tabs"].keys())
                    all_tab_names.update(row_tabs)

                    # Open each sub-tab to capture fields from ALL sections
                    tabs_to_open = sorted(row_tabs)
                    if i == 0 and tabs_to_open:
                        tab_list = ", ".join(tabs_to_open[:8])
                        more = f" +{len(tabs_to_open)-8} more" if len(tabs_to_open) > 8 else ""
                        print(f"    Sub-tabs found: {tab_list}{more}")

                    for tab_name in tabs_to_open:
                        try:
                            tab_raw = client.call_tool(
                                "form_open_or_close_tab", {"tabName": tab_name, "tabAction": "Open"}
                            )
                            if isinstance(tab_raw, str) and tab_raw.strip():
                                tab_parsed = json.loads(tab_raw)
                                if isinstance(tab_parsed, dict):
                                    tab_form = tab_parsed.get("FormState", {}).get("Form", {})
                                    if tab_form:
                                        tab_detail = {"fields": {}, "grids": {}, "tabs": {}, "buttons": {}}
                                        self._parse_form_obj(tab_form, tab_detail, depth=0)
                                        # Resolve tab label
                                        tab_lbl = tab_label_map.get(tab_name, tab_name)
                                        # Also check if the tab response has a better label
                                        for tn, ti in tab_detail["tabs"].items():
                                            if tn == tab_name:
                                                lbl = ti.get("label", "")
                                                if lbl:
                                                    tab_lbl = lbl
                                                    tab_label_map[tab_name] = lbl
                                        # Use ONLY fields from this specific tab's children
                                        # (not the full FormState which includes all tabs)
                                        tab_fields = tab_detail["tabs"].get(tab_name, {}).get("fields", {})
                                        if not tab_fields:
                                            # Fallback: if tab not found by name, use all fields
                                            # but only those not already captured
                                            tab_fields = tab_detail["fields"]
                                        for fname, fdata in tab_fields.items():
                                            val = fdata.get("value", "")
                                            if val and str(val).strip() and fname not in row_fields:
                                                fdata["_tab"] = tab_lbl
                                                row_fields[fname] = fdata
                                        # Same for sub-grids: prefer tab-specific grids
                                        tab_grids = tab_detail["tabs"].get(tab_name, {}).get("grids", {})
                                        if not tab_grids:
                                            tab_grids = tab_detail["grids"]
                                        for sgname, sgdata in tab_grids.items():
                                            if sgname != grid_name and sgdata.get("rows") and sgname not in row_grids:
                                                row_grids[sgname] = sgdata
                        except Exception:
                            pass

                    row_details[i] = row_fields
                    if row_grids:
                        row_sub_grids[i] = row_grids
                    time.sleep(0.1)
                except Exception as e:
                    consecutive_fails += 1
                    print(f"      Row {i} detail extraction failed: {e}")

            print(f"      Page {page + 1}/{total_pages}: details for rows {start_row}-{end_row - 1}")

        # Merge sub-grid data into row_details as flattened fields
        for ri, sub_grids in row_sub_grids.items():
            if ri not in row_details:
                row_details[ri] = {}
            for sgname, sgdata in sub_grids.items():
                # Flatten sub-grid rows into a summary field
                sg_rows = sgdata.get("rows", [])
                if sg_rows:
                    col_labels = [c.get("label", c.get("name", "")) for c in sgdata.get("columns", [])]
                    # Create a compact text summary of sub-grid
                    summaries = []
                    for row in sg_rows[:20]:  # limit to 20 rows
                        parts = [str(row.get(cl, "")) for cl in col_labels if row.get(cl)]
                        if parts:
                            summaries.append(" | ".join(parts))
                    if summaries:
                        row_details[ri][f"__subgrid_{sgname}"] = {
                            "label": f"[SubGrid] {sgname} ({len(sg_rows)} rows)",
                            "value": "; ".join(summaries),
                            "type": "sub_grid_summary",
                        }

        if all_tab_names:
            print(f"    Grid '{grid_name}': opened {len(all_tab_names)} sub-tabs per row")
        print(f"    Grid '{grid_name}': extracted details for {len(row_details)}/{ginfo['total_rows']} rows")
        return row_details

    # ── OData Enrichment for CustTable FastTab Fields ──────────────────

    # CBA/CHB fields in the CustomersV3 OData entity (Acme Asia FastTab)
    _CUSTTABLE_ODATA_FIELDS = [
        "CBAAccountGroup", "CBACPGNameId",
        "CBACommercialHierarchyIdOne", "CBACommercialHierarchyIdTwo", "CBACommercialHierarchyIdThree",
        "CBACountryRegionId", "CBACustomAssociationFive", "CBACustomAssociationOne",
        "CBACustomAssociationThree", "CBACustomAssociationTwo",
        "CBAEInvoiceIndicator", "CBAExciseDutyType", "CBAExpirationRuleGroupId",
        "CBAFChannelCode", "CBAFChannelLFive", "CBAFChannelLFour",
        "CBAFChannelLOne", "CBAFChannelLThree", "CBAFChannelLTwo",
        "CBAFGeographyCode", "CBAFGeographyLFive", "CBAFGeographyLFour",
        "CBAFGeographyLOne", "CBAFGeographyLThree", "CBAFGeographyLTwo",
        "CBAFGlobalChannelCode", "CBASalesOffice", "CBASalesOrganization",
        "CBASalesRegionId", "CBASubTradeChannel", "CBATaxRegstrationType",
        "CHBConsignmentWarehouse",
    ]

    # Human-readable labels for CBA/CHB OData fields
    _CUSTTABLE_ODATA_LABELS = {
        "CBAAccountGroup": "Account group",
        "CBACPGNameId": "CPG name",
        "CBACommercialHierarchyIdOne": "Commercial hierarchy L1",
        "CBACommercialHierarchyIdTwo": "Commercial hierarchy L2",
        "CBACommercialHierarchyIdThree": "Commercial hierarchy L3",
        "CBACountryRegionId": "Country/Region",
        "CBACustomAssociationFive": "Custom association 5",
        "CBACustomAssociationOne": "Custom association 1",
        "CBACustomAssociationThree": "Custom association 3",
        "CBACustomAssociationTwo": "Custom association 2",
        "CBAEInvoiceIndicator": "E-Invoice indicator",
        "CBAExciseDutyType": "Excise duty type",
        "CBAExpirationRuleGroupId": "Expiration rule group",
        "CBAFChannelCode": "Channel code",
        "CBAFChannelLFive": "Channel L5",
        "CBAFChannelLFour": "Channel L4",
        "CBAFChannelLOne": "Channel L1",
        "CBAFChannelLThree": "Channel L3",
        "CBAFChannelLTwo": "Channel L2",
        "CBAFGeographyCode": "Geography code",
        "CBAFGeographyLFive": "Geography L5",
        "CBAFGeographyLFour": "Geography L4",
        "CBAFGeographyLOne": "Geography L1",
        "CBAFGeographyLThree": "Geography L3",
        "CBAFGeographyLTwo": "Geography L2",
        "CBAFGlobalChannelCode": "Global channel code",
        "CBASalesOffice": "Sales office",
        "CBASalesOrganization": "Sales organization",
        "CBASalesRegionId": "Sales region",
        "CBASubTradeChannel": "Sub trade channel",
        "CBATaxRegstrationType": "Tax registration type",
        "CHBConsignmentWarehouse": "Consignment warehouse",
    }

    def _enrich_custtable_with_odata(self, client, company, grid_rows, row_details):
        """Enrich CustTable row_details with CBA/CHB fields from OData CustomersV3.

        The form_find_controls approach returns stale/empty values for FastTab fields
        when iterating rows. OData gives correct per-customer values.
        """
        if not grid_rows:
            return row_details

        # Find the CustomerAccount column label in the grid
        acct_col = None
        for row in grid_rows[:1]:
            for col_label in row.keys():
                if "account" in col_label.lower() or "customer" in col_label.lower():
                    acct_col = col_label
                    break
        if not acct_col:
            print(f"    OData enrich: cannot find CustomerAccount column in grid")
            return row_details

        # Collect all customer accounts from grid rows
        accounts = []
        for row in grid_rows:
            acct = str(row.get(acct_col, "")).strip()
            if acct:
                accounts.append(acct)

        if not accounts:
            print(f"    OData enrich: no customer accounts found in grid")
            return row_details

        # Query OData in batches of 15 (OData $filter has URL length limits)
        le_lower = company.lower()
        select_fields = ",".join(["CustomerAccount", "dataAreaId"] + self._CUSTTABLE_ODATA_FIELDS)
        odata_map = {}  # CustomerAccount -> {field: value, ...}

        batch_size = 15
        for batch_start in range(0, len(accounts), batch_size):
            batch = accounts[batch_start:batch_start + batch_size]
            # Build OR filter: CustomerAccount eq '210205' or CustomerAccount eq '1000' ...
            acct_filters = " or ".join([f"CustomerAccount eq '{a}'" for a in batch])
            query = (
                f"$filter=dataAreaId eq '{le_lower}' and ({acct_filters})"
                f"&$select={select_fields}"
                f"&$top={len(batch)}"
            )
            try:
                raw = client.find_entities("CustomersV3", query)
                if raw:
                    data = json.loads(raw) if isinstance(raw, str) else raw
                    for rec in data.get("value", []):
                        acct = rec.get("CustomerAccount", "")
                        if acct:
                            odata_map[acct] = rec
            except Exception as e:
                print(f"    OData enrich batch error: {e}")

        # Merge OData values into row_details
        enriched = 0
        for row_idx, row_data in enumerate(grid_rows):
            acct = str(row_data.get(acct_col, "")).strip()
            if acct and acct in odata_map:
                rec = odata_map[acct]
                if row_idx not in row_details:
                    row_details[row_idx] = {}
                for field_name in self._CUSTTABLE_ODATA_FIELDS:
                    val = rec.get(field_name, "")
                    val_str = str(val).strip() if val else ""
                    # Store with OData field name as control name
                    label = self._CUSTTABLE_ODATA_LABELS.get(field_name, field_name)
                    row_details[row_idx][f"CustTable_{field_name}"] = {
                        "label": label,
                        "value": val_str,
                        "type": "odata_field",
                        "_tab": "(Acme Asia)",
                    }
                enriched += 1

        print(f"    OData enrich: matched {enriched}/{len(grid_rows)} rows with {len(self._CUSTTABLE_ODATA_FIELDS)} CBA/CHB fields")
        return row_details

    # ── form_find_controls Sweep ────────────────────────────────────────

    def _find_controls_sweep(self, client):
        """Broad sweep using form_find_controls to capture all controls."""
        all_controls = {}

        for term in FIND_CONTROLS_TERMS:
            try:
                raw = client.call_tool("form_find_controls", {"controlSearchTerm": term})
                if isinstance(raw, str) and raw.strip():
                    parsed = json.loads(raw)
                    if isinstance(parsed, list):
                        for item in parsed:
                            cname = item.get("Name", "")
                            if cname and cname not in all_controls:
                                props = item.get("Properties", {})
                                all_controls[cname] = self._normalize_control(cname, props)
            except (json.JSONDecodeError, Exception):
                pass

        return all_controls

    def _normalize_control(self, name, props):
        """Normalize a control from form_find_controls into consistent format."""
        if not isinstance(props, dict):
            return {"name": name, "label": "", "value": "", "type": "unknown"}

        # Grid control
        if "Columns" in props:
            cols = [{"name": c.get("Name", ""), "label": c.get("Label", "")}
                    for c in props.get("Columns", [])]
            rows = []
            for row in props.get("Rows", []):
                vals = row.get("Values", {})
                clean = {k: v for k, v in vals.items() if not k.startswith("<")}
                if any(v for v in clean.values()):
                    rows.append(clean)
            return {
                "name": name,
                "label": props.get("Text", props.get("Label", "")),
                "type": "grid",
                "columns": cols,
                "rows": rows,
                "total_rows": len(rows),
            }

        # Regular field
        value = props.get("ValueText", props.get("Value", ""))
        return {
            "name": name,
            "label": props.get("Label", ""),
            "value": str(value).strip() if value else "",
            "type": "field",
            "is_required": props.get("IsRequired", ""),
            "is_editable": props.get("IsEditable", ""),
            "has_lookup": props.get("HasLookup", ""),
            "options": [
                {"label": o.get("Label", ""), "value": o.get("Value", "")}
                for o in props.get("Options", [])
            ],
        }

    # ── Summary ─────────────────────────────────────────────────────────

    def _build_summary(self, data):
        """Build summary counts."""
        total_controls = 0
        with_values = 0
        grid_count = len(data["grids"])
        total_grid_rows = 0

        # Fields
        for fname, fdata in data["fields"].items():
            total_controls += 1
            val = fdata.get("value", "")
            if val and str(val).strip():
                with_values += 1

        # Grids
        for gname, gdata in data["grids"].items():
            total_controls += 1  # grid itself
            total_grid_rows += gdata["total_rows"]
            if gdata["total_rows"] > 0:
                with_values += 1

        # find_controls (deduplicated)
        fc_extra = 0
        for cname, cdata in data["find_controls"].items():
            if cname not in data["fields"] and cname not in data["grids"]:
                fc_extra += 1
                val = cdata.get("value", "")
                if val and str(val).strip():
                    with_values += 1
                total_controls += 1

        return {
            "total_controls": total_controls,
            "controls_with_values": with_values,
            "grid_count": grid_count,
            "total_grid_rows": total_grid_rows,
            "buttons": len(data["buttons"]),
            "tabs": len(data["tabs"]),
            "find_controls_extra": fc_extra,
        }

    # ── Console Output ──────────────────────────────────────────────────

    @staticmethod
    def print_result(result):
        """Print extraction result to console."""
        print(f"\n{'='*70}")
        print(f"FORM: {result['form_name']} - {result['form_caption']}")
        print(f"Menu Item: {result['menu_item']} ({result['mi_type']})")
        print(f"Path: {result['navigation_path']}")
        print(f"Environment: {result['environment']}")
        print(f"{'='*70}")

        for le, le_data in result["legal_entities"].items():
            print(f"\n{'-'*70}")
            print(f"  LEGAL ENTITY: {le}  |  Status: {le_data['status']}")
            print(f"{'-'*70}")

            if le_data["status"] != "OK":
                continue

            s = le_data["summary"]
            print(f"  Total Controls: {s['total_controls']}  |  With Values: {s['controls_with_values']}")
            print(f"  Grids: {s['grid_count']}  |  Grid Rows: {s['total_grid_rows']}")
            print(f"  Tabs: {s['tabs']}  |  Buttons: {s['buttons']}")

            # Grids
            for gname, gdata in le_data["grids"].items():
                print(f"\n  [GRID] {gname} - {gdata.get('text', '')}")
                print(f"    Columns: {', '.join(c['label'] for c in gdata['columns'])}")
                print(f"    Rows: {gdata['total_rows']}")
                for i, row in enumerate(gdata["rows"][:10]):  # show first 10
                    print(f"    Row {i}: {row}")
                if gdata["total_rows"] > 10:
                    print(f"    ... ({gdata['total_rows'] - 10} more rows)")

                # Row details
                for ri, fields in sorted(gdata.get("row_details", {}).items()):
                    if fields:
                        print(f"\n    Row {ri} details:")
                        for fname, fdata in sorted(fields.items()):
                            print(f"      {fname}: {fdata.get('label', '')} = {fdata.get('value', '')}")

            # Fields with values (not from grids)
            print(f"\n  [FIELDS WITH VALUES]")
            for fname, fdata in sorted(le_data["fields"].items()):
                val = fdata.get("value", "")
                if val and str(val).strip():
                    print(f"    {fname}: {fdata.get('label', '')} = {val}")

            # Fields without values
            empty = [(fn, fd) for fn, fd in le_data["fields"].items()
                     if not (fd.get("value", "") and str(fd["value"]).strip())]
            if empty:
                print(f"\n  [FIELDS WITHOUT VALUES] ({len(empty)})")
                for fname, fdata in sorted(empty):
                    print(f"    {fname}: {fdata.get('label', '')}")

            # Extra from find_controls (not already in fields/grids)
            extra = {cn: cd for cn, cd in le_data["find_controls"].items()
                     if cn not in le_data["fields"] and cn not in le_data["grids"]}
            if extra:
                extra_with_val = {cn: cd for cn, cd in extra.items()
                                  if cd.get("value") and str(cd["value"]).strip()}
                extra_no_val = {cn: cd for cn, cd in extra.items()
                                if not (cd.get("value") and str(cd["value"]).strip())}
                if extra_with_val:
                    print(f"\n  [EXTRA CONTROLS WITH VALUES] ({len(extra_with_val)})")
                    for cn, cd in sorted(extra_with_val.items()):
                        print(f"    {cn}: {cd.get('label', '')} = {cd.get('value', '')}")
                if extra_no_val:
                    print(f"\n  [EXTRA CONTROLS WITHOUT VALUES] ({len(extra_no_val)})")
                    for cn, cd in sorted(extra_no_val.items()):
                        print(f"    {cn}: {cd.get('label', '')}")

            # Buttons
            if le_data["buttons"]:
                print(f"\n  [BUTTONS] ({len(le_data['buttons'])})")
                for bname, bdata in sorted(le_data["buttons"].items()):
                    print(f"    {bname}: {bdata.get('label', '')}")

    # ── Excel Export ────────────────────────────────────────────────────

    @staticmethod
    def to_excel(result, output_dir=None):
        """Export result to Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if not output_dir:
            output_dir = r"C:\D365 Configuration Drift Analysis\output\Tables"
        os.makedirs(output_dir, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        form_name = result["form_name"] or result["menu_item"]
        filename = f"FormControls_{form_name}_{ts}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        val_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        no_val_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        def write_header(ws, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font, c.fill, c.alignment, c.border = hf, hfill, Alignment(horizontal="center"), tb

        def auto_width(ws, max_col, max_row):
            for ci in range(1, max_col + 1):
                mx = 10
                for ri in range(1, min(max_row + 1, 200)):
                    v = ws.cell(row=ri, column=ci).value
                    if v:
                        mx = max(mx, min(len(str(v)), 60))
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(mx + 2, 60)

        # ── Summary sheet ──
        ws = wb.create_sheet("Summary")
        hdrs = ["Legal Entity", "Status", "Total Controls", "With Values",
                "Grid Count", "Total Rows", "Tabs", "Buttons"]
        write_header(ws, hdrs)
        ri = 2
        for le, ld in result["legal_entities"].items():
            s = ld.get("summary", {})
            for ci, v in enumerate([
                le, ld["status"], s.get("total_controls", 0), s.get("controls_with_values", 0),
                s.get("grid_count", 0), s.get("total_grid_rows", 0),
                s.get("tabs", 0), s.get("buttons", 0),
            ], 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = tb
                c.fill = val_fill if ld["status"] == "OK" else no_val_fill
            ri += 1
        auto_width(ws, len(hdrs), ri)
        ws.freeze_panes = "A2"

        # ── Controls Detail sheet ──
        ws2 = wb.create_sheet("Controls Detail")
        hdrs2 = ["Legal Entity", "Control Name", "Label", "Value", "Type",
                 "Required", "Editable", "Source"]
        write_header(ws2, hdrs2)
        ri = 2
        for le, ld in result["legal_entities"].items():
            if ld["status"] != "OK":
                continue
            # Fields from FormState
            for fname, fdata in sorted(ld["fields"].items()):
                val = fdata.get("value", "")
                for ci, v in enumerate([
                    le, fname, fdata.get("label", ""), str(val),
                    fdata.get("type", ""), fdata.get("is_required", ""),
                    fdata.get("is_editable", ""), "FormState",
                ], 1):
                    c = ws2.cell(row=ri, column=ci, value=v)
                    c.border = tb
                    c.fill = val_fill if val and str(val).strip() else no_val_fill
                ri += 1
            # Extra from find_controls
            for cname, cdata in sorted(ld["find_controls"].items()):
                if cname in ld["fields"] or cname in ld["grids"]:
                    continue
                val = cdata.get("value", "")
                for ci, v in enumerate([
                    le, cname, cdata.get("label", ""), str(val),
                    cdata.get("type", ""), cdata.get("is_required", ""),
                    cdata.get("is_editable", ""), "FindControls",
                ], 1):
                    c = ws2.cell(row=ri, column=ci, value=v)
                    c.border = tb
                    c.fill = val_fill if val and str(val).strip() else no_val_fill
                ri += 1
        auto_width(ws2, len(hdrs2), ri)
        ws2.freeze_panes = "A2"

        # ── Grid Data sheet (one per grid) ──
        for le, ld in result["legal_entities"].items():
            if ld["status"] != "OK":
                continue
            for gname, gdata in ld["grids"].items():
                sheet_name = f"Grid_{le}_{gname}"[:31]  # Excel 31 char limit
                ws3 = wb.create_sheet(sheet_name)
                cols = [c["label"] for c in gdata["columns"]]
                write_header(ws3, cols)
                ri = 2
                for row in gdata["rows"]:
                    for ci, col_label in enumerate(cols, 1):
                        c = ws3.cell(row=ri, column=ci, value=row.get(col_label, ""))
                        c.border = tb
                    ri += 1
                auto_width(ws3, len(cols), ri)
                ws3.freeze_panes = "A2"

        wb.save(filepath)
        print(f"\nExcel saved: {filepath}")
        return filepath


# ── CLI ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="D365 Form Control Extractor")
    parser.add_argument("--path", help="Navigation path (e.g., 'Inventory Management -> Setup -> ...')")
    parser.add_argument("--mi", help="Direct menu item name (e.g., 'Interest')")
    parser.add_argument("--mi-type", default="Display", choices=["Display", "Action", "Output"])
    parser.add_argument("--le", nargs="+", help="Legal entities (e.g., MY30 MY60 SG60)")
    parser.add_argument("--env", default="ENV1", help="Environment key (ENV1, ENV4, etc.)")
    parser.add_argument("--excel", action="store_true", help="Export to Excel")
    parser.add_argument("--output-dir", help="Output directory for Excel")

    args = parser.parse_args()

    # Interactive mode if no args
    if not args.path and not args.mi:
        args.path = input("Navigation path (or menu item name): ").strip()
    if not args.le:
        le_input = input("Legal entities (comma/space separated, e.g., MY30,MY60,SG60): ").strip()
        args.le = [x.strip() for x in le_input.replace(",", " ").split() if x.strip()]
    if not args.le:
        args.le = ["MY30"]

    path_or_mi = args.path or args.mi
    mi_type = args.mi_type

    print(f"\n{'='*70}")
    print(f"D365 Form Control Extractor")
    print(f"{'='*70}")
    print(f"  Path/MI:  {path_or_mi}")
    print(f"  MI Type:  {mi_type}")
    print(f"  LEs:      {', '.join(args.le)}")
    print(f"  Env:      {args.env}")

    extractor = FormControlExtractor(env_key=args.env)
    result = extractor.extract(path_or_mi, args.le, mi_type)

    # Print to console
    FormControlExtractor.print_result(result)

    # Excel
    if args.excel or True:  # always export
        FormControlExtractor.to_excel(result, args.output_dir)


if __name__ == "__main__":
    main()
