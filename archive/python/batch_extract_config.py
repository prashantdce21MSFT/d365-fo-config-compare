"""
Batch form data extraction — ALL 276 MY/SG CDDs with Screenshot Path.
Reads form data from Config (ENV4) for MY30, MY60, SG60, DAT.
Uses form tools with pagination (>25 rows).
Outputs a single Excel file with all fetched data.
Checkpoints progress so it can resume after interruption.
"""
import sys
import os
import json
import time
import traceback
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from d365_mcp_client import D365McpClient
from form_reader import (
    open_form, close_form, read_form_records, find_menu_item,
    resolve_menu_item_from_path,
)

# ── Config ────────────────────────────────────────────────────────────
CONFIG_FILE = r"C:\D365DataValidator\config.json"
EXCEL_FILE = r"C:\D365 Configuration Drift Analysis\output\MYSG_CDD_FormPaths_20260428_230658.xlsx"
OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output"
CHECKPOINT_FILE = os.path.join(OUTPUT_DIR, "batch_extract_checkpoint.json")
MI_CACHE_FILE = os.path.join(OUTPUT_DIR, "batch_extract_mi_cache.json")
DATA_DIR = os.path.join(OUTPUT_DIR, "form_data")

ENV_KEY = "ENV4"
ENV_DISPLAY = "Env4 Config"
COMPANIES = ["MY30", "MY60", "SG60", "DAT"]


def load_config():
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)


def load_excel_items():
    """Load all rows with Screenshot Path populated."""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb["CDD Form Paths"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        screenshot_path = row[10]  # Screenshot Path (from images)
        if not screenshot_path or not str(screenshot_path).strip():
            continue
        items.append({
            "ado_id": str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0]),
            "title": row[1] or "",
            "screenshot_path": str(screenshot_path).strip(),
            "area_path": row[2] or "",
            "state": row[3] or "",
        })
    wb.close()
    return items


def load_checkpoint():
    if os.path.isfile(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r") as f:
            return json.load(f)
    return {"completed": {}, "failed": {}, "skipped": {}}


def save_checkpoint(cp):
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(cp, f, indent=2)


def load_mi_cache():
    if os.path.isfile(MI_CACHE_FILE):
        with open(MI_CACHE_FILE, "r") as f:
            return json.load(f)
    return {}


def save_mi_cache(cache):
    with open(MI_CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def save_item_data(ado_id, company, records, columns):
    """Save records for one ADO+company to a JSON file."""
    os.makedirs(DATA_DIR, exist_ok=True)
    fpath = os.path.join(DATA_DIR, f"{ado_id}_{company}.json")
    with open(fpath, "w") as f:
        json.dump({"records": records, "columns": columns}, f, indent=1)


def load_item_data(ado_id, company):
    """Load previously saved records."""
    fpath = os.path.join(DATA_DIR, f"{ado_id}_{company}.json")
    if os.path.isfile(fpath):
        with open(fpath, "r") as f:
            d = json.load(f)
            return d.get("records", []), d.get("columns", [])
    return None, None


def resolve_mi(client, screenshot_path, mi_cache):
    """Resolve menu item, using cache if available."""
    cache_key = screenshot_path.strip()
    if cache_key in mi_cache:
        cached = mi_cache[cache_key]
        return cached.get("mi_name"), cached.get("mi_type"), cached.get("hint")

    mi_name, mi_type, hint = resolve_menu_item_from_path(client, screenshot_path, "MY30")
    mi_cache[cache_key] = {
        "mi_name": mi_name,
        "mi_type": mi_type,
        "hint": hint,
    }
    return mi_name, mi_type, hint


def read_form_data(env_config, company, mi_name, mi_type):
    """Connect, open form, read all records with pagination + detail fields.

    Two-pass for >25 rows: paginate grid first, then re-open and select
    each row page-by-page to read Input/Checkbox/Tab detail fields.
    """
    from form_reader import (
        mcp_call, _collect_all_grid_rows, _extract_fields, extract_form_data,
    )

    client = D365McpClient(env_config)
    client.connect()

    try:
        form_result = open_form(client, mi_name, mi_type, company)

        # Pass 1: collect all grid rows via pagination
        grid_name, grid_rows, last_form = _collect_all_grid_rows(client, form_result)
        total_rows = len(grid_rows)

        if total_rows == 0:
            # Parameter/settings form — no grid
            form_data = extract_form_data(client, form_result)
            record = {}
            for fname, finfo in form_data["fields"].items():
                record[finfo.get("label", fname)] = finfo.get("value", "")
            for tab_name, tab_info in form_data["tabs"].items():
                for fname, finfo in tab_info["fields"].items():
                    record[finfo.get("label", fname)] = finfo.get("value", "")
            close_form(client)
            if record:
                return [record], list(record.keys())
            return [], []

        if total_rows <= 25:
            # Single page — detail read directly
            records = []
            columns = set()
            for r in grid_rows:
                columns.update(r.keys())

            for row_idx in range(1, total_rows + 1):
                try:
                    sel_result = mcp_call(client, "form_select_grid_row", {
                        "gridName": grid_name,
                        "rowNumber": row_idx,
                        "marking": "Unmarked",
                    })
                    record = dict(grid_rows[row_idx - 1])
                    if "raw" not in sel_result:
                        sel_form = sel_result.get("FormState", {}).get("Form", {})
                        detail_fields = {}
                        _extract_fields(sel_form, detail_fields)
                        for fname, finfo in detail_fields.items():
                            label = finfo.get("label", fname)
                            val = finfo.get("value", "")
                            if label not in record:
                                record[label] = val
                                columns.add(label)
                        for tab_name, tab_info in sel_form.get("Tab", {}).items():
                            if not isinstance(tab_info, dict):
                                continue
                            children = tab_info.get("Children", {})
                            tab_fields = {}
                            _extract_fields(children, tab_fields)
                            for fname, finfo in tab_fields.items():
                                label = finfo.get("label", fname)
                                val = finfo.get("value", "")
                                if label not in record:
                                    record[label] = val
                                    columns.add(label)
                    records.append(record)
                    time.sleep(0.15)
                except Exception:
                    records.append(dict(grid_rows[row_idx - 1]))

            close_form(client)
            return records, sorted(columns)

        # >25 rows — re-open and walk page by page
        close_form(client)

        PAGE_SIZE = 25
        all_records = []
        all_columns = set()
        for r in grid_rows:
            all_columns.update(r.keys())

        page_num = 0
        rows_processed = 0

        while rows_processed < total_rows:
            client2 = D365McpClient(env_config)
            client2.connect()
            form_result2 = open_form(client2, mi_name, mi_type, company)

            # Navigate to correct page
            if page_num > 0:
                for p in range(page_num):
                    frm = form_result2.get("FormState", {}).get("Form", {})
                    gn = None
                    for g, gi in frm.get("Grid", {}).items():
                        if isinstance(gi, dict) and "Rows" in gi:
                            gn = g
                            break
                    if not gn:
                        break
                    next_r = mcp_call(client2, "form_click_control", {
                        "controlName": gn, "actionId": "LoadNextPage",
                    })
                    if "raw" in next_r:
                        break
                    form_result2 = next_r

            page_start = page_num * PAGE_SIZE
            page_end = min(page_start + PAGE_SIZE, total_rows)

            for row_idx in range(1, page_end - page_start + 1):
                global_idx = page_start + row_idx - 1
                try:
                    sel_result = mcp_call(client2, "form_select_grid_row", {
                        "gridName": grid_name,
                        "rowNumber": row_idx,
                        "marking": "Unmarked",
                    })
                    record = dict(grid_rows[global_idx]) if global_idx < len(grid_rows) else {}
                    if "raw" not in sel_result:
                        sel_form = sel_result.get("FormState", {}).get("Form", {})
                        detail_fields = {}
                        _extract_fields(sel_form, detail_fields)
                        for fname, finfo in detail_fields.items():
                            label = finfo.get("label", fname)
                            val = finfo.get("value", "")
                            if label not in record:
                                record[label] = val
                                all_columns.add(label)
                        for tab_name, tab_info in sel_form.get("Tab", {}).items():
                            if not isinstance(tab_info, dict):
                                continue
                            children = tab_info.get("Children", {})
                            tab_fields = {}
                            _extract_fields(children, tab_fields)
                            for fname, finfo in tab_fields.items():
                                label = finfo.get("label", fname)
                                val = finfo.get("value", "")
                                if label not in record:
                                    record[label] = val
                                    all_columns.add(label)
                    all_records.append(record)
                    time.sleep(0.15)
                except Exception:
                    if global_idx < len(grid_rows):
                        all_records.append(dict(grid_rows[global_idx]))
                    else:
                        all_records.append({})

            rows_processed = page_end
            page_num += 1
            close_form(client2)

        return all_records, sorted(all_columns)

    except Exception as e:
        try:
            close_form(client)
        except:
            pass
        raise


def generate_excel(items, checkpoint, mi_cache):
    """Generate final Excel with all collected data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["MYSG CDD Form Data Extraction"])
    ws_sum.append(["Environment", ENV_DISPLAY])
    ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_sum.append(["Total items with Screenshot Path", len(items)])
    ws_sum.append(["Successfully extracted", len(checkpoint.get("completed", {}))])
    ws_sum.append(["Failed", len(checkpoint.get("failed", {}))])
    ws_sum.append(["Skipped (no MI resolved)", len(checkpoint.get("skipped", {}))])
    ws_sum.append([])
    ws_sum.append(["ADO ID", "Title", "Screenshot Path", "Menu Item", "MY30 Rows", "MY60 Rows", "SG60 Rows", "DAT Rows", "Status"])

    for item in items:
        ado_id = item["ado_id"]
        cached_mi = mi_cache.get(item["screenshot_path"], {})
        mi_name = cached_mi.get("mi_name", "")

        row_counts = {}
        for co in COMPANIES:
            recs, _ = load_item_data(ado_id, co)
            row_counts[co] = len(recs) if recs else 0

        status = "OK"
        if ado_id in checkpoint.get("failed", {}):
            status = f"FAILED: {checkpoint['failed'][ado_id][:60]}"
        elif ado_id in checkpoint.get("skipped", {}):
            status = "No MI resolved"
        elif ado_id not in checkpoint.get("completed", {}):
            status = "Not processed"

        ws_sum.append([
            int(ado_id), item["title"], item["screenshot_path"], mi_name or "",
            row_counts.get("MY30", 0), row_counts.get("MY60", 0),
            row_counts.get("SG60", 0), row_counts.get("DAT", 0),
            status,
        ])

    for col in ["A", "B", "C", "D"]:
        ws_sum.column_dimensions[col].width = 30
    for col in ["E", "F", "G", "H"]:
        ws_sum.column_dimensions[col].width = 12
    ws_sum.column_dimensions["I"].width = 40

    # ── Data sheets — one per ADO item ──
    sheet_count = 0
    for item in items:
        ado_id = item["ado_id"]
        if ado_id not in checkpoint.get("completed", {}):
            continue

        # Collect all data for this ADO across companies
        all_columns = set()
        company_data = {}
        for co in COMPANIES:
            recs, cols = load_item_data(ado_id, co)
            if recs:
                company_data[co] = recs
                all_columns.update(cols)

        if not company_data:
            continue

        sorted_cols = sorted(all_columns)
        # Sheet name: "36293_Warehouses" (max 31 chars)
        title_short = item["title"].split(">")[-1].strip()[:15].replace("[", "").replace("]", "")
        sheet_name = f"{ado_id}_{title_short}"[:31]

        # Avoid duplicate sheet names
        existing = {ws.title for ws in wb.worksheets}
        if sheet_name in existing:
            sheet_name = f"{ado_id}"[:31]

        ws = wb.create_sheet(sheet_name)

        # Header: Company + all column names
        header = ["Company"] + sorted_cols
        for ci, h in enumerate(header, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows grouped by company
        row_num = 2
        for co in COMPANIES:
            recs = company_data.get(co, [])
            for rec in recs:
                ws.cell(row=row_num, column=1, value=co).border = thin_border
                for ci, col_name in enumerate(sorted_cols, 2):
                    val = rec.get(col_name, "")
                    ws.cell(row=row_num, column=ci, value=val).border = thin_border
                row_num += 1

        ws.freeze_panes = "B2"

        # Auto-width
        for ci, col_name in enumerate(header, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(len(str(col_name)) + 2, 10), 40)

        sheet_count += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"MYSG_CDD_FormData_Config_{ts}.xlsx")
    wb.save(output_path)
    print(f"\nExcel saved: {output_path} ({sheet_count} data sheets)")
    return output_path


def main():
    config = load_config()
    env_config = config["environments"][ENV_KEY]

    items = load_excel_items()
    print(f"Loaded {len(items)} items with Screenshot Path")

    checkpoint = load_checkpoint()
    mi_cache = load_mi_cache()

    # Create a persistent client for MI resolution
    resolve_client = D365McpClient(env_config)
    resolve_client.connect()

    done = len(checkpoint.get("completed", {}))
    failed = len(checkpoint.get("failed", {}))
    skipped = len(checkpoint.get("skipped", {}))
    print(f"Checkpoint: {done} done, {failed} failed, {skipped} skipped, {len(items) - done - failed - skipped} remaining")

    # Batch limit — process N new items per run, then stop
    BATCH_LIMIT = int(os.environ.get("BATCH_LIMIT", "10"))
    batch_processed = 0
    batch_completed_ids = set()  # ADO IDs handled in this batch run

    for idx, item in enumerate(items):
        ado_id = item["ado_id"]

        # Skip already processed
        if ado_id in checkpoint.get("completed", {}):
            continue
        if ado_id in checkpoint.get("skipped", {}):
            continue

        if batch_processed >= BATCH_LIMIT:
            print(f"\n--- Batch limit ({BATCH_LIMIT}) reached. Run again to continue. ---")
            break

        progress = done + failed + skipped
        print(f"\n[{progress + 1}/{len(items)}] ADO {ado_id}: {item['screenshot_path'][:80]}")

        # Step 1: Resolve MI
        try:
            mi_name, mi_type, hint = resolve_mi(resolve_client, item["screenshot_path"], mi_cache)
        except Exception as e:
            # Reconnect and retry once
            print(f"  MI resolve error, reconnecting: {e}")
            try:
                resolve_client = D365McpClient(env_config)
                resolve_client.connect()
                mi_name, mi_type, hint = resolve_mi(resolve_client, item["screenshot_path"], mi_cache)
            except Exception as e2:
                print(f"  MI resolve failed: {e2}")
                checkpoint.setdefault("failed", {})[ado_id] = f"MI resolve: {str(e2)[:100]}"
                save_checkpoint(checkpoint)
                failed += 1
                batch_processed += 1
                batch_completed_ids.add(ado_id)
                continue

        save_mi_cache(mi_cache)

        if not mi_name:
            print(f"  SKIP — no menu item resolved")
            checkpoint.setdefault("skipped", {})[ado_id] = "No MI resolved"
            save_checkpoint(checkpoint)
            skipped += 1
            batch_processed += 1
            batch_completed_ids.add(ado_id)
            continue

        print(f"  MI: {mi_name} ({mi_type}), hint: {hint}")

        # Step 2: Read form for each company
        item_ok = True
        for company in COMPANIES:
            # Check if already have data
            existing_recs, _ = load_item_data(ado_id, company)
            if existing_recs is not None:
                print(f"  {company}: cached ({len(existing_recs)} rows)")
                continue

            print(f"  {company}: reading...", end=" ", flush=True)
            try:
                records, columns = read_form_data(env_config, company, mi_name, mi_type)
                save_item_data(ado_id, company, records, columns)
                print(f"{len(records)} rows, {len(columns)} cols")
            except Exception as e:
                err_msg = str(e)[:120]
                print(f"ERROR: {err_msg}")
                # Save empty so we don't retry
                save_item_data(ado_id, company, [], [])
                # Don't mark whole item as failed — continue with other companies

            time.sleep(0.5)

        checkpoint.setdefault("completed", {})[ado_id] = datetime.now().isoformat()
        checkpoint.get("failed", {}).pop(ado_id, None)
        save_checkpoint(checkpoint)
        done += 1
        batch_processed += 1
        batch_completed_ids.add(ado_id)

        # Save MI cache periodically
        if done % 10 == 0:
            save_mi_cache(mi_cache)
            print(f"  --- Progress: {done} done, {failed} failed, {skipped} skipped ---")

    # Final save
    save_mi_cache(mi_cache)
    save_checkpoint(checkpoint)

    print(f"\n{'='*60}")
    print(f"BATCH COMPLETE")
    print(f"  This batch: {batch_processed} items")
    print(f"  Total completed: {done}")
    print(f"  Total failed: {failed}")
    print(f"  Total skipped: {skipped}")
    print(f"  Remaining: {len(items) - done - failed - skipped}")
    print(f"{'='*60}")

    # Generate Excel for THIS batch only (items completed in this run)
    if batch_processed > 0:
        print("\nGenerating Excel for this batch...")
        generate_batch_excel(items, checkpoint, mi_cache, batch_completed_ids, BATCH_LIMIT)


def generate_batch_excel(items, checkpoint, mi_cache, batch_ids, batch_limit=10):
    """Generate Excel for just the items completed in this batch."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    company_fills = {
        "MY30": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
        "MY60": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "SG60": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "DAT":  PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
    }

    # Filter to only batch items
    batch_items = [it for it in items if it["ado_id"] in batch_ids]

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    total_done = len(checkpoint.get("completed", {}))
    total_all = len(items)
    ws_sum.append(["MYSG CDD Form Data Extraction — Batch"])
    ws_sum.append(["Environment", ENV_DISPLAY])
    ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_sum.append(["Items in this batch", len(batch_items)])
    ws_sum.append(["Overall progress", f"{total_done}/{total_all} completed"])
    ws_sum.append([])
    ws_sum.append(["ADO ID", "Title", "Screenshot Path", "Menu Item",
                    "MY30 Rows", "MY60 Rows", "SG60 Rows", "DAT Rows", "Status"])

    for item in batch_items:
        ado_id = item["ado_id"]
        cached_mi = mi_cache.get(item["screenshot_path"], {})
        mi_name = cached_mi.get("mi_name", "")
        row_counts = {}
        for co in COMPANIES:
            recs, _ = load_item_data(ado_id, co)
            row_counts[co] = len(recs) if recs else 0
        status = "OK"
        if ado_id in checkpoint.get("failed", {}):
            status = f"FAILED: {checkpoint['failed'][ado_id][:60]}"
        elif ado_id in checkpoint.get("skipped", {}):
            status = "Skipped (no MI)"
        ws_sum.append([
            int(ado_id), item["title"], item["screenshot_path"], mi_name or "",
            row_counts.get("MY30", 0), row_counts.get("MY60", 0),
            row_counts.get("SG60", 0), row_counts.get("DAT", 0), status,
        ])

    for col in ["A", "B", "C", "D"]:
        ws_sum.column_dimensions[col].width = 30
    for col in ["E", "F", "G", "H"]:
        ws_sum.column_dimensions[col].width = 12
    ws_sum.column_dimensions["I"].width = 40

    # ── One data sheet per ADO item ──
    sheet_count = 0
    for item in batch_items:
        ado_id = item["ado_id"]
        if ado_id not in checkpoint.get("completed", {}):
            continue

        all_columns = set()
        company_data = {}
        for co in COMPANIES:
            recs, cols = load_item_data(ado_id, co)
            if recs:
                company_data[co] = recs
                for r in recs:
                    all_columns.update(r.keys())

        if not company_data:
            continue

        sorted_cols = sorted(all_columns)
        title_short = item["title"].split(">")[-1].strip()[:15].replace("[", "").replace("]", "")
        sheet_name = f"{ado_id}_{title_short}"[:31]
        existing = {ws.title for ws in wb.worksheets}
        if sheet_name in existing:
            sheet_name = f"{ado_id}"[:31]

        ws = wb.create_sheet(sheet_name)
        header = ["Company"] + sorted_cols
        for ci, h in enumerate(header, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        row_num = 2
        for co in COMPANIES:
            fill = company_fills.get(co)
            for rec in company_data.get(co, []):
                ws.cell(row=row_num, column=1, value=co).border = thin_border
                if fill:
                    ws.cell(row=row_num, column=1).fill = fill
                for ci, col_name in enumerate(sorted_cols, 2):
                    val = rec.get(col_name, "")
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill
                row_num += 1

        ws.freeze_panes = "B2"
        if row_num > 1:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(header))}{row_num - 1}"
        for ci, col_name in enumerate(header, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(
                max(len(str(col_name)) + 2, 10), 40
            )
        sheet_count += 1

    # Move Summary first
    wb.move_sheet("Summary", offset=-sheet_count)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    batch_num = (total_done - 1) // batch_limit + 1
    output_path = os.path.join(OUTPUT_DIR, f"MYSG_FormData_Batch{batch_num}_{ts}.xlsx")
    wb.save(output_path)
    print(f"Excel saved: {output_path} ({sheet_count} data sheets)")
    return output_path


if __name__ == "__main__":
    main()
