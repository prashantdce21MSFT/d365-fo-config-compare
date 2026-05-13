"""
Batch extraction for Asset Management & Fixed Assets ADO items.

Reads Correct_URL from CDD_FormMapping.xlsx at runtime.
Parses mi= from each URL. Items sharing the same MI are extracted once.
Items with blank Correct_URL are skipped with a warning.

Extraction uses Playwright CDP (connect to existing Chrome on port 9222).
Navigates directly to the URL — no MCP, no D365McpClient.

Output format matches DeepCompare_smmParameters_20260501_170853.xlsx exactly:
  Summary | per-tab sheets (#, Control, Label, UAT Value, Config Value, Match, LE)
  | {TabSheet}_{Grid}_Grid sheets | All Controls flat sheet
"""
import sys
import os
import json
import time
import traceback
import io
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse, urljoin
from datetime import datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# Force unbuffered output so progress shows in CMD immediately
import builtins
_orig_print = builtins.print
builtins.print = lambda *a, **kw: _orig_print(*a, **{**kw, "flush": True})

OUTPUT_DIR   = r"C:\D365 Configuration Drift Analysis\output\Asset  and FA"
CHECKPOINT   = os.path.join(OUTPUT_DIR, "_checkpoint.json")
MAPPING_FILE = r"C:\D365 Configuration Drift Analysis\output\CDD_FormMapping.xlsx"
LES          = ["MY30", "MY60", "SG60"]

CHROME_PROFILE = r"C:\D365 Configuration Drift Analysis\_chrome_profile"
UAT_HOST    = "example-uat.sandbox.operations.dynamics.com"
CONFIG_HOST = "example-config.sandbox.operations.dynamics.com"

# ADO IDs in scope
ADO_IDS = [
    36742, 36743, 36744, 36745, 36749, 36750, 36753, 36754,
    36755, 36756, 36757, 36759, 36760, 36761, 36762, 36763,
    36767, 36768, 36777, 36780, 36788, 36789,
    36861, 36862, 36874, 36875,
    37040, 37041, 37042, 37043, 37044, 37045, 37046, 37047,
    37048, 37049, 37050, 37051, 37052, 37053, 37054, 37055,
    37056, 37066,
    42060, 42061, 42062, 42063, 42064, 42066, 42078, 42085,
    51991,
]


def load_items_from_mapping():
    """Read Correct_URL from CDD_FormMapping sheet. Parse mi= from URL.
    Returns (items, skipped) where items = [(ado_id, mi, title, correct_url), ...]
    and skipped = [(ado_id, title)] for rows with blank Correct_URL."""
    import openpyxl
    wb  = openpyxl.load_workbook(MAPPING_FILE, data_only=True)
    ws  = wb.active

    # Find header row (row 11) and column indexes
    header_row = 11
    headers = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=ci).value
        if v:
            headers[str(v).strip()] = ci

    col_ado   = headers.get("ADO ID")
    col_title = headers.get("ADO Title")
    col_url   = headers.get("Correct_URL")

    items   = []
    skipped = []
    ado_set = set(ADO_IDS)

    for r in range(header_row + 1, ws.max_row + 1):
        ado_val = ws.cell(row=r, column=col_ado).value
        if not ado_val:
            continue
        try:
            ado_id = int(ado_val)
        except (ValueError, TypeError):
            continue
        if ado_id not in ado_set:
            continue

        title   = str(ws.cell(row=r, column=col_title).value or "").strip()
        url_val = ws.cell(row=r, column=col_url).value

        if not url_val:
            skipped.append((ado_id, title))
            continue

        url_str = str(url_val).strip()

        # Parse mi= from URL query string
        qs  = parse_qs(urlparse(url_str).query)
        mi  = qs.get("mi", [None])[0]
        if not mi:
            skipped.append((ado_id, title))
            continue

        items.append((ado_id, mi, title, url_str))

    return items, skipped


def group_by_mi(items):
    """Dedupe by MI — items sharing the same MI are extracted once.
    Keeps the first Correct_URL seen for that MI."""
    seen  = {}
    order = []
    for ado_id, mi, title, url in items:
        if mi not in seen:
            seen[mi]  = {"mi": mi, "title": title, "ado_ids": [], "url": url}
            order.append(mi)
        seen[mi]["ado_ids"].append(ado_id)
    return [
        (seen[k]["mi"], seen[k]["ado_ids"], seen[k]["title"], seen[k]["url"])
        for k in order
    ]


def build_url(base_url, le, env):
    """Substitute cmp=LE and hostname for the target environment."""
    parsed = urlparse(base_url)
    qs     = parse_qs(parsed.query, keep_blank_values=True)

    # Swap legal entity
    qs["cmp"] = [le]

    # Swap hostname for config env
    if env == "config":
        netloc = CONFIG_HOST
    else:
        netloc = UAT_HOST

    new_query = urlencode({k: v[0] for k, v in qs.items()})
    new_url   = urlunparse((parsed.scheme, netloc, parsed.path, "", new_query, ""))
    return new_url


# ── Playwright extraction helpers ─────────────────────────────────────────────

def wait_ms(page, ms=1500):
    page.wait_for_timeout(ms)


def get_tab_names(page):
    return page.evaluate("""() => {
        const candidates = [
            ...document.querySelectorAll('[class*="verticalTabs"] button'),
            ...document.querySelectorAll('[class*="sysTabControl"] [role="tab"]'),
            ...document.querySelectorAll('[role="tablist"] [role="tab"]'),
            ...document.querySelectorAll('[class*="tabPage"] button'),
            ...document.querySelectorAll('[class*="pivot"] button'),
        ];
        const seen = new Set();
        const tabs = [];
        candidates.forEach((el, idx) => {
            const name = el.textContent.trim().replace(/\\s+/g, ' ');
            if (name && name.length < 120 && !seen.has(name)) {
                seen.add(name);
                tabs.push({ name, idx });
            }
        });
        return tabs;
    }""")


def click_tab(page, tab_idx):
    """Click tab by index; return its aria-controls panel ID (or True/False)."""
    try:
        return page.evaluate(f"""() => {{
            const candidates = [
                ...document.querySelectorAll('[class*="verticalTabs"] button'),
                ...document.querySelectorAll('[class*="sysTabControl"] [role="tab"]'),
                ...document.querySelectorAll('[role="tablist"] [role="tab"]'),
                ...document.querySelectorAll('[class*="tabPage"] button'),
                ...document.querySelectorAll('[class*="pivot"] button'),
            ];
            const seen = new Set();
            const tabs = [];
            candidates.forEach(el => {{
                const name = el.textContent.trim().replace(/\\s+/g, ' ');
                if (name && name.length < 120 && !seen.has(name)) {{
                    seen.add(name);
                    tabs.push(el);
                }}
            }});
            const tab = tabs[{tab_idx}];
            if (!tab) return null;
            tab.click();
            // Return aria-controls so caller can scope extraction to exact panel
            return tab.getAttribute('aria-controls') || '__clicked__';
        }}""")
    except Exception:
        return None


def expand_all_fasttabs(page):
    try:
        page.evaluate("""() => {
            const sels = [
                '[class*="sectionHeader"][aria-expanded="false"]',
                '[class*="groupHeader"][aria-expanded="false"]',
                '[class*="headerExpander"][aria-expanded="false"]',
                'button[aria-expanded="false"][class*="section"]',
                '[class*="expandButton"][aria-expanded="false"]',
                '[class*="fastTabHeader"] button[aria-expanded="false"]',
                '[data-dyn-controltype="Group"] button[aria-expanded="false"]',
            ];
            sels.forEach(s => document.querySelectorAll(s).forEach(el => el.click()));
        }""")
        page.wait_for_timeout(1000)
        # Second pass — expanding one section can reveal nested collapsed ones
        page.evaluate("""() => {
            document.querySelectorAll('[aria-expanded="false"]').forEach(el => {
                if ((el.tagName === 'BUTTON' || el.getAttribute('role') === 'button') &&
                    el.closest('[class*="section"], [class*="group"], [class*="fastTab"]')) {
                    el.click();
                }
            });
        }""")
        page.wait_for_timeout(800)
    except Exception:
        pass


def extract_fields_js(page, panel_id=None):
    """Extract D365 leaf-level fields from the active tab panel.

    panel_id: aria-controls value returned by click_tab — used to scope to the
    exact panel the clicked tab owns (most reliable). Falls back to aria-hidden/
    display detection when panel_id is None or element not found.
    """
    pid_js = f'"{panel_id}"' if panel_id and panel_id != "__clicked__" else "null"
    return page.evaluate(f"""() => {{
        const panelId = {pid_js};
        // Primary: scope to the exact panel the clicked tab controls
        let activePanel = panelId ? document.getElementById(panelId) : null;
        if (!activePanel) {{
            // Fallback 1: aria-hidden + computed display
            activePanel = [...document.querySelectorAll('[role="tabpanel"]')]
                .find(p => {{
                    if (p.getAttribute('aria-hidden') === 'true') return false;
                    const st = window.getComputedStyle(p);
                    return st.display !== 'none' && st.visibility !== 'hidden';
                }});
        }}
        if (!activePanel) {{
            // Fallback 2: aria-selected tab → aria-controls
            const selTab = [...document.querySelectorAll('[role="tab"]')]
                .find(t => t.getAttribute('aria-selected') === 'true');
            const cid = selTab?.getAttribute('aria-controls');
            if (cid) activePanel = document.getElementById(cid);
        }}
        activePanel = activePanel || document.querySelector('[role="main"]') || document.body;

        const results = [];
        const seen    = new Set();

        const allCtrls = [...activePanel.querySelectorAll('[data-dyn-controlname]')];

        allCtrls.forEach(ctrl => {{
            const st = window.getComputedStyle(ctrl);
            if (st.display === 'none' || st.visibility === 'hidden') return;
            if (ctrl.closest('[aria-hidden="true"]')) return;
            if (ctrl.closest('[role="grid"], [class*="gridView"]')) return;

            const name = ctrl.getAttribute('data-dyn-controlname') || '';

            const directLabel = [...ctrl.querySelectorAll('label')].find(lbl => {{
                return lbl.closest('[data-dyn-controlname]') === ctrl;
            }});
            if (!directLabel) return;
            if (directLabel.classList.contains('checkBox-container')) return;

            let label = directLabel.textContent.trim().replace(/\\s+/g, ' ');
            if (!label || label.length > 200) return;

            let value = '';
            const toggleVal = ctrl.querySelector('.toggle-value');
            if (toggleVal) {{
                value = toggleVal.textContent.trim();
            }} else {{
                const inp = ctrl.querySelector(
                    'input[type="text"], input[type="number"], input[type="date"]'
                );
                const chk = ctrl.querySelector('input[type="checkbox"]');
                const sel = ctrl.querySelector('select');
                if (chk) value = chk.checked ? 'Yes' : 'No';
                else if (inp) value = inp.value;
                else if (sel) value = sel.options[sel.selectedIndex]?.text || '';
            }}

            value = String(value).replace(/\\s+/g, ' ').trim();
            const key = (name || label) + '::' + value;
            if (!seen.has(key)) {{
                seen.add(key);
                results.push({{ label, value, name }});
            }}
        }});

        return results;
    }}""")


def extract_grids_js(page, panel_id=None):
    """Extract all visible grids scoped to the active (or panel_id) tab panel."""
    pid_js = f'"{panel_id}"' if panel_id and panel_id != "__clicked__" else "null"
    return page.evaluate(f"""() => {{
        const panelId = {pid_js};
        let activePanel = panelId ? document.getElementById(panelId) : null;
        if (!activePanel) {{
            activePanel = [...document.querySelectorAll('[role="tabpanel"]')]
                .find(p => {{
                    if (p.getAttribute('aria-hidden') === 'true') return false;
                    const st = window.getComputedStyle(p);
                    return st.display !== 'none' && st.visibility !== 'hidden';
                }});
        }}
        if (!activePanel) {{
            const selTab = [...document.querySelectorAll('[role="tab"]')]
                .find(t => t.getAttribute('aria-selected') === 'true');
            const cid = selTab?.getAttribute('aria-controls');
            if (cid) activePanel = document.getElementById(cid);
        }}
        activePanel = activePanel || document.querySelector('[role="main"]') || document.body;

        const grids = [];
        const gridEls = [...activePanel.querySelectorAll('[role="grid"], [class*="gridView"]')];

        gridEls.forEach(grid => {{
            const st = window.getComputedStyle(grid);
            if (st.display === 'none' || st.visibility === 'hidden') return;
            if (grid.closest('[aria-hidden="true"]')) return;

            const headers = Array.from(
                grid.querySelectorAll('[role="columnheader"], th, [class*="colHeader"]')
            ).map(h => h.textContent.trim().replace(/\\s+/g, ' ')).filter(h => h);

            const rows = [];
            grid.querySelectorAll('[role="row"]').forEach(row => {{
                const cells = [];
                row.querySelectorAll('[role="gridcell"]').forEach(cell => {{
                    const inp = cell.querySelector('input[type="text"]');
                    const chk = cell.querySelector('input[type="checkbox"]');
                    if (chk) cells.push(chk.checked ? 'Yes' : 'No');
                    else if (inp) cells.push(inp.value);
                    else cells.push(cell.textContent.trim().replace(/\\s+/g, ' '));
                }});
                if (cells.some(c => c.length > 0)) rows.push(cells);
            }});

            if (headers.length > 0 || rows.length > 0) {{
                grids.push({{ headers, rows }});
            }}
        }});
        return grids;
    }}""")


def extract_one_le(page, url):
    """Navigate to url, walk all tabs, extract fields + grids."""
    from playwright.sync_api import TimeoutError as PWTimeout

    print(f"      Navigating: {url[:100]}...")
    page.goto(url)

    # Wait for login/MFA to complete if redirected to Microsoft auth
    for _ in range(120):
        cur = page.url
        if any(x in cur for x in ("login.microsoftonline", "login.live.com", "msft.sts", "/signin")):
            print("      Waiting for login to complete (sign in if prompted)...")
            page.wait_for_timeout(2000)
        else:
            break

    try:
        page.wait_for_selector(
            'main form, [class*="formBody"], [role="main"], [class*="defaultPanel"]',
            timeout=90_000
        )
    except PWTimeout:
        print("      WARNING: timed out waiting for form load")

    wait_ms(page, 2000)

    tabs = get_tab_names(page)
    print(f"      Found {len(tabs)} tabs")

    all_fields = {}
    all_grids  = {}
    tab_data   = {}

    if not tabs:
        # Single-page form — extract without tab context
        fields = extract_fields_js(page)
        grids  = extract_grids_js(page)
        for f in fields:
            key = f["label"]
            all_fields[key] = {"label": f["label"], "value": f["value"], "type": "field", "_tab": ""}
        for gi, g in enumerate(grids):
            gname = f"Grid_{gi}"
            cols  = g.get("headers", [])
            rows  = []
            for row_cells in g.get("rows", []):
                row_dict = {cols[ci]: v for ci, v in enumerate(row_cells) if ci < len(cols)}
                rows.append(row_dict)
            all_grids[gname] = {"columns": [{"label": c, "name": c} for c in cols], "rows": rows}
        tab_data["(Main)"] = {
            "label": "(Main)",
            "fields": {f["label"]: {"label": f["label"], "value": f["value"]} for f in fields},
            "grids": all_grids,
        }
    else:
        for i, tab in enumerate(tabs):
            tab_name  = tab["name"]
            tab_label = tab_name
            print(f"        Tab [{i+1}/{len(tabs)}]: {tab_name}")

            panel_id = click_tab(page, i)
            if not panel_id:
                print(f"        WARNING: could not click tab '{tab_name}'")
                continue

            wait_ms(page, 2500)
            expand_all_fasttabs(page)

            # Scroll active panel to trigger lazy-loaded controls
            pid_scroll = panel_id if panel_id != "__clicked__" else None
            page.evaluate(f"""() => {{
                const panel = {"document.getElementById('" + pid_scroll + "')" if pid_scroll else "null"} ||
                    [...document.querySelectorAll('[role="tabpanel"]')]
                        .find(p => {{
                            if (p.getAttribute('aria-hidden') === 'true') return false;
                            const st = window.getComputedStyle(p);
                            return st.display !== 'none';
                        }}) ||
                    document.querySelector('[class*="formBody"]');
                if (panel) panel.scrollTop = panel.scrollHeight;
            }}""")
            wait_ms(page, 800)
            page.evaluate(f"""() => {{
                const panel = {"document.getElementById('" + pid_scroll + "')" if pid_scroll else "null"} ||
                    [...document.querySelectorAll('[role="tabpanel"]')]
                        .find(p => {{
                            if (p.getAttribute('aria-hidden') === 'true') return false;
                            const st = window.getComputedStyle(p);
                            return st.display !== 'none';
                        }}) ||
                    document.querySelector('[class*="formBody"]');
                if (panel) panel.scrollTop = 0;
            }}""")
            wait_ms(page, 500)

            fields = extract_fields_js(page, panel_id)
            grids  = extract_grids_js(page, panel_id)
            print(f"          panel_id={panel_id!r}  fields={len(fields)}  grids={len(grids)}")

            tab_fields = {}
            for f in fields:
                key = f["label"]
                tab_fields[key] = {"label": f["label"], "value": f["value"]}
                all_fields[key] = {"label": f["label"], "value": f["value"],
                                   "type": "field", "_tab": tab_label}

            tab_grids = {}
            for gi, g in enumerate(grids):
                gname = f"Grid_{gi}"
                cols  = g.get("headers", [])
                rows  = []
                for row_cells in g.get("rows", []):
                    row_dict = {cols[ci]: v for ci, v in enumerate(row_cells) if ci < len(cols)}
                    rows.append(row_dict)
                gdata = {"columns": [{"label": c, "name": c} for c in cols], "rows": rows}
                tab_grids[gname] = gdata
                all_grids[gname] = gdata

            tab_data[tab_name] = {
                "label": tab_label,
                "fields": tab_fields,
                "grids":  tab_grids,
            }

    return {
        "status":     "ok",
        "tabs":       tab_data,
        "all_fields": all_fields,
        "all_grids":  all_grids,
    }


# ── Per-environment extraction ────────────────────────────────────────────────

def extract_env(page, base_url, legal_entities, env_key, env_label):
    """Extract one form from one environment across all LEs using an existing Playwright page."""
    print(f"\n  {'='*56}")
    print(f"  Environment: {env_label} ({env_key})")
    print(f"  {'='*56}")

    results = {}
    for le in legal_entities:
        print(f"\n    --- Legal Entity: {le} ---")
        url = build_url(base_url, le, env_key)
        try:
            results[le] = extract_one_le(page, url)
        except Exception as e:
            print(f"      ERROR extracting {le}: {e}")
            traceback.print_exc()
            results[le] = {"status": f"error: {e}", "tabs": {}, "all_fields": {}, "all_grids": {}}

    return results


# ── Checkpoint helpers ────────────────────────────────────────────────────────
def load_checkpoint():
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT) as f:
            return json.load(f)
    return {}

def save_checkpoint(cp):
    with open(CHECKPOINT, "w") as f:
        json.dump(cp, f, indent=2)


# ── Excel writer (DeepCompare format) ─────────────────────────────────────────
def to_excel_deepcompare(uat_results, config_results, legal_entities,
                          mi_name, ado_label, area_path, out_path):
    """Write comparison Excel in exact DeepCompare format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = openpyxl.Workbook()

    # Styles
    hf          = Font(bold=True, color="FFFFFF", size=10)
    hfill       = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    grid_hfill  = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    tb          = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    bold        = Font(bold=True, size=10)
    match_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    diff_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    only_fill   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    le_sep_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    le_sep_font = Font(bold=True, size=11, color="1F3864")

    def write_header(ws, headers, row=1, fill=None):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hf
            c.fill = fill or hfill
            c.border = tb
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws, ncols, max_row):
        for ci in range(1, ncols + 1):
            mx = 8
            for r in range(1, min(max_row + 1, 200)):
                v = ws.cell(row=r, column=ci).value
                if v:
                    mx = max(mx, min(len(str(v)), 55))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 58)

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    info_rows = [
        ["ADO ID",        ado_label],
        ["Area Path",     area_path],
        ["Form",          mi_name],
        ["UAT Env",       "Env1 UAT (ENV1)"],
        ["Config Env",    "Env4 Config (ENV4)"],
        ["Extracted",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["Legal Entity", "UAT Status", "Config Status",
         "UAT Fields", "Config Fields", "UAT Grids", "Config Grids"],
    ]
    header_idx = 8
    for le in legal_entities:
        uat_le = uat_results.get(le, {})
        cfg_le = config_results.get(le, {})
        info_rows.append([
            le,
            uat_le.get("status", "N/A"),
            cfg_le.get("status", "N/A"),
            len(uat_le.get("all_fields", {})),
            len(cfg_le.get("all_fields", {})),
            len(uat_le.get("all_grids", {})),
            len(cfg_le.get("all_grids", {})),
        ])

    for ri, row_data in enumerate(info_rows, 1):
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = tb
            if ri <= 6:
                c.font = bold
            if ri == header_idx:
                c.font, c.fill = hf, hfill
    for ci in range(1, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # ── Collect all tabs across all LEs/envs ───────────────────────────────────
    all_tabs = {}
    for le in legal_entities:
        for results in (uat_results, config_results):
            le_data = results.get(le, {})
            for tname, tinfo in le_data.get("tabs", {}).items():
                if tname not in all_tabs:
                    all_tabs[tname] = tinfo.get("label", tname)

    # ── Per-tab comparison sheets ──────────────────────────────────────────────
    for tab_name in sorted(all_tabs.keys()):
        tab_label = all_tabs[tab_name]
        sheet_name = tab_label[:28] if len(tab_label) > 28 else tab_label
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:25] + f"_{len(existing)}"

        ws_tab = wb.create_sheet(sheet_name)

        # Union of all fields in this tab
        tab_fields_union = {}
        tab_grids_union  = {}
        for le in legal_entities:
            for results in (uat_results, config_results):
                tinfo = results.get(le, {}).get("tabs", {}).get(tab_name, {})
                for fname, fdata in tinfo.get("fields", {}).items():
                    if fname not in tab_fields_union:
                        tab_fields_union[fname] = fdata.get("label", fname)
                for gname, gdata in tinfo.get("grids", {}).items():
                    if gname not in tab_grids_union:
                        tab_grids_union[gname] = [
                            c.get("label", c.get("name", ""))
                            for c in gdata.get("columns", [])
                        ]

        # Field comparison table
        headers = ["#", "Control", "Label", "UAT Value", "Config Value", "Match", "LE"]
        write_header(ws_tab, headers)
        ri = 2
        seq = 0

        for le in legal_entities:
            uat_fields = uat_results.get(le, {}).get("tabs", {}).get(tab_name, {}).get("fields", {})
            cfg_fields = config_results.get(le, {}).get("tabs", {}).get(tab_name, {}).get("fields", {})

            # LE separator
            for ci in range(1, len(headers) + 1):
                c = ws_tab.cell(row=ri, column=ci)
                c.fill = le_sep_fill
                c.border = tb
            ws_tab.cell(row=ri, column=1, value=le).font = le_sep_font
            ws_tab.cell(row=ri, column=1).fill = le_sep_fill
            ws_tab.cell(row=ri, column=2,
                        value=f"UAT: {len(uat_fields)} | Config: {len(cfg_fields)} fields"
                        ).font = le_sep_font
            ws_tab.cell(row=ri, column=2).fill = le_sep_fill
            ri += 1

            for fname in list(tab_fields_union.keys()):
                uat_val = str(uat_fields.get(fname, {}).get("value", "")) if fname in uat_fields else ""
                cfg_val = str(cfg_fields.get(fname, {}).get("value", "")) if fname in cfg_fields else ""
                label   = tab_fields_union[fname]

                if not uat_val and not cfg_val:
                    continue

                seq += 1
                if uat_val == cfg_val:
                    match, mfill = "Match",       match_fill
                elif uat_val and cfg_val:
                    match, mfill = "DIFF",        diff_fill
                elif uat_val and not cfg_val:
                    match, mfill = "UAT only",    only_fill
                else:
                    match, mfill = "Config only", only_fill

                ws_tab.cell(row=ri, column=1, value=seq).border = tb
                ws_tab.cell(row=ri, column=2, value=fname).border = tb
                ws_tab.cell(row=ri, column=3, value=label).border = tb
                ws_tab.cell(row=ri, column=4, value=uat_val).border = tb
                ws_tab.cell(row=ri, column=5, value=cfg_val).border = tb
                mc = ws_tab.cell(row=ri, column=6, value=match)
                mc.border = tb
                mc.fill   = mfill
                ws_tab.cell(row=ri, column=7, value=le).border = tb
                ri += 1

        auto_width(ws_tab, len(headers), ri)
        ws_tab.freeze_panes = "B2"
        if ri > 2:
            ws_tab.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ri - 1}"

        # Sub-grid sheets for this tab
        for gname, col_labels in tab_grids_union.items():
            if not col_labels:
                continue
            g_sheet_name = f"{sheet_name[:18]}_{gname[:8]}_Grid"
            existing_names = [s.title for s in wb.worksheets]
            if g_sheet_name in existing_names:
                g_sheet_name = g_sheet_name[:25] + f"_{len(existing_names)}"
            ws_g = wb.create_sheet(g_sheet_name)

            g_headers = ["#"] + col_labels + ["Env", "LE"]
            write_header(ws_g, g_headers, fill=grid_hfill)
            gri  = 2
            gseq = 0

            for le in legal_entities:
                le_rows = []
                for env_lbl, results in [("UAT", uat_results), ("Config", config_results)]:
                    gdata = results.get(le, {}).get("tabs", {}).get(tab_name, {}).get("grids", {}).get(gname, {})
                    for row in gdata.get("rows", []):
                        le_rows.append((env_lbl, row))

                if not le_rows:
                    continue

                for ci in range(1, len(g_headers) + 1):
                    c = ws_g.cell(row=gri, column=ci)
                    c.fill = le_sep_fill
                    c.border = tb
                ws_g.cell(row=gri, column=1, value=le).font = le_sep_font
                ws_g.cell(row=gri, column=1).fill = le_sep_fill
                gri += 1

                for env_lbl, row_data in le_rows:
                    gseq += 1
                    ws_g.cell(row=gri, column=1, value=gseq).border = tb
                    for ci, cl in enumerate(col_labels, 2):
                        ws_g.cell(row=gri, column=ci, value=row_data.get(cl, "")).border = tb
                    ws_g.cell(row=gri, column=len(col_labels) + 2, value=env_lbl).border = tb
                    ws_g.cell(row=gri, column=len(col_labels) + 3, value=le).border = tb
                    gri += 1

            auto_width(ws_g, len(g_headers), gri)
            ws_g.freeze_panes = "B2"

    # ── All Controls flat sheet ────────────────────────────────────────────────
    ws_all    = wb.create_sheet("All Controls")
    all_hdrs  = ["#", "Control", "Label", "Value", "Type", "Tab", "Env", "LE"]
    write_header(ws_all, all_hdrs)
    ari  = 2
    aseq = 0

    for le in legal_entities:
        for env_lbl, results in [("UAT", uat_results), ("Config", config_results)]:
            le_data = results.get(le, {})
            for fname, fdata in sorted(le_data.get("all_fields", {}).items()):
                aseq += 1
                ws_all.cell(row=ari, column=1, value=aseq).border = tb
                ws_all.cell(row=ari, column=2, value=fname).border = tb
                ws_all.cell(row=ari, column=3, value=fdata.get("label", "")).border = tb
                ws_all.cell(row=ari, column=4, value=str(fdata.get("value", ""))).border = tb
                ws_all.cell(row=ari, column=5, value=fdata.get("type", "")).border = tb
                ws_all.cell(row=ari, column=6, value=fdata.get("_tab", "")).border = tb
                ws_all.cell(row=ari, column=7, value=env_lbl).border = tb
                ws_all.cell(row=ari, column=8, value=le).border = tb
                ari += 1

    auto_width(ws_all, len(all_hdrs), ari)
    ws_all.freeze_panes = "B2"
    if ari > 2:
        ws_all.auto_filter.ref = f"A1:{get_column_letter(len(all_hdrs))}{ari - 1}"

    # ── Also add top-level (non-tab) grids sheet if any ───────────────────────
    tabbed_grids = set()
    for le in legal_entities:
        for results in (uat_results, config_results):
            for tinfo in results.get(le, {}).get("tabs", {}).values():
                tabbed_grids.update(tinfo.get("grids", {}).keys())

    toplevel_grids = {}
    for le in legal_entities:
        for results in (uat_results, config_results):
            for gname, gdata in results.get(le, {}).get("all_grids", {}).items():
                if gname not in tabbed_grids and gname not in toplevel_grids:
                    toplevel_grids[gname] = [
                        c.get("label", c.get("name", ""))
                        for c in gdata.get("columns", [])
                    ]

    for gname, col_labels in toplevel_grids.items():
        if not col_labels:
            continue
        g_sheet = f"{gname[:24]}_Grid"
        existing = [s.title for s in wb.worksheets]
        if g_sheet in existing:
            g_sheet = g_sheet[:25] + f"_{len(existing)}"
        ws_g = wb.create_sheet(g_sheet)
        g_headers = ["#"] + col_labels + ["Env", "LE"]
        write_header(ws_g, g_headers, fill=grid_hfill)
        gri = 2; gseq = 0

        for le in legal_entities:
            le_rows = []
            for env_lbl, results in [("UAT", uat_results), ("Config", config_results)]:
                gdata = results.get(le, {}).get("all_grids", {}).get(gname, {})
                for row in gdata.get("rows", []):
                    le_rows.append((env_lbl, row))
            if not le_rows:
                continue
            for ci in range(1, len(g_headers) + 1):
                c = ws_g.cell(row=gri, column=ci)
                c.fill = le_sep_fill; c.border = tb
            ws_g.cell(row=gri, column=1, value=le).font = le_sep_font
            ws_g.cell(row=gri, column=1).fill = le_sep_fill
            gri += 1
            for env_lbl, row_data in le_rows:
                gseq += 1
                ws_g.cell(row=gri, column=1, value=gseq).border = tb
                for ci, cl in enumerate(col_labels, 2):
                    ws_g.cell(row=gri, column=ci, value=row_data.get(cl, "")).border = tb
                ws_g.cell(row=gri, column=len(col_labels) + 2, value=env_lbl).border = tb
                ws_g.cell(row=gri, column=len(col_labels) + 3, value=le).border = tb
                gri += 1
        auto_width(ws_g, len(g_headers), gri)
        ws_g.freeze_panes = "B2"

    wb.save(out_path)
    print(f"  Saved: {os.path.basename(out_path)}")
    return out_path


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    import argparse
    from playwright.sync_api import sync_playwright

    parser = argparse.ArgumentParser()
    parser.add_argument("--only", help="Run only this MI name (e.g. EntAssetParameters)")
    parser.add_argument("--le", help="Comma-separated LEs to run (e.g. MY30 or MY30,SG60)")
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    cp = load_checkpoint()

    # Determine which LEs to process
    legal_entities = LES
    if args.le:
        legal_entities = [x.strip().upper() for x in args.le.split(",") if x.strip()]
        print(f"Restricting to LEs: {legal_entities}")

    print(f"\nReading Correct_URL from: {os.path.basename(MAPPING_FILE)}")
    items, skipped = load_items_from_mapping()

    if skipped:
        print(f"WARNING — {len(skipped)} items have no Correct_URL, skipped:")
        for ado_id, title in skipped:
            print(f"  ADO {ado_id}: {title}")

    groups = group_by_mi(items)

    if args.only:
        groups = [(mi, ids, title, url) for mi, ids, title, url in groups
                  if mi.lower() == args.only.lower()]
        if not groups:
            print(f"ERROR: MI '{args.only}' not found in mapping. Check spelling.")
            return

    total  = len(groups)
    print(f"\n{'='*70}")
    print(f" Asset & FA Batch Deep-Extraction — {total} unique forms, {len(items)} ADO items")
    print(f" Output : {OUTPUT_DIR}")
    print(f" Format : DeepCompare (per-tab comparison + All Controls)")
    print(f" Engine : Playwright CDP -> http://localhost:9222")
    print(f"{'='*70}\n")

    done   = 0
    errors = []

    with sync_playwright() as pw:
        # Try to reuse already-open Chrome first
        try:
            browser = pw.chromium.connect_over_cdp("http://localhost:9222")
            ctx  = browser.contexts[0]
            page = ctx.new_page()
            print("Connected to existing Chrome via CDP.")
        except Exception:
            # Chrome not open with CDP — launch it with real profile + CDP port
            print(f"\nLaunching Chrome with your existing profile...")
            print("  Close other Chrome windows first if you see a profile-locked error.")
            ctx = pw.chromium.launch_persistent_context(
                CHROME_PROFILE,
                channel="chrome",
                headless=False,
                viewport={"width": 1600, "height": 900},
                args=[
                    "--remote-debugging-port=9222",
                    "--profile-directory=Default",
                    "--disable-blink-features=AutomationControlled",
                ],
            )
            page = ctx.new_page()
            print("Chrome launched.")
        page.set_viewport_size({"width": 1600, "height": 900})

        for idx, (mi, ado_ids, title, base_url) in enumerate(groups, 1):
            cp_key    = mi
            ado_label = "_".join(str(a) for a in ado_ids)

            if cp.get(cp_key) == "done":
                print(f"[{idx}/{total}] SKIP (done): {mi}")
                done += 1
                continue

            print(f"\n[{idx}/{total}] ADO {ado_label.replace('_','/')}  |  MI: {mi}")
            print(f"  Path: {title}")

            try:
                print(f"\n  >>> Extracting UAT (ENV1)...")
                uat_results = extract_env(page, base_url, legal_entities, "uat", "Env1 UAT")

                print(f"\n  >>> Extracting Config (ENV4)...")
                cfg_results = extract_env(page, base_url, legal_entities, "config", "Env4 Config")

                ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
                fname  = f"DeepCompare_{mi}_{ado_label}_{ts}.xlsx"
                out    = os.path.join(OUTPUT_DIR, fname)

                print(f"\n  >>> Writing Excel...")
                to_excel_deepcompare(
                    uat_results, cfg_results, legal_entities,
                    mi_name   = mi,
                    ado_label = ado_label.replace("_", "/"),
                    area_path = title,
                    out_path  = out,
                )

                cp[cp_key] = "done"
                save_checkpoint(cp)
                done += 1

            except Exception as e:
                msg = f"ERROR [{ado_label}] {mi}: {e}"
                print(f"  ERROR: {msg}")
                traceback.print_exc()
                errors.append(msg)
                cp[cp_key] = f"error: {str(e)[:200]}"
                save_checkpoint(cp)
                time.sleep(2)

        page.close()

    print(f"\n{'='*70}")
    print(f" DONE: {done}/{total} forms extracted")
    if errors:
        print(f" ERRORS ({len(errors)}):")
        for e in errors:
            print(f"   - {e}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
