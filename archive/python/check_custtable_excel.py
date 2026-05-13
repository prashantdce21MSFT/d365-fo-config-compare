"""Check what the Excel shows for customer 210205."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import glob, os

# Find the latest CustTable Excel
pattern = r"C:\D365 Configuration Drift Analysis\output\Comparison\Compare_CustTable_*.xlsx"
files = sorted(glob.glob(pattern))
if not files:
    print("No CustTable Excel found!")
    sys.exit(1)

latest = files[-1]
print(f"Reading: {latest}\n")

wb = openpyxl.load_workbook(latest, read_only=True)
for sheet_name in wb.sheetnames:
    if sheet_name == "Summary":
        continue
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")

    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value or "")

    # Find column indices for key fields
    acct_cols = [i for i, h in enumerate(headers) if h and ("account" in h.lower() or "customer" in h.lower())]
    consign_cols = [i for i, h in enumerate(headers) if h and "consignment" in h.lower()]
    warehouse_cols = [i for i, h in enumerate(headers) if h and h.lower() == "[d] (fasttab) > warehouse"]
    env_cols = [i for i, h in enumerate(headers) if h and h.lower() == "env"]
    le_cols = [i for i, h in enumerate(headers) if h and h.lower() == "le"]

    print(f"  Headers ({len(headers)} cols): {headers[:10]}...")
    print(f"  Account cols: {[(i, headers[i]) for i in acct_cols]}")
    print(f"  Consignment cols: {[(i, headers[i]) for i in consign_cols]}")
    print(f"  Warehouse cols: {[(i, headers[i]) for i in warehouse_cols]}")
    print(f"  Env cols: {[(i, headers[i]) for i in env_cols]}")

    # Search for customer 210205
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        # Check if 210205 appears in any cell
        for v in vals:
            if str(v or "").strip() == "210205":
                found = True
                print(f"\n  FOUND 210205 at row {row[0].row}:")
                for i, h in enumerate(headers):
                    v = vals[i] if i < len(vals) else ""
                    if v and str(v).strip():
                        print(f"    {h} = {v}")
                break
        if found:
            break

    if not found:
        # Show first few MY30 UAT rows
        print("\n  210205 NOT found. Showing first 5 data rows:")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            vals = [cell.value for cell in row]
            if not vals[0]:
                continue
            # Skip separator rows
            try:
                int(vals[0])
            except:
                continue
            count += 1
            if count > 5:
                break
            key_info = []
            for i in acct_cols[:1]:
                key_info.append(f"Acct={vals[i]}")
            for i in consign_cols[:1]:
                key_info.append(f"Consignment={vals[i]}")
            for i in env_cols[:1]:
                key_info.append(f"Env={vals[i]}")
            for i in le_cols[:1]:
                key_info.append(f"LE={vals[i]}")
            print(f"    Row {row[0].row}: {', '.join(key_info)}")

wb.close()
print("\nDone.")
