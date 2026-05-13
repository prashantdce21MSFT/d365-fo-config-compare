from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

# ── 1. Load CDD_FormMapping ───────────────────────────────────────────────────
wb1 = load_workbook('output/CDD_FormMapping.xlsx', read_only=True, data_only=True)
ws1 = wb1['CDD Form Mapping']
cdd_rows = []
for r in ws1.iter_rows(values_only=True):
    if r[0] is not None and isinstance(r[0], int):
        cdd_rows.append({
            'ado_id':     r[0],
            'title':      r[1] or '',
            'form_name':  r[2] or '',
            'mi_name':    r[3] or '',
            'confidence': r[4] or '',
            'state':      r[5] or '',
            'type':       r[6] or '',
        })
wb1.close()
print(f'CDD rows loaded: {len(cdd_rows)}')

# ── 2. Load CDD_Form_Tables ───────────────────────────────────────────────────
wb2 = load_workbook('output/CDD_Form_Tables.xlsx', read_only=True, data_only=True)
ws2 = wb2['Form \u2192 Tables']
tables_map = {}
header_passed = False
for r in ws2.iter_rows(values_only=True):
    if not header_passed:
        if r[0] == 'MI Name (Form)':
            header_passed = True
        continue
    if r[0] and isinstance(r[0], str) and r[0].strip():
        mi = r[0].strip()
        tables_map[mi] = {
            'form_type':     r[3] or '',
            'cba_extended':  r[4] or '',
            'cdd_count':     r[5] or 0,
            'table_count':   r[6] or 0,
            'primary_table': r[7] or '',
            'all_tables':    r[8] or '',
            'source':        r[9] or '',
        }
wb2.close()
print(f'Forms in tables map: {len(tables_map)}')

# ── 3. Merge ──────────────────────────────────────────────────────────────────
merged = []
for c in cdd_rows:
    mi = c['mi_name'].strip()
    t  = tables_map.get(mi, {})
    merged.append({**c, **t})
print(f'Merged rows: {len(merged)}')

# ── Styles ────────────────────────────────────────────────────────────────────
hdr_font  = Font(bold=True, color="FFFFFF", size=10)
hdr_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
orange    = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
blue_fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
thin      = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'),  bottom=Side(style='thin'))

def row_fill(form_type, confidence):
    if 'Custom CBA' in (form_type or ''):
        return orange
    if confidence == 'H':
        return green
    if confidence == 'M':
        return yellow
    return red_fill

# ── 4. Build Excel ────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: CDD → Form → Tables (all 606 CDDs, one row each) ────────────────
ws = wb.active
ws.title = "CDD to Form to Tables"

ws.merge_cells('A1:N1')
ws['A1'] = (f"MY/SG Configuration Deliverables  —  CDD  ›  Form  ›  D365 Tables  |  "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(merged)} CDDs")
ws['A1'].font = Font(bold=True, size=13)

headers = [
    "ADO ID", "CDD Title", "Form Display Name", "MI Name (Form)",
    "Confidence", "State", "CDD Type",
    "Form Type", "CBA Extended?",
    "# Tables", "Primary Table", "All Tables (pipe-sep)",
    "Table Source", "# CDDs on Form"
]
HR = 3
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=HR, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = thin
    c.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, m in enumerate(merged, HR + 1):
    vals = [
        m['ado_id'], m['title'], m['form_name'], m['mi_name'],
        m['confidence'], m['state'], m['type'],
        m.get('form_type', ''), m.get('cba_extended', ''),
        m.get('table_count', ''), m.get('primary_table', ''),
        m.get('all_tables', ''), m.get('source', ''), m.get('cdd_count', ''),
    ]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.border = thin
        c.alignment = Alignment(vertical='top', wrap_text=(ci in (2, 12)))
        c.font = Font(size=9)
    fill = row_fill(m.get('form_type'), m.get('confidence'))
    for ci in range(1, len(headers) + 1):
        ws.cell(row=ri, column=ci).fill = fill
    if m.get('cba_extended') == 'YES':
        ws.cell(row=ri, column=9).fill = blue_fill

col_widths = [9, 72, 38, 32, 11, 28, 18, 24, 13, 9, 30, 65, 12, 14]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.auto_filter.ref = f"A{HR}:{get_column_letter(len(headers))}{HR + len(merged)}"
ws.freeze_panes = f"A{HR + 1}"
ws.row_dimensions[HR].height = 30

# ── Sheet 2: By Form (one row per unique form, CDD IDs listed) ───────────────
ws2 = wb.create_sheet("By Form")
ws2.merge_cells('A1:J1')
ws2['A1'] = "One row per form — shows all CDD IDs that reference each form"
ws2['A1'].font = Font(bold=True, size=12)

h2 = ["MI Name (Form)", "Form Display Name", "Confidence", "Form Type",
      "CBA Extended?", "# CDDs", "# Tables", "Primary Table",
      "All Tables (pipe-sep)", "CDD IDs"]
for ci, h in enumerate(h2, 1):
    c = ws2.cell(row=3, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = thin
    c.alignment = Alignment(horizontal='center', wrap_text=True)

form_groups = defaultdict(list)
form_meta   = {}
for m in merged:
    mi = m['mi_name'] or '(blank)'
    form_groups[mi].append(m['ado_id'])
    if mi not in form_meta:
        form_meta[mi] = m

sorted_forms = sorted(form_meta.keys(), key=lambda x: (
    0 if 'Custom CBA' in form_meta[x].get('form_type', '') else 1, x))

for ri, mi in enumerate(sorted_forms, 4):
    m    = form_meta[mi]
    ids  = ', '.join(str(i) for i in sorted(form_groups[mi]))
    vals = [
        mi, m.get('form_name', ''), m.get('confidence', ''),
        m.get('form_type', ''), m.get('cba_extended', ''),
        len(form_groups[mi]), m.get('table_count', ''),
        m.get('primary_table', ''), m.get('all_tables', ''), ids,
    ]
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.border = thin
        c.alignment = Alignment(vertical='top', wrap_text=(ci in (9, 10)))
        c.font = Font(size=9)
    fill = row_fill(m.get('form_type'), m.get('confidence'))
    for ci in range(1, len(h2) + 1):
        ws2.cell(row=ri, column=ci).fill = fill
    if m.get('cba_extended') == 'YES':
        ws2.cell(row=ri, column=5).fill = blue_fill

w2 = [32, 38, 11, 24, 13, 9, 9, 30, 65, 60]
for ci, w in enumerate(w2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.auto_filter.ref = f"A3:{get_column_letter(len(h2))}{3 + len(sorted_forms)}"
ws2.freeze_panes = "A4"

# ── Sheet 3: Summary ──────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Summary")
ws3['A1'] = "Summary"; ws3['A1'].font = Font(bold=True, size=14)

total_cdds   = len(merged)
unique_forms = len(form_groups)
with_tables  = sum(1 for m in merged if m.get('all_tables'))
cba_cdds     = sum(1 for m in merged if 'Custom CBA' in m.get('form_type', ''))
std_cdds     = total_cdds - cba_cdds
h_conf       = sum(1 for m in merged if m.get('confidence') == 'H')
m_conf       = sum(1 for m in merged if m.get('confidence') == 'M')
l_conf       = sum(1 for m in merged if m.get('confidence') == 'L')

summary_rows = [
    ("Total CDDs", total_cdds),
    ("Unique forms", unique_forms),
    ("CDDs with table mapping", with_tables),
    ("", ""),
    ("CBA custom form CDDs", cba_cdds),
    ("Standard D365 form CDDs", std_cdds),
    ("", ""),
    ("High confidence (H)", h_conf),
    ("Medium confidence (M)", m_conf),
    ("Low confidence (L)", l_conf),
]
for ri, (lbl, val) in enumerate(summary_rows, 3):
    ws3.cell(row=ri, column=1, value=lbl).font = Font(bold=bool(lbl))
    ws3.cell(row=ri, column=2, value=val)
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 12

# ── Save ──────────────────────────────────────────────────────────────────────
out = r'output\CDD_Form_Tables_Merged.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'CDDs: {total_cdds} | Unique forms: {unique_forms} | CBA: {cba_cdds} | Std: {std_cdds}')
print(f'Confidence  H:{h_conf}  M:{m_conf}  L:{l_conf}')
