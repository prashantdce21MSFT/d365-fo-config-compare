"""Generate Excel with all 606 CDDs, form paths, and screenshot paths."""
import json
import os
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], shell=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output"
SS_DIR = os.path.join(OUTPUT_DIR, "screenshots")

# Company mapping
_COMPANY_MAP = {"MY": "MY30", "SG": "SG60"}
_SKIP_TAGS = {"export", "import", "empties", "cutover"}

def extract_company(title):
    for m in re.finditer(r'\[([A-Za-z0-9]{2,6})\]', title):
        tag = m.group(1)
        if tag.lower() in _SKIP_TAGS or tag.lower().startswith("fdd"):
            continue
        return _COMPANY_MAP.get(tag, tag)
    return ""

def extract_nav_path(title):
    """Extract D365 navigation path from title. Many titles ARE the path."""
    if ">" in title:
        return title.strip()
    return ""

def get_screenshot_paths(ado_id):
    """Find all screenshot files for an ADO ID."""
    paths = []
    if not os.path.isdir(SS_DIR):
        return paths
    for f in sorted(os.listdir(SS_DIR)):
        if f.startswith(f"{ado_id}_img") and f.endswith(".png"):
            paths.append(os.path.join(SS_DIR, f))
    return paths

def main():
    with open(os.path.join(OUTPUT_DIR, "all_ado_items.json")) as f:
        items = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "CDD Form Paths"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(size=11, italic=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Title rows
    ws.merge_cells('A1:H1')
    ws['A1'] = "MY/SG Configuration Deliverables — Form Path Mapping"
    ws['A1'].font = title_font
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = subtitle_font

    # Summary
    total = len(items)
    with_path = sum(1 for i in items if i["nav_path"])
    with_ss = sum(1 for i in items if i["has_screenshots"])
    no_path_no_ss = sum(1 for i in items if not i["nav_path"] and not i["has_screenshots"])

    ws['A4'] = "Total CDDs:"
    ws['B4'] = total
    ws['B4'].font = Font(bold=True)
    ws['A5'] = "With nav path in title:"
    ws['B5'] = with_path
    ws['B5'].fill = green_fill
    ws['A6'] = "With screenshots:"
    ws['B6'] = with_ss
    ws['A7'] = "No path & no screenshots:"
    ws['B7'] = no_path_no_ss
    ws['B7'].fill = red_fill

    # Headers
    headers = [
        "ADO ID", "Title", "Navigation Path (from title)", 
        "Area Path", "Company", "State",
        "Has Screenshots", "Screenshot Count", 
        "Screenshot Path (from images)"
    ]
    header_row = 9
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, item in enumerate(items, header_row + 1):
        ado_id = item["ado_id"]
        title = item["title"]
        nav_path = item["nav_path"]
        area = item["area_path"]
        company = item["company"] or extract_company(title)
        state = item["state"]
        has_ss = item["has_screenshots"]
        ss_count = item["screenshot_count"]
        
        # Get screenshot file paths
        ss_paths = get_screenshot_paths(ado_id)
        ss_path_str = "\n".join(ss_paths) if ss_paths else ""

        vals = [ado_id, title, nav_path, area, company, state,
                "Yes" if has_ss else "No", ss_count, ss_path_str]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 9))

        # Color coding
        if not nav_path and not has_ss:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill
        elif not nav_path and has_ss:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = yellow_fill

    # Column widths
    widths = [10, 80, 60, 40, 10, 15, 14, 14, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:I{header_row + len(items)}"

    # Freeze panes
    ws.freeze_panes = f"A{header_row + 1}"

    # Save
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = os.path.join(OUTPUT_DIR, f"CDD_Form_Paths_{ts}.xlsx")
    wb.save(out_file)
    print(f"Saved: {out_file}")
    print(f"Total rows: {len(items)}")

if __name__ == "__main__":
    main()
