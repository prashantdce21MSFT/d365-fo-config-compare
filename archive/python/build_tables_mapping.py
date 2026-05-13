"""
Build Form -> D365 Tables mapping for all 606 CDD items.
Sources: (1) CB_CBA_Customisation_Analysis_v2.xlsx for CBA forms,
         (2) D365 F&O knowledge for standard forms.
Output: output/CDD_Form_Tables.xlsx
"""
import json, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook as lw

# ── 1. Read CBA datasources from reference file ────────────────────────────
REF_FILE = r"C:\Users\prverma\Downloads\CB_CBA_Customisation_Analysis_v2.xlsx"
wb_ref = lw(REF_FILE, read_only=True, data_only=True)
ws_ref = wb_ref['1. Custom Forms']
rows_ref = list(ws_ref.iter_rows(values_only=True))
CBA_DS = {}
for r in rows_ref[1:]:
    if not r[2]: continue
    fn = str(r[2]).strip()
    ds = [x.strip() for x in str(r[4] or '').split('\n') if x.strip()]
    CBA_DS[fn] = ds

# Also read Form Extensions to flag standard forms that have CBA extensions
ws_ext = wb_ref['2. Form Extensions']
rows_ext = list(ws_ext.iter_rows(values_only=True))
CBA_EXTENDED = set()
for r in rows_ext[1:]:
    if r[2]: CBA_EXTENDED.add(str(r[2]).strip())

wb_ref.close()

# ── 2. Standard D365 form -> tables mapping ────────────────────────────────
STD_DS = {
    # Asset Management
    'EntAssetParameters':                   ['EntAssetParameters'],
    'EntAssetObjectType':                    ['EntAssetObjectType','EntAssetObjectTypeLifecycleState'],
    'EntAssetObjectTypeDefault':             ['EntAssetObjectTypeDefault'],
    'EntAssetCriticality':                   ['EntAssetCriticality'],
    'EntAssetCriticalityType':               ['EntAssetCriticalityType'],
    'EntAssetServiceLevel':                  ['EntAssetServiceLevel'],
    'EntAssetCounter':                       ['EntAssetCounter','EntAssetCounterTrans'],
    'EntAssetLifecycleState':                ['EntAssetLifecycleState'],
    'EntAssetLifecycleModel':                ['EntAssetLifecycleModel','EntAssetLifecycleState'],
    'EntAssetFaultType':                     ['EntAssetFaultType'],
    'EntAssetFaultArea':                     ['EntAssetFaultArea'],
    'EntAssetFaultCause':                    ['EntAssetFaultCause'],
    'EntAssetFaultSymptom':                  ['EntAssetFaultSymptom'],
    'EntAssetFaultRemedy':                   ['EntAssetFaultRemedy'],
    'EntAssetFaultDesigner':                 ['EntAssetFaultDesigner'],
    'EntAssetDiagnosticType':                ['EntAssetDiagnosticType'],
    'EntAssetMaintenanceJobType':            ['EntAssetMaintenanceJobType','EntAssetMaintenanceJobTypeSetup'],
    'EntAssetMaintenanceJobTypeCategory':    ['EntAssetMaintenanceJobTypeCategory'],
    'EntAssetMaintenanceJobTypeTrade':       ['EntAssetMaintenanceJobTypeTrade'],
    'EntAssetMaintenanceJobTypeVariant':     ['EntAssetMaintenanceJobTypeVariant'],
    'EntAssetMaintenancePlan':               ['EntAssetMaintenancePlan','EntAssetMaintenancePlanLine'],
    'EntAssetMaintenanceRequestType':        ['EntAssetMaintenanceRequestType'],
    'EntAssetMaintenanceDowntimeReason':     ['EntAssetMaintenanceDowntimeReason'],
    'EntAssetMaintenanceChecklistTemplate':  ['EntAssetMaintenanceChecklistTemplate','EntAssetMaintenanceChecklistLine'],
    'EntAssetMaintenanceChecklistVariable':  ['EntAssetMaintenanceChecklistVariable','EntAssetMaintenanceChecklistVariableLine'],
    'EntAssetMaintenanceResponsibleWorkers': ['EntAssetMaintenanceResponsibleWorkers'],
    'EntAssetWorkerGroup':                   ['EntAssetWorkerGroup','EntAssetWorkerGroupLine'],
    'EntAssetWorkerPreferred':               ['EntAssetWorkerPreferred'],
    'EntAssetWorkOrderType':                 ['EntAssetWorkOrderType'],
    'EntAssetScheduledExecution':            ['EntAssetScheduledExecution'],
    'EntAssetSparePartChangeReason':         ['EntAssetSparePartChangeReason'],
    'EntAssetSpareParts':                    ['EntAssetSpareParts'],
    'EntAssetManufacturerModel':             ['EntAssetManufacturer','EntAssetManufacturerModel'],
    'EntAssetMajorType':                     ['EntAssetMajorType'],
    'EntAssetFunctionalLocationType':        ['EntAssetFunctionalLocationType'],
    'EntAssetConditionAssessmentTemplate':   ['EntAssetConditionAssessmentTemplate','EntAssetConditionAssessmentLine'],
    # General Ledger
    'LedgerParameters':                ['LedgerParameters','LedgerParametersDimAttribute'],
    'LedgerChartOfAccounts':           ['LedgerChartOfAccounts','LedgerChartOfAccountsTranslation'],
    'LedgerTable':                     ['LedgerTable','LedgerFiscalCalendarPeriod'],
    'MainAccount':                     ['MainAccount','MainAccountLegalEntityOverride','MainAccountCategory'],
    'MainAccountCategory':             ['MainAccountCategory'],
    'LedgerJournalName':               ['LedgerJournalName'],
    'LedgerAccrualTable':              ['LedgerAccrualTable','LedgerAccrualTrans'],
    'LedgerSystemAccounts':            ['LedgerSystemAccounts'],
    'LedgerDimensionLink':             ['LedgerDimensionLink'],
    'LedgerForeignCurrencyRevaluation':['LedgerForeignCurrencyRevaluation'],
    'LedgerIntercompany':              ['LedgerIntercompany','LedgerIntercompanyAccount'],
    'LedgerNetting':                   ['LedgerNetting','LedgerNettingLine'],
    'LedgerPeriodTemplate':            ['LedgerPeriodTemplate','LedgerPeriodTemplateLine'],
    'LedgerSettlement':                ['LedgerSettlement'],
    'LedgerTransactionCode':           ['LedgerTransactionCode'],
    'LedgerPostingDefinition':         ['LedgerPostingDefinition','LedgerPostingDefinitionEntry'],
    'FiscalCalendar':                  ['FiscalCalendar','FiscalCalendarYear','FiscalCalendarPeriod'],
    'DimensionAttribute':              ['DimensionAttribute','DimensionAttributeValue','DimensionAttributeValueTranslation'],
    'DimensionAttributeValueTranslation':['DimensionAttributeValueTranslation','DimensionAttribute'],
    'DimensionHierarchy':              ['DimensionHierarchy','DimensionHierarchyLevel','DimensionHierarchyNode'],
    'DimensionRuleStructure':          ['DimensionRuleStructure','DimensionRule'],
    'DimensionSet':                    ['DimensionSet','DimensionSetSegment'],
    'DimensionIntegrationConfiguration':['DimensionIntegrationConfiguration','DimensionIntegrationField'],
    'ReasonTable':                     ['ReasonTable'],
    # Accounts Receivable
    'CustParameters':               ['CustParameters','NumberSequenceReference'],
    'CustGroup':                    ['CustGroup'],
    'CustPostingProfile':           ['CustPostingProfile','CustPostingProfileLine'],
    'CustPool':                     ['CustPool'],
    'CustPaymMode':                 ['CustPaymMode'],
    'CustInterestTable':            ['CustInterestTable','CustInterestLine'],
    'CustCollectionLetterLine':     ['CustCollectionLetterTable','CustCollectionLetterLine'],
    'CustVendAgePeriodDef':         ['CustVendAgePeriodDef'],
    'CustVendBlockingCause':        ['CustVendBlockingCause'],
    'CustBlockingRule':             ['CustBlockingRule'],
    'CustWriteOffReason':           ['CustWriteOffReason'],
    'CustCreditGroup':              ['CustCreditGroup'],
    'CustCreditMgmtReason':         ['CustCreditMgmtReason'],
    'CustCreditMgmtWorkflow':       ['CustCreditMgmtWorkflow'],
    'CustCreditMgmtAccountStatus':  ['CustCreditMgmtAccountStatus'],
    'CustStatisticsGroup':          ['CustStatisticsGroup'],
    'CustAdvanceInvoice':           ['CustAdvanceInvoiceTmp'],
    'CustPriceGroup':               ['CustPriceGroup'],
    'CustTable':                    ['CustTable','DirParty','LogisticsPostalAddress'],
    'SumUpdateParameters':          ['SumUpdateParameters'],
    # Accounts Payable
    'VendParameters':               ['VendParameters','NumberSequenceReference'],
    'VendGroup':                    ['VendGroup'],
    'VendPostingProfile':           ['VendPostingProfile','VendPostingProfileLine'],
    'VendEditInvoice':              ['VendInvoiceInfoTable','VendInvoiceInfoLine','VendInvoiceInfoSubLine'],
    'VendRequestVendorUpdate':      ['VendRequestVendorUpdate'],
    'VendVendorEmailContactInfoApprovalWorkflow':['VendVendorContactInfoApprovalWorkflow'],
    'VendVendorPortalFieldValidation':['VendVendorPortalFieldValidation'],
    # Cash & Bank
    'BankParameters':               ['BankParameters'],
    'BankGroup':                    ['BankGroup'],
    'BankTransType':                ['BankTransType'],
    'BankStatementFormat':          ['BankStatementFormat'],
    'BankPaymentPurposeCode':       ['BankPaymentPurposeCode'],
    'BankPositivePayFormat':        ['BankPositivePayFormat'],
    'BankReconciliationMatchRule':  ['BankReconciliationMatchRule','BankReconciliationMatchRuleLine'],
    'BankReconciliationMatchRuleSet':['BankReconciliationMatchRuleSet','BankReconciliationMatchRuleSetLine'],
    'BankLCPostingProfile':         ['BankLCPostingProfile'],
    # Fixed Assets
    'AssetParameters':              ['AssetParameters'],
    'AssetGroup':                   ['AssetGroup','AssetGroupBook'],
    'AssetBook':                    ['AssetBook','AssetBookTable'],
    'AssetDepreciationProfile':     ['AssetDepreciationProfile'],
    'AssetLedgerAccounts':          ['AssetLedgerAccounts'],
    # Tax
    'TaxTable':                     ['TaxTable','TaxData','TaxGroupData'],
    'TaxData':                      ['TaxData','TaxTable'],
    'TaxGroupData':                 ['TaxGroupData','TaxGroup'],
    'TaxItemGroup':                 ['TaxItemGroup','TaxItemGroupHeading'],
    'TaxSetup':                     ['TaxSetup'],
    'TaxPeriodHead':                ['TaxPeriodHead','TaxPeriod'],
    'TaxAuthority':                 ['TaxAuthority'],
    'TaxRegistrationType':          ['TaxRegistrationType'],
    'TaxTransactionCode':           ['TaxTransactionCode'],
    'TaxWithholdTable':             ['TaxWithholdTable','TaxWithholdData'],
    'TaxWithholdGroup':             ['TaxWithholdGroup','TaxWithholdGroupData'],
    'TaxWithholdItemGroup':         ['TaxWithholdItemGroup'],
    'TaxWithholdAuthority':         ['TaxWithholdAuthority'],
    'TaxWithholdPeriodHead':        ['TaxWithholdPeriodHead','TaxWithholdPeriod'],
    'TaxLedgerAccountGroup':        ['TaxLedgerAccountGroup'],
    # Inventory
    'InventParameters':             ['InventParameters','InventParametersJournalName'],
    'InventLocation':               ['InventLocation','InventSiteBalanceSetup'],
    'InventSite':                   ['InventSite'],
    'InventItemGroup':              ['InventItemGroup'],
    'InventModelGroup':             ['InventModelGroup'],
    'InventCostGroup':              ['InventCostGroup'],
    'InventCostVersion':            ['InventCostVersion'],
    'InventCostSheetDesigner':      ['InventCostSheetNode'],
    'InventCostVariance':           ['InventCostVariance'],
    'InventAdjustmentType':         ['InventAdjustmentType'],
    'InventMovType':                ['InventMovType'],
    'InventJournalName':            ['InventJournalName'],
    'InventPosting':                ['InventPosting','InventAccountType'],
    'InventDimCombination':         ['InventDimCombination','DimensionAttribute'],
    'InventItemChargeGroup':        ['InventItemChargeGroup'],
    'InventValueReport':            ['InventValueReport'],
    'InventCountingReasonCode':     ['InventCountingReasonCode'],
    'InventCountingReasonCodePolicy':['InventCountingReasonCodePolicy'],
    # Quality
    'InventTestAssociation':        ['InventTestAssociation'],
    'InventTestTable':              ['InventTestTable'],
    'InventTestGroup':              ['InventTestGroup','InventTestGroupMember'],
    'InventTestVariable':           ['InventTestVariable','InventTestVariableOutcome'],
    'InventTestOperation':          ['InventTestOperation'],
    'InventTestOperationGroup':     ['InventTestOperationGroup'],
    'InventTestWorkGroup':          ['InventTestWorkGroup'],
    'InventTestInstrument':         ['InventTestInstrument'],
    'InventTestItemSampling':       ['InventTestItemSampling'],
    'InventTestItemQualityGroup':   ['InventTestItemQualityGroup','InventTestItemQualityGroupLine'],
    'InventTestQualityCharge':      ['InventTestQualityCharge'],
    'InventTestQualityGroup':       ['InventTestQualityGroup'],
    'InventTestQualityWorker':      ['InventTestQualityWorker'],
    'InventTestCalibration':        ['InventTestCalibration'],
    'InventTestCalibrationGroup':   ['InventTestCalibrationGroup'],
    'InventTestClassificationCode': ['InventTestClassificationCode'],
    'InventTestModifyReasonCode':   ['InventTestModifyReasonCode'],
    'InventTestSampleNameCode':     ['InventTestSampleNameCode'],
    'InventNonConformanceType':     ['InventNonConformanceType'],
    'PdsBatchAttrib':               ['PdsBatchAttrib','PdsBatchAttribEnumValue'],
    'PdsBatchDispositionMaster':    ['PdsBatchDispositionMaster'],
    'PdsBatchNumConfigTable':       ['PdsBatchNumConfigTable'],
    # WHS
    'WHSParameters':                ['WHSParameters'],
    'WHSInventStatus':              ['WHSInventStatus'],
    'WHSLocDirTable':               ['WHSLocDirTable','WHSLocDirLine','WHSLocDirAction'],
    'WHSLocationProfile':           ['WHSLocationProfile'],
    'WHSLocationLimit':             ['WHSLocationLimit'],
    'WHSReservationHierarchy':      ['WHSReservationHierarchy','WHSReservationHierarchyElement'],
    'WHSUnitSeqGroupTable':         ['WHSUnitSeqGroupTable','WHSUnitSeqGroupLine'],
    'WHSDispositionTable':          ['WHSDispositionTable'],
    'WHSZoneTable':                 ['WHSZoneTable'],
    'WHSZoneGroup':                 ['WHSZoneGroup'],
    'WHSCycleCountPlan':            ['WHSCycleCountPlan','WHSCycleCountPlanLine'],
    'WHSLoadTemplate':              ['WHSLoadTemplate'],
    'WHSWaveTemplateTable':         ['WHSWaveTemplateTable','WHSWaveTemplateLine'],
    'WHSWaveMethod':                ['WHSWaveMethod'],
    'WHSWorkClass':                 ['WHSWorkClass'],
    'WHSWorkTemplateTable':         ['WHSWorkTemplateTable','WHSWorkTemplateLine'],
    'WHSWorkPool':                  ['WHSWorkPool'],
    'WHSWorkException':             ['WHSWorkException'],
    'WHSRFMenuTable':               ['WHSRFMenuTable','WHSRFMenuItemTable'],
    'WHSRFMenuItem':                ['WHSRFMenuItem'],
    'WHSMobileAppField':            ['WHSMobileAppField'],
    'WHSMobileAppFieldPriority':    ['WHSMobileAppFieldPriority'],
    'WHSMobileAppStepFieldPriority':['WHSMobileAppStepFieldPriority'],
    'WMSAisle':                     ['WMSAisle'],
    'WMSArrivalOverviewSetup':      ['WMSArrivalOverviewSetup'],
    'WMSLocationFormat':            ['WMSLocationFormat','WMSLocationFormatLine'],
    'WMSLocationType':              ['WMSLocationType'],
    # Sales & Marketing
    'smmParameters':                ['smmParameters'],
    'smmSegment':                   ['smmSegment','smmSubSegment'],
    'SMMActivityChangeReason':      ['SMMActivityChangeReason'],
    'SMMActivityRule':              ['SMMActivityRule'],
    'SMMCustomerClassGroup':        ['SMMCustomerClassGroup','SMMCustomerClassGroupLine'],
    'SalesPool':                    ['SalesPool'],
    'SalesParameters':              ['SalesParameters'],
    'MarkupTable':                  ['MarkupTable','MarkupTrans'],
    'MarkupCode':                   ['MarkupCode'],
    'ReturnAction':                 ['ReturnAction'],
    'ReturnReasonCode':             ['ReturnReasonCode'],
    'ReturnReasonCodeGroup':        ['ReturnReasonCodeGroup'],
    'MCRFulfillmentPolicy':         ['MCRFulfillmentPolicy'],
    'MCRHoldCode':                  ['MCRHoldCode'],
    'GUPParameters':                ['GUPParameters'],
    'GUPPriceAttributeGroup':       ['GUPPriceAttributeGroup','GUPPriceAttributeGroupLine'],
    'GUPPriceComponentCode':        ['GUPPriceComponentCode'],
    'GUPPriceComponentCodeGroup':   ['GUPPriceComponentCodeGroup','GUPPriceComponentCodeGroupLine'],
    'GUPPriceTree':                 ['GUPPriceTree','GUPPriceTreeLine'],
    'PriceDiscAdmName':             ['PriceDiscAdmName'],
    'PriceDiscTable':               ['PriceDiscTable','PriceDiscLine'],
    'PriceDiscTolerance':           ['PriceDiscTolerance'],
    # Procurement
    'PurchParameters':              ['PurchParameters'],
    'PurchPool':                    ['PurchPool'],
    'PurchReqConsolidationRule':    ['PurchReqConsolidationRule'],
    'PurchReqControlRule':          ['PurchReqControlRule'],
    'PurchReqPermission':           ['PurchReqPermission'],
    'PurchReqPurposeRule':          ['PurchReqPurposeRule'],
    'PurchReqBusinessJustification':['PurchReqBusinessJustification'],
    'PurchCategoryAccessPolicyRule':['PurchCategoryAccessPolicyRule'],
    'PurchCategoryPolicyRule':      ['PurchCategoryPolicyRule'],
    'PurchProcurementCategory':     ['EcoResCategoryHierarchy','EcoResCategory'],
    'PurchSigningLimitParameters':  ['PurchSigningLimitParameters','PurchSigningLimit'],
    # Rebate / TAM
    'TAMParameters':                ['TAMParameters'],
    'TAMRebateAgreement':           ['TAMRebateAgreementTable','TAMRebateAgreementLine'],
    'TAMRebateLedgerTable':         ['TAMRebateLedgerTable'],
    # Transportation
    'TMSParameters':                ['TMSParameters'],
    'TMSCarrier':                   ['TMSCarrier','TMSCarrierService'],
    'TMSCarrierGroup':              ['TMSCarrierGroup'],
    'TMSEquipment':                 ['TMSEquipment'],
    'TMSTransportationMethod':      ['TMSTransportationMethod'],
    'TMSTransportDays':             ['TMSTransportDays'],
    'TMSTransportationStatus':      ['TMSTransportationStatus'],
    'TMSZoneMaster':                ['TMSZoneMaster','TMSZoneMasterLine'],
    'TMSLoadBuildStrategy':         ['TMSLoadBuildStrategy'],
    'TMSFreightBillTemplate':       ['TMSFreightBillTemplate'],
    'TMSFreightBillTypeAssignment': ['TMSFreightBillTypeAssignment'],
    # Product
    'EcoResParameters':             ['EcoResParameters'],
    'EcoResCategory':               ['EcoResCategory','EcoResCategoryTranslation'],
    'EcoResCategoryGroup':          ['EcoResCategoryGroup'],
    'EcoResCategoryHierarchy':      ['EcoResCategoryHierarchy','EcoResCategory'],
    'EcoResCategoryHierarchyRole':  ['EcoResCategoryHierarchyRole'],
    'EcoResAttributeType':          ['EcoResAttributeType','EcoResAttributeTypeValue'],
    'EcoResProductLifecycleState':  ['EcoResProductLifecycleState'],
    'EcoResStorageDimensionGroup':  ['EcoResStorageDimensionGroup','EcoResStorageDimensionGroupFldSetup',
                                     'EcoResTrackingDimensionGroup','EcoResTrackingDimensionGroupFldSetup'],
    'EcoResGS1AppIdentifier':       ['EcoResGS1AppIdentifier'],
    'EcoResGS1Policy':              ['EcoResGS1Policy'],
    'EcoResGS1Setup':               ['EcoResGS1Setup'],
    # Production
    'ProdParameters':               ['ProdParameters'],
    'ProdPool':                     ['ProdPool'],
    'ProdJournalName':              ['ProdJournalName'],
    'ProdStdCostVariance':          ['ProdStdCostVariance'],
    'PmfFormula':                   ['PmfFormula','PmfFormulaLine'],
    'RouteTable':                   ['RouteTable','RouteLine'],
    'RouteOprTable':                ['RouteOprTable','RouteOprRelation'],
    'RouteGroup':                   ['RouteGroup'],
    'WrkCtrTable':                  ['WrkCtrTable','WrkCtrGroupTable'],
    'WrkCtrResourceGroup':          ['WrkCtrResourceGroup'],
    'WrkCtrCostCategory':           ['WrkCtrCostCategory'],
    'CostSheetNodeCalcGroup':       ['CostSheetNodeCalcGroup'],
    # Project
    'ProjParameters':               ['ProjParameters'],
    'ProjGroup':                    ['ProjGroup'],
    'ProjCategory':                 ['ProjCategory'],
    'ProjLineProperty':             ['ProjLineProperty'],
    'ProjSharedCategory':           ['ProjSharedCategory','ProjCategoryGroup'],
    'ProjTable':                    ['ProjTable'],
    'ProjLedgerTable':              ['ProjLedgerTable'],
    'ProjCostPriceExpense':         ['ProjCostPriceExpense'],
    'PSABillingClassification':     ['PSABillingClassification'],
    'PSABillingCode':               ['PSABillingCode'],
    # Cost Accounting
    'CAMCostAccountingLedger':      ['CAMCostAccountingLedger'],
    'CAMCostAllocationPolicy':      ['CAMCostAllocationPolicy','CAMCostAllocationPolicyRule'],
    'CAMCostRollupPolicy':          ['CAMCostRollupPolicy'],
    'CAMCostControlWorkspaceConfiguration':['CAMCostControlWorkspaceConfiguration'],
    'CAMDimension':                 ['CAMDimension','CAMDimensionMember'],
    'CAMDimensionHierarchy':        ['CAMDimensionHierarchy','CAMDimensionHierarchyNode'],
    'CAMStatisticalDimension':      ['CAMStatisticalDimension','CAMStatisticalDimensionMember'],
    'CAMStatisticalDimensionMember':['CAMStatisticalDimensionMember'],
    'CAMDataConnector':             ['CAMDataConnector'],
    # Organization
    'OMLegalEntity':                ['OMLegalEntity','DirParty'],
    'OMInternalOrganization':       ['OMInternalOrganization','DirParty'],
    'OMHierarchyTable':             ['OMHierarchyTable','OMHierarchyRelationship'],
    'OMHierarchyType':              ['OMHierarchyType'],
    'OMOperatingUnit':              ['OMOperatingUnit','DirParty'],
    'OMTeam':                       ['OMTeam','OMTeamMember'],
    'DirAddressBook':               ['DirAddressBook'],
    'DirParameters':                ['DirParameters'],
    'DirNameSequence':              ['DirNameSequence'],
    'DirTitleSuffix':               ['DirTitleSuffix'],
    'DirLineOfBusiness':            ['DirLineOfBusiness'],
    'DirPartyBusinessType':         ['DirPartyBusinessType'],
    'DirRelationshipType':          ['DirRelationshipType'],
    'LogisticsLocationRole':        ['LogisticsLocationRole'],
    'LogisticsAddressSetup':        ['LogisticsAddressSetup'],
    'LogisticsAddressZipCode':      ['LogisticsAddressZipCode'],
    'LogisticsIssuingAgency':       ['LogisticsIssuingAgency'],
    'LogisticsElecAddressFormatPopulateRecords':['LogisticsElecAddressFormat'],
    # Number Sequences / Common
    'NumberSequenceTable':          ['NumberSequenceTable','NumberSequenceScope'],
    'NumberSeqExtension':           ['NumberSeqExtension'],
    'NumberSeqConfig':              ['NumberSeqConfig'],
    'Currency':                     ['Currency','CurrencyTxt'],
    'ExchangeRate':                 ['ExchangeRateType','ExchangeRate'],
    'ExchangeRateType':             ['ExchangeRateType'],
    'DateIntervalTable':            ['DateIntervalTable'],
    'Periods':                      ['Periods'],
    'PaymTerm':                     ['PaymTerm'],
    'PaymentCalendar':              ['PaymentCalendar','PaymentCalendarRule'],
    'PaymentSchedule':              ['PaymentSchedule','PaymentScheduleLine'],
    'DlvMode':                      ['DlvMode'],
    'DlvTerm':                      ['DlvTerm'],
    'WorkCalendarTable':            ['WorkCalendarTable','WorkCalendarDate'],
    'WorkTimeTable':                ['WorkTimeTable','WorkTimeLine'],
    'BudgetModel':                  ['BudgetModel'],
    'BudgetControlDimensionAttribute':['BudgetControlDimensionAttribute'],
    'ReqForecastModel':             ['ReqForecastModel'],
    # HR
    'HcmWorker':                    ['HcmWorker','DirPerson','HcmEmployment'],
    'HcmJob':                       ['HcmJob'],
    'HcmPosition':                  ['HcmPosition','HcmPositionWorkerAssignment'],
    'HcmPositionHierarchy':         ['HcmPositionHierarchy','HcmPositionHierarchyLine'],
    'HcmCertificateType':           ['HcmCertificateType'],
    'TrvExpLedgerAccounts':         ['TrvExpLedgerAccounts'],
    # Electronic Reporting
    'ERSolutionTable':              ['ERSolutionTable','ERModelDefinitionTable'],
    'ERFormatDestination':          ['ERFormatDestination','ERFormatDestinationItem'],
    'ERFormatMappingRunJobTable':   ['ERFormatMappingRunJobTable'],
    'ERTableDataSourceParameters':  ['ERTableDataSourceParameters'],
    'ERLegalEntityTable':           ['ERLegalEntityTable'],
    # Electronic Signatures
    'SIGParameters':                ['SIGParameters'],
    'SIGReasonCode':                ['SIGReasonCode'],
    'SIGSignatureRequirement':      ['SIGSignatureRequirement'],
    # System
    'SysEmailTable':                ['SysEmailTable','SysEmailMessageTable'],
    'SysEmailParameters':           ['SysEmailParameters'],
    'SysUserGroup':                 ['SysUserGroup','SysUserGroupInfo'],
    'SysUserInfo':                  ['SysUserInfo','DirPerson'],
    'SysMessageProcessing':         ['SysMessageProcessing'],
    'SysMessageProcessingAction':   ['SysMessageProcessingAction'],
    'SysMessageItemStatus':         ['SysMessageItemStatus'],
    'SysMessageItemType':           ['SysMessageItemType'],
    'SysMessageStatus':             ['SysMessageStatus'],
    'SysProcessAutomationType':     ['SysProcessAutomationType'],
    'SysCorpNetPrinter':            ['SysCorpNetPrinter'],
    'SysExternalAppTable':          ['SysExternalAppTable'],
    'SysExternalExecutable':        ['SysExternalExecutable'],
    'SysLabel':                     ['SysLabel'],
    'SysTableBrowser':              ['SysTableBrowser'],
    'AifWebServiceTableTable':      ['AifWebServiceTableTable'],
    # Document Management
    'DocuType':                     ['DocuType'],
    'DocuParameters':               ['DocuParameters'],
    'DocuTemplate':                 ['DocuTemplate'],
    'DocuActiveTable':              ['DocuActiveTable'],
    'DocuField':                    ['DocuField'],
    'DocuContentType':              ['DocuContentType'],
    # Print
    'PrintMgmtDocInstance':         ['PrintMgmtDocInstance','PrintMgmtSettings'],
    # Exchange
    'ExchangeServerTable':          ['ExchangeServerTable'],
    'ExchangeServerParameters':     ['ExchangeServerParameters'],
    'ExchangeServerMapping':        ['ExchangeServerMapping'],
    # Workflow
    'WorkflowParameters':           ['WorkflowParameters'],
    'WorkflowItemQueue':            ['WorkflowItemQueue'],
    'WorkflowItemQueueGroup':       ['WorkflowItemQueueGroup'],
    'WorkflowItemQueueAssignmentRule':['WorkflowItemQueueAssignmentRule'],
    # Intrastat
    'IntrastatPort':                ['IntrastatPort'],
    'IntrastatSetup':               ['IntrastatSetup'],
    # Misc
    'BatchJobTable':                ['BatchJobTable','BatchJob'],
    'OfficeAppDocuments':           ['OfficeAppDocuments'],
    'KMDescription':                ['KMDescription'],
    'FinancialReports':             ['FinancialReports'],
    'DataMigrationSourceSystem':    ['DataMigrationSourceSystem'],
}


def get_tables(mi_name):
    if mi_name in CBA_DS:
        ds = CBA_DS[mi_name]
        return ('Custom CBA', 'REF', ds if ds else [mi_name])
    if mi_name in STD_DS:
        tables = STD_DS[mi_name]
        ext = ' + CBA ext' if mi_name in CBA_EXTENDED else ''
        return ('Standard D365' + ext, 'Knowledge', tables)
    # Fallback
    return ('Standard D365', 'Inferred', [mi_name])


# ── 3. Load mapped results and build full table ────────────────────────────
with open('/tmp/cdd_mapped.json', encoding='utf-8') as f:
    data = json.load(f)
results = data['results']

# Build per-form summary (deduplicated by mi_name)
from collections import Counter
cdd_count = Counter(r.get('mi_name', '').strip() for r in results if r.get('mi_name', '').strip())

forms_seen = {}
form_rows = []
for r in results:
    mi = r.get('mi_name', '').strip()
    if not mi:
        continue
    if mi in forms_seen:
        forms_seen[mi]['ado_ids'].append(r['id'])
        continue
    kind, source, tables = get_tables(mi)
    entry = {
        'mi_name':    mi,
        'display':    r.get('form_name', ''),
        'confidence': r.get('confidence', ''),
        'kind':       kind,
        'source':     source,
        'tables':     tables,
        'table_count': len(tables),
        'ado_ids':    [r['id']],
        'cba_extended': mi in CBA_EXTENDED,
        'cdd_count':  cdd_count.get(mi, 0),
    }
    forms_seen[mi] = entry
    form_rows.append(entry)

form_rows.sort(key=lambda x: x['mi_name'])

# ── 4. Generate Excel ──────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Form → Tables ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Form → Tables"

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red      = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
orange   = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
purple   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin     = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

# Title
ws1.merge_cells('A1:I1')
ws1['A1'] = "MY/SG CDD — Form to D365 Table Mapping"
ws1['A1'].font = Font(bold=True, size=14)
ws1['A2'] = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
             f"Forms: {len(form_rows)}  |  "
             f"CBA: {sum(1 for f in form_rows if 'CBA' in f['kind'])}  |  "
             f"Standard: {sum(1 for f in form_rows if 'Standard' in f['kind'])}")
ws1['A2'].font = Font(size=11, italic=True)

headers = ["MI Name (Form)", "Display Name", "Confidence", "Type",
           "CBA Extended?", "# CDDs", "# Tables", "Primary Table",
           "All Tables (pipe-sep)", "Source"]
HR = 4
for ci, h in enumerate(headers, 1):
    c = ws1.cell(row=HR, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = thin

for ri, f in enumerate(form_rows, HR + 1):
    tables = f['tables']
    primary = tables[0] if tables else ''
    all_tbl = ' | '.join(tables)
    ext_flag = 'YES' if f['cba_extended'] else ''
    vals = [f['mi_name'], f['display'], f['confidence'], f['kind'],
            ext_flag, f['cdd_count'], f['table_count'], primary, all_tbl, f['source']]
    for ci, v in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=v)
        c.border = thin
        c.alignment = Alignment(vertical='top', wrap_text=(ci in (2, 9)))

    # Row color
    if 'Custom CBA' in f['kind']:
        fill = orange
    elif f['confidence'] == 'H':
        fill = green
    elif f['confidence'] == 'M':
        fill = yellow
    else:
        fill = red
    for ci in range(1, len(headers) + 1):
        ws1.cell(row=ri, column=ci).fill = fill
    if f['cba_extended']:
        ws1.cell(row=ri, column=5).fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")

col_widths = [30, 40, 12, 22, 14, 10, 10, 30, 70, 12]
for ci, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.auto_filter.ref = f"A{HR}:J{HR + len(form_rows)}"
ws1.freeze_panes = f"A{HR + 1}"

# ── Sheet 2: All Tables (flat — one row per table per form) ────────────────
ws2 = wb.create_sheet("All Tables (flat)")
ws2.merge_cells('A1:E1')
ws2['A1'] = "All D365 Tables per Form — flat list (one row per table)"
ws2['A1'].font = Font(bold=True, size=13)

h2 = ["MI Name (Form)", "Display Name", "Table Name", "Table #", "Type"]
for ci, h in enumerate(h2, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = thin
    c.alignment = Alignment(horizontal='center')

row2 = 3
for f in form_rows:
    for ti, tbl in enumerate(f['tables'], 1):
        vals = [f['mi_name'], f['display'], tbl, ti, f['kind']]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=row2, column=ci, value=v)
            c.border = thin
        if 'Custom CBA' in f['kind']:
            for ci in range(1, 6): ws2.cell(row=row2, column=ci).fill = orange
        row2 += 1

for ci, w in zip(range(1, 6), [28, 38, 35, 8, 20]):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.auto_filter.ref = f"A2:E{row2}"
ws2.freeze_panes = "A3"

# ── Sheet 3: CBA Extended forms ────────────────────────────────────────────
ws3 = wb.create_sheet("CBA Extended Forms")
ws3['A1'] = "Standard D365 forms that have CBA model extensions (extra fields/datasources)"
ws3['A1'].font = Font(bold=True, size=12)
h3 = ["MI Name (Form)", "Display Name", "Confidence", "Standard Tables"]
for ci, h in enumerate(h3, 1):
    c = ws3.cell(row=2, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.border = thin
ext_rows = [f for f in form_rows if f['cba_extended'] and 'Custom CBA' not in f['kind']]
for ri, f in enumerate(ext_rows, 3):
    vals = [f['mi_name'], f['display'], f['confidence'], ' | '.join(f['tables'])]
    for ci, v in enumerate(vals, 1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.border = thin
        c.fill = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
for ci, w in zip(range(1, 5), [28, 40, 12, 60]):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.auto_filter.ref = f"A2:D{2 + len(ext_rows)}"
ws3.freeze_panes = "A3"

# ── Sheet 4: Summary ───────────────────────────────────────────────────────
ws4 = wb.create_sheet("Summary")
total_forms = len(form_rows)
total_tables = sum(f['table_count'] for f in form_rows)
cba_forms  = sum(1 for f in form_rows if 'Custom CBA' in f['kind'])
std_forms  = sum(1 for f in form_rows if 'Standard D365' in f['kind'])
ext_forms  = sum(1 for f in form_rows if f['cba_extended'])
multi_tbl  = sum(1 for f in form_rows if f['table_count'] > 1)
from_ref   = sum(1 for f in form_rows if f['source'] == 'REF')
from_know  = sum(1 for f in form_rows if f['source'] == 'Knowledge')
from_inf   = sum(1 for f in form_rows if f['source'] == 'Inferred')

summary_data = [
    ("Unique forms mapped", total_forms),
    ("Total table references", total_tables),
    ("Avg tables per form", round(total_tables / total_forms, 1)),
    ("Forms with >1 table", multi_tbl),
    ("", ""),
    ("Custom CBA forms", cba_forms),
    ("Standard D365 forms", std_forms),
    ("Standard + CBA extended", ext_forms),
    ("", ""),
    ("Source: CBA reference file", from_ref),
    ("Source: D365 knowledge", from_know),
    ("Source: Inferred (same name)", from_inf),
]
ws4['A1'] = "Summary"; ws4['A1'].font = Font(bold=True, size=14)
for ri, (label, val) in enumerate(summary_data, 2):
    ws4.cell(row=ri, column=1, value=label)
    ws4.cell(row=ri, column=2, value=val)
    if label:
        ws4.cell(row=ri, column=1).font = Font(bold=True)
ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 12

# Save
out = r"C:\D365 Configuration Drift Analysis\output\CDD_Form_Tables.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Forms: {total_forms} | Tables: {total_tables} | CBA: {cba_forms} | Std: {std_forms} | CBA-extended: {ext_forms}")
print(f"Inferred (same name): {from_inf}")
