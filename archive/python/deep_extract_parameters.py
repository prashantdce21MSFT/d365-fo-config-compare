"""
Deep extraction of D365 parameters forms — opens every tab, re-parses
FormState, paginates sub-grids, runs find_controls sweep, then writes
a comparison Excel (UAT vs Config) with one sheet per tab.

Usage:
    python deep_extract_parameters.py
"""
import sys
import os
import json
import time
import io
from datetime import datetime

# Fix Windows console encoding for Unicode characters (arrows, etc.)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from d365_mcp_client import D365McpClient
from form_reader import open_form, close_form, mcp_call
from form_control_extractor import FIND_CONTROLS_TERMS

CONFIG_FILE = r"C:\D365DataValidator\config.json"
OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output\Comparison"


def parse_form_obj(form_obj, depth=0):
    """Recursively parse FormState.Form into fields, grids, tabs."""
    data = {"fields": {}, "grids": {}, "tabs": {}}
    if not isinstance(form_obj, dict) or depth > 10:
        return data

    for gname, ginfo in form_obj.get("Grid", {}).items():
        if not isinstance(ginfo, dict):
            continue
        columns = [{"name": c.get("Name", ""), "label": c.get("Label", "")}
                   for c in ginfo.get("Columns", [])]
        rows = []
        for row in ginfo.get("Rows", []):
            vals = row.get("Values", {})
            clean = {k: v for k, v in vals.items() if not k.startswith("<")}
            if any(v for v in clean.values()):
                rows.append(clean)
        pag = ginfo.get("Pagination", {})
        data["grids"][gname] = {
            "columns": columns,
            "rows": rows,
            "total_rows": len(rows),
            "has_next": str(pag.get("HasNextPage", "False")).lower() == "true",
        }

    field_types = {
        "Input": "input", "Checkbox": "checkbox", "Combobox": "combobox",
        "RadioButton": "radiobutton", "RealInput": "realinput",
        "IntInput": "intinput", "DateInput": "dateinput",
        "TimeInput": "timeinput", "SegmentedEntry": "segmentedentry",
    }
    for ftype, ftype_label in field_types.items():
        for fname, finfo in form_obj.get(ftype, {}).items():
            if not isinstance(finfo, dict):
                continue
            if ftype == "Combobox":
                val = finfo.get("ValueText", finfo.get("Value", ""))
            elif ftype == "Checkbox":
                val = finfo.get("IsChecked", finfo.get("Value", ""))
            else:
                val = finfo.get("Value", "")
            data["fields"][fname] = {
                "label": finfo.get("Label", fname),
                "value": val,
                "type": ftype_label,
            }

    for container_key in ("Tab", "TabPage"):
        for tname, tinfo in form_obj.get(container_key, {}).items():
            if not isinstance(tinfo, dict):
                continue
            data["tabs"][tname] = {
                "label": tinfo.get("Label", tinfo.get("Text", tname)),
            }
            children = tinfo.get("Children", tinfo)
            child_data = parse_form_obj(children, depth + 1)
            data["fields"].update(child_data["fields"])
            data["grids"].update(child_data["grids"])
            for tn, tv in child_data["tabs"].items():
                if tn not in data["tabs"]:
                    data["tabs"][tn] = tv

    for container_key in ("Group", "ReferenceGroup"):
        for _gname, ginfo in form_obj.get(container_key, {}).items():
            if isinstance(ginfo, dict):
                child_data = parse_form_obj(ginfo, depth + 1)
                data["fields"].update(child_data["fields"])
                data["grids"].update(child_data["grids"])

    children = form_obj.get("Children", {})
    if isinstance(children, dict):
        child_data = parse_form_obj(children, depth + 1)
        data["fields"].update(child_data["fields"])
        data["grids"].update(child_data["grids"])
        for tn, tv in child_data["tabs"].items():
            if tn not in data["tabs"]:
                data["tabs"][tn] = tv

    return data


def find_grid_in_form(form_obj, grid_name):
    """Recursively find a grid by name in FormState tree."""
    if not isinstance(form_obj, dict):
        return None
    grids = form_obj.get("Grid", {})
    if isinstance(grids, dict) and grid_name in grids:
        gi = grids[grid_name]
        if isinstance(gi, dict):
            return gi
    for key in ("Tab", "TabPage", "Group", "Children"):
        container = form_obj.get(key, {})
        if isinstance(container, dict):
            for _n, child in container.items():
                if isinstance(child, dict):
                    for sub in [child.get("Children", {}), child]:
                        result = find_grid_in_form(sub, grid_name)
                        if result:
                            return result
    return None


def paginate_grid(client, grid_name, initial_rows, form_obj):
    """Paginate a grid to collect all rows."""
    all_rows = list(initial_rows)
    page = 1
    while page < 200:
        try:
            next_result = mcp_call(client, "form_click_control", {
                "controlName": grid_name, "actionId": "LoadNextPage",
            })
            if "raw" in next_result:
                break
            form = next_result.get("FormState", {}).get("Form", {})
            gi = find_grid_in_form(form, grid_name)
            if not gi:
                break
            page_rows = []
            for row in gi.get("Rows", []):
                vals = row.get("Values", {})
                clean = {k: v for k, v in vals.items() if not k.startswith("<")}
                if any(v for v in clean.values()):
                    page_rows.append(clean)
            if not page_rows:
                break
            all_rows.extend(page_rows)
            page += 1
            print(f"        Page {page}: +{len(page_rows)} rows (total: {len(all_rows)})")
            pag = gi.get("Pagination", {})
            if str(pag.get("HasNextPage", "False")).lower() != "true":
                break
        except Exception as e:
            print(f"        Pagination error: {e}")
            break
    return all_rows


def find_controls_sweep(client):
    """Run broad form_find_controls sweep."""
    all_controls = {}
    for term in FIND_CONTROLS_TERMS:
        try:
            raw = client.call_tool("form_find_controls", {"controlSearchTerm": term})
            if isinstance(raw, str) and raw.strip():
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    for item in parsed:
                        cname = item.get("Name", "")
                        if cname and cname not in all_controls:
                            props = item.get("Properties", {})
                            if "Columns" in props:
                                ctype = "grid"
                            elif "IsChecked" in props:
                                ctype = "checkbox"
                            elif "Options" in props:
                                ctype = "combobox"
                            else:
                                ctype = "input"
                            if ctype == "combobox":
                                val = props.get("ValueText", props.get("Value", ""))
                            elif ctype == "checkbox":
                                val = props.get("IsChecked", props.get("Value", ""))
                            elif ctype == "grid":
                                val = f"[grid: {len(props.get('Rows', []))} rows]"
                            else:
                                val = props.get("Value", "")
                            all_controls[cname] = {
                                "label": props.get("Label", cname),
                                "value": val,
                                "type": ctype,
                            }
        except Exception:
            pass
    return all_controls


def extract_le(env_config, mi_name, mi_type, company):
    """Deep extract one legal entity: open form, walk every tab, sweep controls."""
    client = D365McpClient(env_config)
    for attempt in range(3):
        try:
            client.connect()
            break
        except Exception as e:
            if attempt < 2:
                wait = 10 * (attempt + 1)
                print(f"    Connection failed (attempt {attempt+1}/3), retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise

    result = {
        "company": company,
        "form_name": "", "form_caption": "",
        "status": "",
        "tabs": {},       # {tab_name: {label, fields: {}, grids: {}}}
        "sweep": {},      # find_controls sweep results
        "all_fields": {}, # merged deduplicated fields
        "all_grids": {},  # merged grids with pagination
    }

    try:
        # 1. Open form
        print(f"    Opening {mi_name} ({mi_type}) on {company}...")
        form_result = open_form(client, mi_name, mi_type, company)
        fs = form_result.get("FormState", {})
        result["form_name"] = fs.get("Name", "")
        result["form_caption"] = fs.get("Caption", "")
        form_obj = fs.get("Form", {})
        print(f"    Form: {result['form_name']} - {result['form_caption']}")

        # 2. Parse initial FormState
        initial = parse_form_obj(form_obj)
        tab_names = sorted(initial["tabs"].keys())
        print(f"    Tabs discovered: {len(tab_names)}")

        # Build tab label map
        tab_labels = {}
        for tn, tinfo in initial["tabs"].items():
            tab_labels[tn] = tinfo.get("label", tn)

        # Collect initial fields (tagged as "_initial")
        for fname, fdata in initial["fields"].items():
            fdata["_tab"] = "(Initial)"
            result["all_fields"][fname] = fdata

        # Collect initial grids
        for gname, gdata in initial["grids"].items():
            result["all_grids"][gname] = gdata

        # 3. Open each tab and re-parse
        for ti, tab_name in enumerate(tab_names):
            tab_label = tab_labels.get(tab_name, tab_name)
            print(f"    [{ti+1}/{len(tab_names)}] Opening tab: {tab_label} ({tab_name})...")
            try:
                tab_raw = client.call_tool(
                    "form_open_or_close_tab",
                    {"tabName": tab_name, "tabAction": "Open"}
                )
                tab_fields = {}
                tab_grids = {}

                if isinstance(tab_raw, str) and tab_raw.strip():
                    tab_parsed = json.loads(tab_raw)
                    if isinstance(tab_parsed, dict):
                        tab_form = tab_parsed.get("FormState", {}).get("Form", {})
                        if tab_form:
                            tab_data = parse_form_obj(tab_form)

                            # Prefer fields from the tab's own children
                            tab_specific = {}
                            for tn_check in (tab_name,):
                                for container_key in ("Tab", "TabPage"):
                                    for tn, tinfo in tab_form.get(container_key, {}).items():
                                        if tn == tn_check and isinstance(tinfo, dict):
                                            children = tinfo.get("Children", tinfo)
                                            child_data = parse_form_obj(children)
                                            tab_specific.update(child_data["fields"])
                                            for gn, gd in child_data["grids"].items():
                                                tab_grids[gn] = gd

                            # Use tab-specific fields if found, else all new fields
                            if tab_specific:
                                tab_fields = tab_specific
                            else:
                                # Use all fields from re-parse, but only NEW ones
                                for fname, fdata in tab_data["fields"].items():
                                    if fname not in result["all_fields"]:
                                        tab_fields[fname] = fdata

                            # Also pick up new grids
                            for gname, gdata in tab_data["grids"].items():
                                if gname not in tab_grids:
                                    tab_grids[gname] = gdata

                # Tag and store
                for fname, fdata in tab_fields.items():
                    fdata["_tab"] = tab_label
                    result["all_fields"][fname] = fdata

                for gname, gdata in tab_grids.items():
                    gdata["_tab"] = tab_label
                    result["all_grids"][gname] = gdata

                new_count = len(tab_fields)
                grid_count = len(tab_grids)
                extra = f", {grid_count} sub-grids" if grid_count else ""
                print(f"      → {new_count} fields{extra}")

                result["tabs"][tab_name] = {
                    "label": tab_label,
                    "fields": tab_fields,
                    "grids": tab_grids,
                }

                time.sleep(0.2)
            except Exception as e:
                print(f"      Tab open failed: {e}")
                result["tabs"][tab_name] = {
                    "label": tab_label, "fields": {}, "grids": {},
                }

        # 4. Paginate ALL sub-grids — always attempt, paginate_grid stops
        #    naturally when HasNextPage=False (avoids missing rows when
        #    has_next flag is not set correctly on initial tab load)
        for gname, gdata in result["all_grids"].items():
            print(f"    Paginating sub-grid '{gname}'...")
            all_rows = paginate_grid(client, gname, gdata["rows"], form_obj)
            gdata["rows"] = all_rows
            gdata["total_rows"] = len(all_rows)
            print(f"      Total rows: {len(all_rows)}")

        # 5. form_find_controls sweep
        print(f"    Running find_controls sweep...")
        sweep = find_controls_sweep(client)
        result["sweep"] = sweep
        print(f"      → {len(sweep)} controls from sweep")

        # Merge sweep into all_fields (don't overwrite tab-sourced)
        for cname, cdata in sweep.items():
            if cname not in result["all_fields"] and cdata.get("type") != "grid":
                cdata["_tab"] = "(Sweep)"
                result["all_fields"][cname] = cdata

        result["status"] = "OK"
        total_fields = len(result["all_fields"])
        total_grids = len(result["all_grids"])
        total_grid_rows = sum(g.get("total_rows", len(g.get("rows", [])))
                              for g in result["all_grids"].values())
        print(f"    TOTAL: {total_fields} fields, {total_grids} grids ({total_grid_rows} rows)")

        close_form(client)
    except Exception as e:
        result["status"] = f"ERROR: {str(e)[:300]}"
        print(f"    ERROR: {e}")
        try:
            close_form(client)
        except Exception:
            pass

    return result


def extract_env(mi_name, mi_type, legal_entities, env_key, env_label):
    """Extract from one environment across all legal entities."""
    with open(CONFIG_FILE) as f:
        config = json.load(f)
    env_config = config["environments"][env_key]

    print(f"\n{'='*60}")
    print(f"  Environment: {env_label} ({env_key})")
    print(f"{'='*60}")

    results = {}
    for le in legal_entities:
        print(f"\n  --- Legal Entity: {le} ---")
        results[le] = extract_le(env_config, mi_name, mi_type, le)

    return results


def to_excel(uat_results, config_results, legal_entities, mi_name):
    """Write comparison Excel with per-tab sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUT_DIR, f"DeepCompare_{mi_name}_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # Styles
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    grid_hfill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    bold = Font(bold=True, size=10)
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    only_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    le_sep_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    le_sep_font = Font(bold=True, size=11, color="1F3864")

    def write_header(ws, headers, row=1, fill=None):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hf
            c.fill = fill or hfill
            c.border = tb
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws, ncols, max_row):
        for ci in range(1, ncols + 1):
            mx = 8
            for r in range(1, min(max_row + 1, 200)):
                v = ws.cell(row=r, column=ci).value
                if v:
                    mx = max(mx, min(len(str(v)), 55))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 58)

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    info_rows = [
        ["Form", mi_name],
        ["UAT Env", "Env1 UAT (ENV1)"],
        ["Config Env", "Env4 Config (ENV4)"],
        ["Extracted", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["Legal Entity", "UAT Status", "Config Status",
         "UAT Fields", "Config Fields", "UAT Grids", "Config Grids"],
    ]
    header_idx = 6
    for le in legal_entities:
        uat_le = uat_results.get(le, {})
        cfg_le = config_results.get(le, {})
        info_rows.append([
            le,
            uat_le.get("status", "N/A"),
            cfg_le.get("status", "N/A"),
            len(uat_le.get("all_fields", {})),
            len(cfg_le.get("all_fields", {})),
            len(uat_le.get("all_grids", {})),
            len(cfg_le.get("all_grids", {})),
        ])

    for ri, row_data in enumerate(info_rows, 1):
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = tb
            if ri <= 4:
                c.font = bold
            if ri == header_idx:
                c.font, c.fill = hf, hfill
    for ci in range(1, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # ── Collect all tab names across all LEs/envs ──
    all_tabs = {}  # {tab_name: label}
    for le in legal_entities:
        for results in (uat_results, config_results):
            le_data = results.get(le, {})
            for tname, tinfo in le_data.get("tabs", {}).items():
                if tname not in all_tabs:
                    all_tabs[tname] = tinfo.get("label", tname)

    # ── Per-tab comparison sheets ──
    for tab_name in sorted(all_tabs.keys()):
        tab_label = all_tabs[tab_name]
        sheet_name = tab_label[:28] if len(tab_label) > 28 else tab_label
        # Ensure unique sheet name
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:25] + f"_{len(existing)}"

        ws_tab = wb.create_sheet(sheet_name)

        # Collect all field names in this tab across all LEs/envs
        tab_fields_union = {}  # {fname: label}
        tab_grids_union = {}   # {gname: [col_labels]}
        for le in legal_entities:
            for results in (uat_results, config_results):
                le_data = results.get(le, {})
                tinfo = le_data.get("tabs", {}).get(tab_name, {})
                for fname, fdata in tinfo.get("fields", {}).items():
                    if fname not in tab_fields_union:
                        tab_fields_union[fname] = fdata.get("label", fname)
                for gname, gdata in tinfo.get("grids", {}).items():
                    if gname not in tab_grids_union:
                        tab_grids_union[gname] = [
                            c.get("label", c.get("name", ""))
                            for c in gdata.get("columns", [])
                        ]

        # Fields comparison table
        if tab_fields_union:
            headers = ["#", "Control", "Label", "UAT Value", "Config Value", "Match", "LE"]
            write_header(ws_tab, headers)
            ri = 2
            seq = 0

            for le in legal_entities:
                uat_tab = uat_results.get(le, {}).get("tabs", {}).get(tab_name, {})
                cfg_tab = config_results.get(le, {}).get("tabs", {}).get(tab_name, {})
                uat_fields = uat_tab.get("fields", {})
                cfg_fields = cfg_tab.get("fields", {})

                # LE separator
                for ci in range(1, len(headers) + 1):
                    c = ws_tab.cell(row=ri, column=ci)
                    c.fill = le_sep_fill
                    c.border = tb
                ws_tab.cell(row=ri, column=1, value=le).font = le_sep_font
                ws_tab.cell(row=ri, column=1).fill = le_sep_fill
                ws_tab.cell(row=ri, column=2,
                            value=f"UAT: {len(uat_fields)} | Config: {len(cfg_fields)} fields"
                            ).font = le_sep_font
                ws_tab.cell(row=ri, column=2).fill = le_sep_fill
                ri += 1

                # All fields in this tab
                all_fnames = list(tab_fields_union.keys())
                for fname in all_fnames:
                    uat_fdata = uat_fields.get(fname, {})
                    cfg_fdata = cfg_fields.get(fname, {})
                    uat_val = str(uat_fdata.get("value", "")) if uat_fdata else ""
                    cfg_val = str(cfg_fdata.get("value", "")) if cfg_fdata else ""
                    label = tab_fields_union[fname]

                    if not uat_val and not cfg_val:
                        continue

                    seq += 1
                    if uat_val == cfg_val:
                        match = "Match"
                        mfill = match_fill
                    elif uat_val and cfg_val:
                        match = "DIFF"
                        mfill = diff_fill
                    elif uat_val and not cfg_val:
                        match = "UAT only"
                        mfill = only_fill
                    else:
                        match = "Config only"
                        mfill = only_fill

                    ws_tab.cell(row=ri, column=1, value=seq).border = tb
                    ws_tab.cell(row=ri, column=2, value=fname).border = tb
                    ws_tab.cell(row=ri, column=3, value=label).border = tb
                    ws_tab.cell(row=ri, column=4, value=uat_val).border = tb
                    ws_tab.cell(row=ri, column=5, value=cfg_val).border = tb
                    mc = ws_tab.cell(row=ri, column=6, value=match)
                    mc.border = tb
                    mc.fill = mfill
                    ws_tab.cell(row=ri, column=7, value=le).border = tb
                    ri += 1

            auto_width(ws_tab, len(headers), ri)
            ws_tab.freeze_panes = "B2"
            if ri > 2:
                ws_tab.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ri - 1}"

        # Sub-grids in this tab
        for gname, col_labels in tab_grids_union.items():
            if not col_labels:
                continue
            # Add a gap and sub-grid header
            g_sheet_name = f"{sheet_name[:20]}_{gname[:8]}"
            existing_names = [s.title for s in wb.worksheets]
            if g_sheet_name in existing_names:
                g_sheet_name = g_sheet_name[:25] + f"_{len(existing_names)}"
            ws_g = wb.create_sheet(g_sheet_name)

            g_headers = ["#"] + col_labels + ["Env", "LE"]
            write_header(ws_g, g_headers, fill=grid_hfill)
            gri = 2
            gseq = 0

            for le in legal_entities:
                le_rows = []
                for env_label, results in [("UAT", uat_results), ("Config", config_results)]:
                    le_data = results.get(le, {})
                    tinfo = le_data.get("tabs", {}).get(tab_name, {})
                    gdata = tinfo.get("grids", {}).get(gname, {})
                    for row in gdata.get("rows", []):
                        le_rows.append((env_label, row))

                if not le_rows:
                    continue

                # LE separator
                for ci in range(1, len(g_headers) + 1):
                    c = ws_g.cell(row=gri, column=ci)
                    c.fill = le_sep_fill
                    c.border = tb
                ws_g.cell(row=gri, column=1, value=le).font = le_sep_font
                ws_g.cell(row=gri, column=1).fill = le_sep_fill
                gri += 1

                for env_label, row_data in le_rows:
                    gseq += 1
                    ws_g.cell(row=gri, column=1, value=gseq).border = tb
                    for ci, cl in enumerate(col_labels, 2):
                        ws_g.cell(row=gri, column=ci, value=row_data.get(cl, "")).border = tb
                    ws_g.cell(row=gri, column=len(col_labels) + 2, value=env_label).border = tb
                    ws_g.cell(row=gri, column=len(col_labels) + 3, value=le).border = tb
                    gri += 1

            auto_width(ws_g, len(g_headers), gri)
            ws_g.freeze_panes = "B2"

    # ── All Controls flat sheet ──
    ws_all = wb.create_sheet("All Controls")
    all_headers = ["#", "Control", "Label", "Value", "Type", "Tab", "Env", "LE"]
    write_header(ws_all, all_headers)
    ari = 2
    aseq = 0

    for le in legal_entities:
        for env_label, results in [("UAT", uat_results), ("Config", config_results)]:
            le_data = results.get(le, {})
            for fname, fdata in sorted(le_data.get("all_fields", {}).items()):
                aseq += 1
                ws_all.cell(row=ari, column=1, value=aseq).border = tb
                ws_all.cell(row=ari, column=2, value=fname).border = tb
                ws_all.cell(row=ari, column=3, value=fdata.get("label", "")).border = tb
                ws_all.cell(row=ari, column=4, value=str(fdata.get("value", ""))).border = tb
                ws_all.cell(row=ari, column=5, value=fdata.get("type", "")).border = tb
                ws_all.cell(row=ari, column=6, value=fdata.get("_tab", "")).border = tb
                ws_all.cell(row=ari, column=7, value=env_label).border = tb
                ws_all.cell(row=ari, column=8, value=le).border = tb
                ari += 1

    auto_width(ws_all, len(all_headers), ari)
    ws_all.freeze_panes = "B2"
    if ari > 2:
        ws_all.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{ari - 1}"

    wb.save(filepath)
    print(f"\nExcel saved: {filepath}")
    return filepath


def main():
    mi_name = "smmParameters"
    mi_type = "Display"
    legal_entities = ["MY30", "MY60", "SG60"]

    print("=" * 70)
    print("  Deep Extraction: Sales and marketing parameters")
    print("  smmParameters — UAT (ENV1) vs Config (ENV4)")
    print(f"  LEs: {', '.join(legal_entities)}")
    print("=" * 70)

    print("\n>>> Extracting from UAT (ENV1)...")
    uat_results = extract_env(mi_name, mi_type, legal_entities, "ENV1", "Env1 UAT")

    print("\n>>> Extracting from Config (ENV4)...")
    config_results = extract_env(mi_name, mi_type, legal_entities, "ENV4", "Env4 Config")

    print("\n>>> Writing comparison Excel...")
    filepath = to_excel(uat_results, config_results, legal_entities, mi_name)
    print(f"\nDone! Open: {filepath}")


if __name__ == "__main__":
    main()
