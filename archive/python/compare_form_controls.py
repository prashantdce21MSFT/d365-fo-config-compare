"""
Extract D365 form controls from UAT and Config environments and dump
all values into Excel. One row per grid record — grid columns + detail
fields side by side. No comparison logic — just raw data with Env/LE
columns for filtering.
"""
import sys
import os
import json
from datetime import datetime

sys.path.insert(0, r"C:\D365DataValidator")
sys.path.insert(0, r"C:\D365 Configuration Drift Analysis")

from form_control_extractor import FormControlExtractor

OUTPUT_DIR = r"C:\D365 Configuration Drift Analysis\output\Comparison"


def collect_grid_data(env_results, legal_entities):
    """Collect grid info across all envs/LEs.
    env_results = [("UAT", uat_result), ("Config", config_result)]
    Returns: {grid_name: {col_labels, detail_labels}}
    detail_labels are tab-prefixed: "Tab Label > Field Label"
    """
    grids = {}
    for env_name, env_result in env_results:
        for le in legal_entities:
            le_data = env_result["legal_entities"].get(le, {})
            for gname, gdata in le_data.get("grids", {}).items():
                if gname not in grids:
                    grids[gname] = {"col_labels": [], "col_seen": set(),
                                    "detail_labels": [], "detail_seen": set(),
                                    "detail_label_map": {}}
                g = grids[gname]
                for col in gdata.get("columns", []):
                    lbl = col.get("label", col.get("name", ""))
                    if lbl and lbl not in g["col_seen"]:
                        g["col_labels"].append(lbl)
                        g["col_seen"].add(lbl)
                for ri, fields in gdata.get("row_details", {}).items():
                    for fname, fdata in fields.items():
                        raw_lbl = fdata.get("label", fname)
                        tab_lbl = fdata.get("_tab", "")
                        # Build prefixed label: "Tab > Field"
                        if tab_lbl:
                            dlbl = f"{tab_lbl} > {raw_lbl}"
                        else:
                            dlbl = raw_lbl
                        if dlbl and dlbl not in g["detail_seen"] and raw_lbl not in g["col_seen"]:
                            g["detail_labels"].append(dlbl)
                            g["detail_seen"].add(dlbl)
                            # Map (fname) -> prefixed label for lookup
                            g["detail_label_map"][fname] = dlbl
    return grids


def to_excel(env_results, legal_entities, uat_result, config_result,
             ado_id="", area_path="", out_path=None):
    """Write all extracted values to Excel — no comparison, just data dump."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    form_name = uat_result.get("form_name") or uat_result.get("menu_item", "Form")
    if out_path:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        filepath = out_path
    else:
        filepath = os.path.join(OUTPUT_DIR, f"Compare_{form_name}_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # ── Styles ──
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    detail_hfill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    bold = Font(bold=True, size=10)
    uat_font = Font(bold=True, color="2F5496", size=10)
    cfg_font = Font(bold=True, color="548235", size=10)
    le_sep_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    le_sep_font = Font(bold=True, size=11, color="1F3864")

    def write_header(ws, headers, row=1, fills=None):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hf
            c.fill = fills[ci - 1] if fills else hfill
            c.border = tb
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws, ncols, nrows):
        for ci in range(1, ncols + 1):
            mx = 8
            for r in range(1, min(nrows + 1, 200)):
                v = ws.cell(row=r, column=ci).value
                if v:
                    mx = max(mx, min(len(str(v)), 55))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 58)

    # ── Summary sheet ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_rows = [
        ["ADO ID", ado_id],
        ["Area Path", area_path],
        ["Form Name", uat_result.get("form_name", "")],
        ["Form Caption", uat_result.get("form_caption", "")],
        ["Menu Item", uat_result.get("menu_item", "")],
        ["MI Type", uat_result.get("mi_type", "")],
        ["Navigation Path", uat_result.get("navigation_path", "")],
        ["UAT Environment", "Env1 UAT (ENV1)"],
        ["Config Environment", "Env4 Config (ENV4)"],
        ["Extracted At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
        ["Legal Entity", "UAT Status", "Config Status",
         "UAT Grid Rows", "Config Grid Rows"],
    ]
    for le in legal_entities:
        uat_le = uat_result["legal_entities"].get(le, {})
        cfg_le = config_result["legal_entities"].get(le, {})
        uat_gr = sum(g.get("total_rows", 0) for g in uat_le.get("grids", {}).values())
        cfg_gr = sum(g.get("total_rows", 0) for g in cfg_le.get("grids", {}).values())
        sum_rows.append([le, uat_le.get("status", "N/A"), cfg_le.get("status", "N/A"),
                         uat_gr, cfg_gr])

    header_row_idx = 12
    for ri, rd in enumerate(sum_rows, 1):
        for ci, v in enumerate(rd, 1):
            c = ws_sum.cell(row=ri, column=ci, value=v)
            c.border = tb
            if ri in (1, 2):
                c.font = bold
            if ri == header_row_idx:
                c.font, c.fill = hf, hfill
    for ci in range(1, 6):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 28

    # ── Grid sheets: one row per record, grid cols + detail fields ──
    grids_info = collect_grid_data(env_results, legal_entities)

    for grid_name in sorted(grids_info.keys()):
        ginfo = grids_info[grid_name]
        col_labels = ginfo["col_labels"]
        detail_labels = ginfo["detail_labels"]
        if not col_labels:
            continue

        sheet_name = grid_name[:28] if len(grid_name) > 28 else grid_name
        ws = wb.create_sheet(sheet_name)

        # Headers: #, [grid cols], [D] detail cols, Env, LE
        headers = (["#"] + col_labels
                   + [f"[D] {dl}" for dl in detail_labels]
                   + ["Env", "LE"])
        ncols = len(headers)
        fills = ([hfill]
                 + [hfill] * len(col_labels)
                 + [detail_hfill] * len(detail_labels)
                 + [hfill, hfill])
        write_header(ws, headers, fills=fills)

        ri = 2
        seq = 0

        for le in legal_entities:
            # Collect rows from both envs for this LE
            le_rows = []  # (env_name, row_data, detail_data)
            for env_name, env_result in env_results:
                le_data = env_result["legal_entities"].get(le, {})
                gdata = le_data.get("grids", {}).get(grid_name, {})
                rows = gdata.get("rows", [])
                details = gdata.get("row_details", {})
                for row_idx, row in enumerate(rows):
                    det = {}
                    for rk in [row_idx, str(row_idx)]:
                        if rk in details:
                            det = details[rk]
                            break
                    le_rows.append((env_name, row, det))

            if not le_rows:
                continue

            # LE separator
            for ci in range(1, ncols + 1):
                c = ws.cell(row=ri, column=ci)
                c.fill = le_sep_fill
                c.border = tb
            ws.cell(row=ri, column=1, value=le).font = le_sep_font
            ws.cell(row=ri, column=1).fill = le_sep_fill
            # Count per env
            uat_cnt = sum(1 for e, _, _ in le_rows if e == "UAT")
            cfg_cnt = sum(1 for e, _, _ in le_rows if e == "Config")
            ws.cell(row=ri, column=2,
                    value=f"UAT: {uat_cnt} rows | Config: {cfg_cnt} rows"
                    ).font = le_sep_font
            ws.cell(row=ri, column=2).fill = le_sep_fill
            ri += 1

            # Write each row
            for env_name, row_data, det_data in le_rows:
                seq += 1
                ws.cell(row=ri, column=1, value=seq).border = tb
                ci = 2
                # Grid columns
                for cl in col_labels:
                    c = ws.cell(row=ri, column=ci, value=row_data.get(cl, ""))
                    c.border = tb
                    ci += 1
                # Detail fields — match by prefixed label (Tab > Field)
                for dl in detail_labels:
                    val = ""
                    for fname, fdata in det_data.items():
                        # Build the same prefixed label used during collection
                        raw_lbl = fdata.get("label", fname)
                        tab_lbl = fdata.get("_tab", "")
                        prefixed = f"{tab_lbl} > {raw_lbl}" if tab_lbl else raw_lbl
                        if prefixed == dl:
                            val = str(fdata.get("value", ""))
                            break
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = tb
                    ci += 1
                # Env
                c_env = ws.cell(row=ri, column=ci, value=env_name)
                c_env.border = tb
                c_env.font = uat_font if env_name == "UAT" else cfg_font
                # LE
                ws.cell(row=ri, column=ci + 1, value=le).border = tb
                ri += 1

        auto_width(ws, ncols, ri)
        ws.freeze_panes = "B2"
        if ri > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ri - 1}"

    # ── Fields sheet: non-grid controls from find_controls + fields ──
    ws_f = wb.create_sheet("Fields")
    f_headers = ["#", "Control Name", "Label", "Value", "Type", "Env", "LE"]
    write_header(ws_f, f_headers)
    fri = 2
    fseq = 0

    for le in legal_entities:
        # Collect field controls from both envs
        le_field_rows = []
        for env_name, env_result in env_results:
            le_data = env_result["legal_entities"].get(le, {})
            seen = set()
            # find_controls sweep results (primary source for parameters forms)
            for cname, cdata in le_data.get("find_controls", {}).items():
                if cdata.get("type") == "grid":
                    continue
                seen.add(cname)
                le_field_rows.append((env_name, cname,
                                      cdata.get("label", cname),
                                      str(cdata.get("value", "")),
                                      cdata.get("type", "")))
            # fields from FormState parsing (if not already covered)
            for cname, cdata in le_data.get("fields", {}).items():
                if cname in seen:
                    continue
                le_field_rows.append((env_name, cname,
                                      cdata.get("label", cname),
                                      str(cdata.get("value", "")),
                                      cdata.get("type", "")))

        if not le_field_rows:
            continue

        # LE separator
        for ci in range(1, len(f_headers) + 1):
            c = ws_f.cell(row=fri, column=ci)
            c.fill = le_sep_fill
            c.border = tb
        ws_f.cell(row=fri, column=1, value=le).font = le_sep_font
        ws_f.cell(row=fri, column=1).fill = le_sep_fill
        uat_cnt = sum(1 for e, *_ in le_field_rows if e == "UAT")
        cfg_cnt = sum(1 for e, *_ in le_field_rows if e == "Config")
        ws_f.cell(row=fri, column=2,
                  value=f"UAT: {uat_cnt} fields | Config: {cfg_cnt} fields"
                  ).font = le_sep_font
        ws_f.cell(row=fri, column=2).fill = le_sep_fill
        fri += 1

        for env_name, cname, label, value, ctype in le_field_rows:
            fseq += 1
            ws_f.cell(row=fri, column=1, value=fseq).border = tb
            ws_f.cell(row=fri, column=2, value=cname).border = tb
            ws_f.cell(row=fri, column=3, value=label).border = tb
            ws_f.cell(row=fri, column=4, value=value).border = tb
            ws_f.cell(row=fri, column=5, value=ctype).border = tb
            c_env = ws_f.cell(row=fri, column=6, value=env_name)
            c_env.border = tb
            c_env.font = uat_font if env_name == "UAT" else cfg_font
            ws_f.cell(row=fri, column=7, value=le).border = tb
            fri += 1

    auto_width(ws_f, len(f_headers), fri)
    ws_f.freeze_panes = "B2"
    if fri > 2:
        ws_f.auto_filter.ref = f"A1:{get_column_letter(len(f_headers))}{fri - 1}"

    wb.save(filepath)
    print(f"\nExcel saved: {filepath}")
    return filepath


def run_one(mi_name, mi_type, area_path, legal_entities, ado_id=""):
    """Extract one form from UAT + Config and write Excel."""
    print("=" * 70)
    print("D365 Form Control Extraction: UAT + Config")
    print("=" * 70)
    print(f"  Form:  {mi_name}")
    print(f"  ADO:   {ado_id}")
    print(f"  LEs:   {', '.join(legal_entities)}")
    print(f"  UAT:   ENV1 (Env1 UAT)")
    print(f"  Config: ENV4 (Env4 Config)")
    print()

    print(">>> Extracting from UAT (ENV1)...")
    uat_extractor = FormControlExtractor(env_key="ENV1")
    uat_result = uat_extractor.extract(mi_name, legal_entities, mi_type)

    print("\n>>> Extracting from Config (ENV4)...")
    config_extractor = FormControlExtractor(env_key="ENV4")
    config_result = config_extractor.extract(mi_name, legal_entities, mi_type)

    env_results = [("UAT", uat_result), ("Config", config_result)]

    print("\n>>> Writing Excel...")
    filepath = to_excel(env_results, legal_entities, uat_result, config_result,
                        ado_id=ado_id, area_path=area_path)
    print(f"\nDone! Open: {filepath}")
    return filepath


def main():
    legal_entities = ["MY30", "MY60", "SG60"]

    forms = [
        ("CBAFGeographyL5", "Display",
         r"Sales and marketing > Setup > Acme asia > Geography L5"),
    ]

    for mi_name, mi_type, area_path in forms:
        run_one(mi_name, mi_type, area_path, legal_entities)
        print("\n\n")


if __name__ == "__main__":
    main()
