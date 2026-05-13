"""
MY/SG Reporter — wraps the main reporter and adds ADO Work Item ID column.
Generates separate reports for OData and Form-based comparisons.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def generate_report(results, source_name, target_name, output_dir, report_tag=""):
    """
    Generate an Excel report with ADO Work Item ID column.

    Args:
        results: dict {display_key: comparison_result}
                 each result must have 'item_id', 'category', 'summary', etc.
        source_name: source environment display name
        target_name: str or list of target names
        output_dir: output directory
        report_tag: e.g. "OData" or "Form" — added to filename
    """
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    multi_target = isinstance(target_name, list) and len(target_name) > 1

    tag = f"_{report_tag}" if report_tag else ""
    if multi_target:
        tgt_display = "Multiple"
    else:
        tgt_display = target_name[0] if isinstance(target_name, list) else target_name

    filename = f"D365_Config{tag}_{source_name}_vs_{tgt_display}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb = Workbook()

    if multi_target:
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _write_multi_summary(ws_summary, results, source_name, target_name)

        sheet_count = 0
        for entity_name, pair_results in sorted(results.items()):
            for pair_label, result in pair_results.items():
                s = result.get("summary", {})
                has_diffs = s.get("only_in_source", 0) > 0 or s.get("only_in_target", 0) > 0 or s.get("differences", 0) > 0
                if has_diffs and sheet_count < 200:
                    pair_tag = pair_label.replace(" vs ", "vs").replace(" ", "")
                    ws = wb.create_sheet(title=_safe_sheet_name(f"{entity_name} {pair_tag}"))
                    parts = pair_label.split(" vs ")
                    src_n = parts[0] if len(parts) == 2 else source_name
                    tgt_n = parts[1] if len(parts) == 2 else pair_label
                    _write_diff_sheet(ws, entity_name, result, src_n, tgt_n)
                    sheet_count += 1
    else:
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _write_summary(ws_summary, results, source_name, tgt_display)

        sheet_count = 0
        for entity_name, result in sorted(results.items()):
            s = result.get("summary", {})
            has_diffs = s.get("only_in_source", 0) > 0 or s.get("only_in_target", 0) > 0 or s.get("differences", 0) > 0
            if has_diffs and sheet_count < 200:
                # Only create diff sheet if there's actual detail data
                detail_count = len(result.get("only_in_source", [])) + len(result.get("only_in_target", [])) + len(result.get("differences", []))
                if detail_count > 0:
                    ado_id = result.get("item_id", "")
                    sheet_label = f"CD {ado_id}" if ado_id else _safe_sheet_name(entity_name)
                    ws = wb.create_sheet(title=_safe_sheet_name(sheet_label))
                    _write_diff_sheet(ws, entity_name, result, source_name, tgt_display)
                    sheet_count += 1

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}")
    print(f"  Summary + {sheet_count} diff sheets")
    return filepath


def _write_summary(ws, results, source_name, target_name):
    """Summary sheet with ADO ID column."""
    ws.merge_cells("A1:J1")
    ws["A1"] = f"D365 Configuration Comparison: {source_name} vs {target_name}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    total_entities = len(results)
    entities_match = sum(1 for r in results.values()
                         if r.get("summary", {}).get("only_in_source", 0) == 0
                         and r.get("summary", {}).get("only_in_target", 0) == 0
                         and r.get("summary", {}).get("differences", 0) == 0)
    entities_differ = total_entities - entities_match

    ws["A4"] = "Total items compared:"
    ws["C4"] = total_entities
    ws["A5"] = "Items matching:"
    ws["C5"] = entities_match
    ws["C5"].fill = GREEN_FILL
    ws["A6"] = "Items with differences:"
    ws["C6"] = entities_differ
    if entities_differ > 0:
        ws["C6"].fill = RED_FILL
    for r in range(4, 7):
        ws[f"A{r}"].font = BOLD_FONT

    headers = [
        "ADO ID", "Entity / Form", "Category",
        f"{source_name} Count", f"{target_name} Count",
        f"Only in {source_name}", f"Only in {target_name}",
        "Value Differences", "Match %", "Status"
    ]
    row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row = 9
    for entity_name in sorted(results.keys()):
        result = results[entity_name]
        s = result.get("summary", {})

        src_count = s.get("source_count", 0)
        tgt_count = s.get("target_count", 0)
        only_src = s.get("only_in_source", 0)
        only_tgt = s.get("only_in_target", 0)
        diffs = s.get("differences", 0)
        matching = s.get("matching", 0)
        total = max(src_count, tgt_count, 1)
        match_pct = round(matching / total * 100, 1) if total > 0 else 100.0
        status = "MATCH" if (only_src == 0 and only_tgt == 0 and diffs == 0) else "DIFF"

        ado_id = result.get("item_id", "")

        values = [
            ado_id, entity_name, result.get("category", ""),
            src_count, tgt_count,
            only_src, only_tgt, diffs,
            f"{match_pct}%", status,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 10:  # Status
                cell.fill = GREEN_FILL if status == "MATCH" else RED_FILL
                cell.font = BOLD_FONT
            elif col in (6, 7, 8) and isinstance(val, (int, float)) and val > 0:
                cell.fill = YELLOW_FILL

        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:{get_column_letter(len(headers))}{row - 1}"


def _write_multi_summary(ws, results, source_name, target_names):
    """Multi-target summary with ADO ID column."""
    tgt_display = ", ".join(target_names)
    ws.merge_cells("A1:L1")
    ws["A1"] = f"D365 Configuration Comparison: {source_name} vs {tgt_display}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    all_pairs = []
    for pair_results in results.values():
        for pk in pair_results:
            if pk not in all_pairs:
                all_pairs.append(pk)

    total_entities = len(results)
    entities_with_diffs = 0
    for entity_pairs in results.values():
        for pair_result in entity_pairs.values():
            s = pair_result.get("summary", {})
            if s.get("only_in_source", 0) > 0 or s.get("only_in_target", 0) > 0 or s.get("differences", 0) > 0:
                entities_with_diffs += 1
                break

    ws["A4"] = "Total items compared:"
    ws["C4"] = total_entities
    ws["A5"] = "Items with differences (any pair):"
    ws["C5"] = entities_with_diffs
    ws["C5"].fill = RED_FILL if entities_with_diffs > 0 else GREEN_FILL
    ws["A6"] = "Comparison pairs:"
    ws["C6"] = ", ".join(all_pairs)
    for r in range(4, 7):
        ws[f"A{r}"].font = BOLD_FONT

    headers = ["ADO ID", "Entity / Form", "Category"]
    for pair in all_pairs:
        short = pair.replace(" vs ", "/")
        headers.extend([f"{short} Only Src", f"{short} Only Tgt", f"{short} Diffs", f"{short} Match%"])
    headers.append("Status")

    row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row = 9
    for entity_name in sorted(results.keys()):
        pair_results = results[entity_name]
        category = ""
        ado_id = ""
        any_diff = False
        col = 1

        for pr in pair_results.values():
            if not category:
                category = pr.get("category", "")
            if not ado_id:
                ado_id = pr.get("item_id", "")

        ws.cell(row=row, column=col, value=ado_id).border = THIN_BORDER
        col += 1
        ws.cell(row=row, column=col, value=entity_name).border = THIN_BORDER
        col += 1
        ws.cell(row=row, column=col, value=category).border = THIN_BORDER
        col += 1

        for pair in all_pairs:
            pr = pair_results.get(pair, {})
            s = pr.get("summary", {})
            only_src = s.get("only_in_source", 0)
            only_tgt = s.get("only_in_target", 0)
            diffs = s.get("differences", 0)
            matching = s.get("matching", 0)
            total = max(s.get("source_count", 0), s.get("target_count", 0), 1)
            match_pct = round(matching / total * 100, 1) if total > 0 else 100.0

            if only_src > 0 or only_tgt > 0 or diffs > 0:
                any_diff = True

            for val in [only_src, only_tgt, diffs]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                if isinstance(val, (int, float)) and val > 0:
                    cell.fill = YELLOW_FILL
                col += 1
            cell = ws.cell(row=row, column=col, value=f"{match_pct}%")
            cell.border = THIN_BORDER
            col += 1

        status = "DIFF" if any_diff else "MATCH"
        cell = ws.cell(row=row, column=col, value=status)
        cell.border = THIN_BORDER
        cell.fill = RED_FILL if any_diff else GREEN_FILL
        cell.font = BOLD_FONT
        row += 1

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:{get_column_letter(len(headers))}{row - 1}"


def _write_diff_sheet(ws, entity_name, result, source_name, target_name):
    """Per-entity diff sheet."""
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Entity: {entity_name}"
    ws["A1"].font = TITLE_FONT

    ado_id = result.get("item_id", "")
    if ado_id:
        ws["A2"] = f"ADO Work Item: {ado_id}"
        ws["A2"].font = BOLD_FONT

    row = 4

    only_source = result.get("only_in_source", [])
    if only_source:
        ws.cell(row=row, column=1, value=f"Records only in {source_name} ({len(only_source)})").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = RED_FILL
        row += 1
        row = _write_records_table(ws, only_source, row)
        row += 1

    only_target = result.get("only_in_target", [])
    if only_target:
        ws.cell(row=row, column=1, value=f"Records only in {target_name} ({len(only_target)})").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = BLUE_FILL
        row += 1
        row = _write_records_table(ws, only_target, row)
        row += 1

    differences = result.get("differences", [])
    if differences:
        ws.cell(row=row, column=1, value=f"Records with differences ({len(differences)})").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = YELLOW_FILL
        row += 1

        diff_headers = ["Key", "Field", f"{source_name} Value", f"{target_name} Value"]
        for col, h in enumerate(diff_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        row += 1

        for diff in differences[:500]:
            key_str = " | ".join(str(k) for k in diff.get("key", ()))
            for field, vals in diff.get("changed_fields", {}).items():
                ws.cell(row=row, column=1, value=key_str).border = THIN_BORDER
                ws.cell(row=row, column=2, value=field).border = THIN_BORDER
                src_cell = ws.cell(row=row, column=3, value=str(vals.get("source", "")))
                src_cell.border = THIN_BORDER
                src_cell.fill = YELLOW_FILL
                tgt_cell = ws.cell(row=row, column=4, value=str(vals.get("target", "")))
                tgt_cell.border = THIN_BORDER
                tgt_cell.fill = YELLOW_FILL
                row += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 25


def _write_records_table(ws, records, start_row):
    if not records:
        return start_row
    fields = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen and not k.startswith("@odata"):
                fields.append(k)
                seen.add(k)
    fields = fields[:20]

    for col, f in enumerate(fields, 1):
        cell = ws.cell(row=start_row, column=col, value=f)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = start_row + 1
    for record in records[:200]:
        for col, f in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col, value=str(record.get(f, "")))
            cell.border = THIN_BORDER
        row += 1
    return row


def _safe_sheet_name(name):
    safe = name.replace("/", "_").replace("\\", "_").replace("*", "")
    safe = safe.replace("?", "").replace("[", "").replace("]", "").replace(":", "")
    if len(safe) > 31:
        safe = safe[:28] + "..."
    return safe
